# -*- coding: utf-8 -*-
"""
SOA (Statement of Account) 生成服务
用于生成客户对账单 - 基于 Project 数据
"""

import json
from datetime import datetime
from App_new.exts import db
from App_new.business.projects.models.project import ProjectHeader, CustomerCompany
from App_new.business.projects.models.ref import ProjectRef
from App_new.business.projects.services.ref_extra_info import resolve_ref_start_date
from App_new.business.projects.models.receipt import ProjectReceipt, ReceiptInvoiceAllocation
from App_new.business.projects.models.invoice import ProjectInvoice
from App_new.business.flight.models.flight import ProjectFlightSegment
from sqlalchemy.orm import contains_eager
from flask import render_template
import io
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from sqlalchemy import union_all


class SOAService:
    """SOA生成服务 - 基于 Project 数据"""

    def __init__(self):
        pass

    def _resolve_date_range(self, month=None, start_date=None, end_date=None):
        """根据 month / start_date / end_date 解析出 BK.DATE 的过滤区间 [start, end)。

        优先级：start_date / end_date 任一存在时按区间过滤（忽略 month）；
        否则回退到 month（按整月过滤）。返回 (start, end_exclusive)；
        任一边界为 None 表示该侧不限。解析失败的字段视为未提供。
        """
        from datetime import date, timedelta

        def _parse(s):
            if not s:
                return None
            try:
                return datetime.strptime(s, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                return None

        start = _parse(start_date)
        end_inclusive = _parse(end_date)

        if start is not None or end_inclusive is not None:
            # 用户显式指定区间：end_date 是包含日，转为半开区间右端 +1 天
            end_exclusive = (end_inclusive + timedelta(days=1)) if end_inclusive else None
            return start, end_exclusive

        if month:
            try:
                year, month_num = month.split('-')
                year = int(year)
                month_num = int(month_num)
                month_start = date(year, month_num, 1)
                month_end = date(year + 1, 1, 1) if month_num == 12 else date(year, month_num + 1, 1)
                return month_start, month_end
            except (ValueError, TypeError):
                pass

        return None, None

    def _calculate_project_balance(self, project_id, total_selling):
        """计算项目余额（销售金额 - 已收款金额）"""
        try:
            # 方式1: 通过发票分配表统计（项目级别收款）
            alloc_total = db.session.query(
                db.func.coalesce(db.func.sum(ReceiptInvoiceAllocation.allocated_amount), 0)
            ).join(
                ProjectInvoice, ReceiptInvoiceAllocation.invoice_id == ProjectInvoice.id
            ).join(
                ProjectReceipt, ReceiptInvoiceAllocation.receipt_id == ProjectReceipt.id
            ).filter(
                ProjectInvoice.header_id == project_id,
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id == None
            ).scalar() or 0

            # 方式2: REF级别直接收款
            ref_receipt_total = db.session.query(
                db.func.coalesce(db.func.sum(ProjectReceipt.amount), 0)
            ).filter(
                ProjectReceipt.header_id == project_id,
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id.isnot(None)
            ).scalar() or 0

            # 方式3: 旧项目级别收款（没有分配记录也没有ref_id）
            old_receipts_subq = db.session.query(
                ReceiptInvoiceAllocation.receipt_id
            ).distinct().scalar_subquery()

            old_receipt_total = db.session.query(
                db.func.coalesce(db.func.sum(ProjectReceipt.amount), 0)
            ).filter(
                ProjectReceipt.header_id == project_id,
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id == None,
                ~ProjectReceipt.id.in_(old_receipts_subq)
            ).scalar() or 0

            total_received = float(alloc_total) + float(ref_receipt_total) + float(old_receipt_total)
            balance = float(total_selling) - total_received
            return balance
        except Exception as e:
            print(f"计算项目 {project_id} 余额时出错: {e}")
            return float(total_selling)

    def _get_project_departure_date(self, ref_id):
        """从REF的航段中获取出发日期"""
        try:
            # 获取第一个航段的出发时间
            first_segment = ProjectFlightSegment.query.filter_by(
                ref_id=ref_id
            ).order_by(ProjectFlightSegment.departure_time).first()

            if first_segment and first_segment.departure_time:
                return first_segment.departure_time.strftime('%Y-%m-%d')
        except Exception:
            pass
        return ''

    def _query_invoice_rows(self, search=None, month=None, company=None, group=None,
                             balance_positive=False, profit_loss=None,
                             start_date=None, end_date=None):
        """查询符合条件的发票行（每张 confirmed 发票一行）。

        - 日期过滤作用于 invoice_date
        - 没有 confirmed 发票的项目自然不会出现
        - 结果按公司名稳定排序便于下载时分组求小计；公司内部按发票日期倒序
        """
        # 主查询：confirmed 发票 + 关联项目
        q = ProjectInvoice.query.join(
            ProjectHeader, ProjectInvoice.header_id == ProjectHeader.id
        ).options(
            contains_eager(ProjectInvoice.header).joinedload(ProjectHeader.company)
        ).filter(ProjectInvoice.status == 'confirmed')

        if search:
            s = search.lower()
            q = q.filter(db.or_(
                ProjectInvoice.invoice_number.ilike(f'%{s}%'),
                ProjectHeader.hid.ilike(f'%{s}%'),
                ProjectHeader.desc.ilike(f'%{s}%'),
                ProjectHeader.contact.ilike(f'%{s}%'),
                ProjectHeader.leader_name.ilike(f'%{s}%')
            ))

        if company:
            q = q.join(CustomerCompany).filter(CustomerCompany.company_name == company)
        elif group:
            q = q.join(CustomerCompany).filter(CustomerCompany.group_name == group)

        # 日期过滤：作用于发票日期（invoice_date）
        range_start, range_end = self._resolve_date_range(month=month, start_date=start_date, end_date=end_date)
        if range_start is not None:
            q = q.filter(ProjectInvoice.invoice_date >= range_start)
        if range_end is not None:
            q = q.filter(ProjectInvoice.invoice_date < range_end)

        # 排序：按集团筛选时先按公司名分组，否则按发票日期倒序
        if group:
            q = q.order_by(
                CustomerCompany.company_name,
                ProjectInvoice.invoice_date.desc(),
                ProjectInvoice.id.desc()
            )
        else:
            q = q.order_by(
                ProjectInvoice.invoice_date.desc(),
                ProjectInvoice.id.desc()
            )

        invoices = q.all()
        if not invoices:
            return []

        # 解析每张发票直接关联的 REF（ProjectInvoice.ref_ids 是 JSON 数组），
        # 这样描述/出发日期能反映该发票本身的内容，而不是项目的第一条 REF
        invoice_ref_ids = {}  # invoice.id -> 保留 JSON 中顺序的 ref id 列表
        all_invoice_ref_ids = set()
        for inv in invoices:
            ids = []
            if inv.ref_ids:
                try:
                    parsed = json.loads(inv.ref_ids)
                    if isinstance(parsed, list):
                        for x in parsed:
                            try:
                                ids.append(int(x))
                            except (TypeError, ValueError):
                                continue
                except (json.JSONDecodeError, TypeError):
                    ids = []
            invoice_ref_ids[inv.id] = ids
            all_invoice_ref_ids.update(ids)

        # 项目第一条 REF 的 id（仅用于发票未挂任何 REF 时的回退展示）
        header_ids = list({inv.header_id for inv in invoices if inv.header_id})
        first_ref_id_by_header = {}
        if header_ids:
            from sqlalchemy import func
            for hid, rid in db.session.query(
                ProjectRef.header_id,
                func.min(ProjectRef.id)
            ).filter(ProjectRef.header_id.in_(header_ids)).group_by(ProjectRef.header_id).all():
                first_ref_id_by_header[hid] = rid

        # 一次性加载所有要用到的 REF（发票挂的 + 回退用的）
        needed_ref_ids = set(all_invoice_ref_ids)
        needed_ref_ids.update(first_ref_id_by_header.values())
        refs_by_id = {}
        if needed_ref_ids:
            for r in ProjectRef.query.filter(ProjectRef.id.in_(list(needed_ref_ids))).all():
                refs_by_id[r.id] = r

        # 一次性加载这些 REF 对应的最早出发时间（机票/航段）
        dep_date_by_ref = {}
        if needed_ref_ids:
            for row in db.session.query(
                ProjectFlightSegment.ref_id,
                db.func.min(ProjectFlightSegment.departure_time).label('first_dep')
            ).filter(ProjectFlightSegment.ref_id.in_(list(needed_ref_ids))).group_by(ProjectFlightSegment.ref_id).all():
                if row.first_dep:
                    dep_date_by_ref[row.ref_id] = row.first_dep.strftime('%Y-%m-%d')

        # 没有航段日期的 REF（酒店/保险/景点等）从 extra_info 里回退，
        # 各业务类型的日期键名不同，统一走 resolve_ref_start_date 解析
        for rid, ref in refs_by_id.items():
            if rid in dep_date_by_ref or not ref.extra_info:
                continue
            try:
                info = json.loads(ref.extra_info)
            except (json.JSONDecodeError, TypeError):
                continue
            d = resolve_ref_start_date(info)
            if d:
                dep_date_by_ref[rid] = d

        # 项目盈亏（仅当用到 profit_loss 过滤时才查）
        project_pl = {}
        if profit_loss and header_ids:
            for r in db.session.query(
                ProjectRef.header_id,
                db.func.sum(ProjectRef.selling_price).label('total_selling'),
                db.func.sum(ProjectRef.cost_price).label('total_cost')
            ).filter(ProjectRef.header_id.in_(header_ids)).group_by(ProjectRef.header_id).all():
                project_pl[r.header_id] = float(r.total_selling or 0) - float(r.total_cost or 0)

        rows = []
        for inv in invoices:
            header = inv.header
            if not header:
                continue

            amount = float(inv.amount or 0)
            paid = float(inv.paid_amount or 0)
            unpaid = amount - paid

            if balance_positive and unpaid <= 0:
                continue

            if profit_loss:
                project_profit = project_pl.get(header.id, 0.0)
                if profit_loss == 'profit' and project_profit <= 0:
                    continue
                elif profit_loss == 'loss' and project_profit >= 0:
                    continue
                elif profit_loss == 'even' and project_profit != 0:
                    continue

            # 取本发票挂的 REF 列表（按 ref_ids 中的顺序）
            inv_refs = [refs_by_id[rid] for rid in invoice_ref_ids.get(inv.id, []) if rid in refs_by_id]

            # ITIN 描述：用本发票的 REF 描述拼接（多个用 / 分隔）；无则回退到项目第一条 REF / header.desc
            descs = []
            for ref in inv_refs:
                d = (ref.description or getattr(ref, 'detailed_description', None) or '').strip()
                if d:
                    descs.append(d)

            if descs:
                itin_desc = ' / '.join(descs)
            else:
                first_ref_id = first_ref_id_by_header.get(header.id)
                first_ref = refs_by_id.get(first_ref_id) if first_ref_id else None
                itin_desc = (first_ref.description if first_ref and first_ref.description else header.desc) or ''

            # 出发日期：先取本发票第一条挂的 REF 的最早出发时间；没有再回退项目第一条 REF
            dep_date = ''
            for ref in inv_refs:
                d = dep_date_by_ref.get(ref.id, '')
                if d:
                    dep_date = d
                    break
            if not dep_date:
                first_ref_id = first_ref_id_by_header.get(header.id)
                if first_ref_id:
                    dep_date = dep_date_by_ref.get(first_ref_id, '')

            rows.append({
                'invoice_id': inv.id,
                'invoice_number': inv.invoice_number,
                'invoice_date': inv.invoice_date.strftime('%Y-%m-%d') if inv.invoice_date else '',
                'header_id': header.id,
                'hid': header.hid,
                'corporate_name': header.company.company_name if header.company else None,
                'client_name': header.leader_member_name or header.leader_name or header.contact,
                'itin_desc': itin_desc,
                'dep_date': dep_date,
                'currency': inv.currency or 'SGD',
                'amount': amount,
                'paid_amount': paid,
                'unpaid_amount': unpaid,
            })

        return rows

    # 可排序的列 -> 行字典里的键
    SORTABLE_FIELDS = {
        'invoice_number': 'invoice_number',  # INV NO
        'hid': 'hid',                        # HID
    }

    @staticmethod
    def _alnum_sort_key(value):
        """字母前缀 + 数字的编号排序键

        INV NO 是纯数字字符串（'12017'），HID 形如 'H2181'。直接按字符串排会得到
        H999 > H2181 这种结果，所以拆成 (前缀, 数字, 原串) 让数字部分按数值比较。
        """
        import re
        s = str(value or '')
        m = re.search(r'\d+', s)
        if m:
            return (s[:m.start()].upper(), int(m.group()), s.upper())
        return (s.upper(), 0, s.upper())

    @classmethod
    def _apply_sort(cls, rows, sort_by, sort_dir='asc'):
        """按 INV NO / HID 就地排序（列表页、Excel、PDF 共用同一套口径）

        Excel/PDF 之后还会按公司名做**稳定**排序来分组加小计，
        所以这里排完，公司组内部就是用户选的顺序，分组结构不受影响。
        """
        field = cls.SORTABLE_FIELDS.get(sort_by)
        if not field:
            return rows
        rows.sort(
            key=lambda r: cls._alnum_sort_key(r.get(field)),
            reverse=(str(sort_dir).lower() == 'desc')
        )
        return rows

    def get_soa_list(self, page=1, per_page=20, search=None, month=None, company=None, group=None, balance_positive=False, profit_loss=None, start_date=None, end_date=None, sort_by=None, sort_dir='asc'):
        """获取可生成SOA的发票列表（每张 confirmed 发票一行）

        sort_by: 'invoice_number' / 'hid'，为空则保持原有默认排序
        sort_dir: 'asc' 正序 / 'desc' 倒序
        """
        try:
            rows = self._query_invoice_rows(
                search=search, month=month, company=company, group=group,
                balance_positive=balance_positive, profit_loss=profit_loss,
                start_date=start_date, end_date=end_date
            )

            # 分页在内存里做，排序也放在切片之前，保证是"全量排序后再翻页"
            self._apply_sort(rows, sort_by, sort_dir)

            total_amount = sum(r['amount'] for r in rows)
            total_balance = sum(r['unpaid_amount'] for r in rows)
            total_count = len(rows)

            total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1
            start_idx = (page - 1) * per_page
            page_rows = rows[start_idx:start_idx + per_page]

            return {
                'invoices': page_rows,
                'total': total_count,
                'pages': total_pages,
                'current_page': page,
                'per_page': per_page,
                'statistics': {
                    'total_balance': total_balance,
                    'total_amount': total_amount,
                    'total_count': total_count
                }
            }

        except Exception as e:
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'message': f'获取SOA列表失败: {str(e)}'
            }

    def get_company_list(self, group=None):
        """获取所有公司列表，支持按集团筛选"""
        try:
            # 获取所有有项目的不重复公司名称
            query = db.session.query(CustomerCompany.company_name).join(
                ProjectHeader, ProjectHeader.company_id == CustomerCompany.id
            ).filter(
                CustomerCompany.company_name.isnot(None),
                CustomerCompany.company_name != ''
            )

            # 如果指定了集团，则只返回该集团下的公司
            if group:
                query = query.filter(CustomerCompany.group_name == group)

            companies = query.distinct().all()
            company_list = [company[0] for company in companies if company[0]]

            return {
                'success': True,
                'companies': company_list
            }

        except Exception as e:
            return {
                'success': False,
                'message': f'获取公司列表失败: {str(e)}'
            }

    def get_group_list(self):
        """获取所有集团/关联标签列表"""
        try:
            # 获取所有有项目的不重复集团名称
            groups = db.session.query(CustomerCompany.group_name).join(
                ProjectHeader, ProjectHeader.company_id == CustomerCompany.id
            ).filter(
                CustomerCompany.group_name.isnot(None),
                CustomerCompany.group_name != ''
            ).distinct().all()

            group_list = [group[0] for group in groups if group[0]]

            return {
                'success': True,
                'groups': group_list
            }

        except Exception as e:
            return {
                'success': False,
                'message': f'获取集团列表失败: {str(e)}'
            }

    def batch_download_soa(self, group=None, company=None, month=None, search=None, balance_positive=None, profit_loss=None, format='excel', start_date=None, end_date=None, sort_by=None, sort_dir='asc'):
        """批量下载SOA - 生成Excel表格（每张 confirmed 发票一行）"""
        try:
            import pandas as pd

            invoice_rows = self._query_invoice_rows(
                search=search, month=month, company=company, group=group,
                balance_positive=bool(balance_positive), profit_loss=profit_loss,
                start_date=start_date, end_date=end_date
            )

            if not invoice_rows:
                return None, "没有找到符合条件的SOA数据"

            # 与列表页同一套排序（下面按公司分组是稳定排序，组内即此顺序）
            self._apply_sort(invoice_rows, sort_by, sort_dir)

            # 准备 Excel 行：列顺序固定为 INV NO / HID / COMPANY / INV.DATE / CLIENT / DP.DATE / ITIN / CCY / AMOUNT / BAL
            def _data_row(r):
                return {
                    'INV NO': r['invoice_number'] or '',
                    'HID': r['hid'] or '',
                    'COMPANY': r['corporate_name'] or '',
                    'INV.DATE': r['invoice_date'] or '',
                    'CLIENT NAME': r['client_name'] or '',
                    'DP.DATE': r['dep_date'] or '',
                    'ITIN DESCRIPTION': r['itin_desc'] or '',
                    'CCY': r['currency'] or 'SGD',
                    'AMOUNT': r['amount'],
                    'BAL': r['unpaid_amount'],
                }

            excel_data = [_data_row(r) for r in invoice_rows]

            # 按公司分组：稳定排序仅按公司名，公司内部保留 invoice_date 倒序
            excel_data.sort(key=lambda row: (row['COMPANY'] or ''))

            final_rows = []
            subtotal_row_indices = []
            companies_seen = []
            current_company = None
            company_amount_subtotal = 0.0
            company_bal_subtotal = 0.0
            company_count = 0
            grand_amount = 0.0
            grand_bal = 0.0

            def _append_subtotal(company, amount_sub, bal_sub, count):
                final_rows.append({
                    'INV NO': '',
                    'HID': '',
                    'COMPANY': company or '',
                    'INV.DATE': '',
                    'CLIENT NAME': '',
                    'DP.DATE': '',
                    'ITIN DESCRIPTION': f'SUBTOTAL ({count} item{"s" if count > 1 else ""})',
                    'CCY': 'SGD',
                    'AMOUNT': amount_sub,
                    'BAL': bal_sub,
                })
                subtotal_row_indices.append(len(final_rows) - 1)

            for row in excel_data:
                if current_company is None:
                    current_company = row['COMPANY']
                elif row['COMPANY'] != current_company:
                    _append_subtotal(current_company, company_amount_subtotal, company_bal_subtotal, company_count)
                    companies_seen.append(current_company)
                    current_company = row['COMPANY']
                    company_amount_subtotal = 0.0
                    company_bal_subtotal = 0.0
                    company_count = 0

                final_rows.append(row)
                company_amount_subtotal += row['AMOUNT']
                company_bal_subtotal += row['BAL']
                company_count += 1
                grand_amount += row['AMOUNT']
                grand_bal += row['BAL']

            # 最后一家公司的小计
            _append_subtotal(current_company, company_amount_subtotal, company_bal_subtotal, company_count)
            companies_seen.append(current_company)

            # 仅当有多家公司时才显示总计行
            grand_total_row_index = None
            if len(companies_seen) > 1:
                final_rows.append({
                    'INV NO': '',
                    'HID': '',
                    'COMPANY': '',
                    'INV.DATE': '',
                    'CLIENT NAME': '',
                    'DP.DATE': '',
                    'ITIN DESCRIPTION': 'GRAND TOTAL',
                    'CCY': 'SGD',
                    'AMOUNT': grand_amount,
                    'BAL': grand_bal,
                })
                grand_total_row_index = len(final_rows) - 1

            df = pd.DataFrame(final_rows)

            # 生成Excel文件
            excel_buffer = io.BytesIO()

            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='SOA Data', index=False)

                worksheet = writer.sheets['SOA Data']

                # 设置列宽（A..J 对应 INV NO, HID, COMPANY, INV.DATE, CLIENT, DP.DATE, ITIN, CCY, AMOUNT, BAL）
                column_widths = {
                    'A': 12,  # INV NO
                    'B': 10,  # HID
                    'C': 25,  # COMPANY
                    'D': 12,  # INV.DATE
                    'E': 25,  # CLIENT NAME
                    'F': 12,  # DP.DATE
                    'G': 30,  # ITIN DESCRIPTION
                    'H': 8,   # CCY
                    'I': 12,  # AMOUNT
                    'J': 12,  # BAL
                }
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width

                # 设置表头样式
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # AMOUNT(第 9 列, idx 8) 与 BAL(第 10 列, idx 9) 设置货币格式
                for row in worksheet.iter_rows(min_row=2):
                    if len(row) > 9:
                        row[8].number_format = '$#,##0.00'
                        row[9].number_format = '$#,##0.00'

                # 小计行样式（浅黄底、加粗）
                subtotal_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                subtotal_font = Font(bold=True, color="664D03")
                for idx in subtotal_row_indices:
                    excel_row_num = idx + 2
                    row_cells = worksheet[excel_row_num]
                    for cell in row_cells:
                        cell.fill = subtotal_fill
                        cell.font = subtotal_font
                    if len(row_cells) > 9:
                        row_cells[8].number_format = '$#,##0.00'
                        row_cells[9].number_format = '$#,##0.00'

                # 总计行样式（仅当有多家公司时存在）
                if grand_total_row_index is not None:
                    excel_row_num = grand_total_row_index + 2
                    total_row = worksheet[excel_row_num]
                    total_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    total_font = Font(bold=True, color="000080")
                    for cell in total_row:
                        cell.fill = total_fill
                        cell.font = total_font
                    if len(total_row) > 9:
                        total_row[8].number_format = '$#,##0.00'
                        total_row[9].number_format = '$#,##0.00'

                # 添加公司付款账户信息
                bank_info_start_row = worksheet.max_row + 3

                bank_info_font = Font(bold=False, color="000000")
                bank_title_font = Font(bold=True, color="000000")

                worksheet.cell(row=bank_info_start_row, column=1, value="JOYFUL ESCAPES PTE. LTD Bank Details:")
                worksheet.cell(row=bank_info_start_row, column=1).font = bank_title_font

                worksheet.cell(row=bank_info_start_row + 1, column=1, value="• OCBC Bank Singapore, AC No.: 5956-7793-1001")
                worksheet.cell(row=bank_info_start_row + 1, column=1).font = bank_info_font

                worksheet.cell(row=bank_info_start_row + 2, column=1, value="• PayNow UEN: 202337627W")
                worksheet.cell(row=bank_info_start_row + 2, column=1).font = bank_info_font

            excel_buffer.seek(0)
            return excel_buffer.getvalue(), None

        except Exception as e:
            import traceback
            traceback.print_exc()
            return None, f"生成Excel文件时出错: {str(e)}"

    def batch_download_soa_pdf(self, group=None, company=None, month=None, search=None, balance_positive=None, profit_loss=None, start_date=None, end_date=None, sort_by=None, sort_dir='asc'):
        """批量下载SOA - 生成PDF文件（每张 confirmed 发票一行）"""
        try:
            invoice_rows = self._query_invoice_rows(
                search=search, month=month, company=company, group=group,
                balance_positive=bool(balance_positive), profit_loss=profit_loss,
                start_date=start_date, end_date=end_date
            )

            if not invoice_rows:
                return None, "没有找到符合条件的SOA数据"

            # 与列表页同一套排序（下面按公司分组是稳定排序，组内即此顺序）
            self._apply_sort(invoice_rows, sort_by, sort_dir)

            # 创建批量PDF换行样式
            batch_wrap_style = ParagraphStyle(
                'BatchWrapStyle',
                parent=getSampleStyleSheet()['Normal'],
                fontSize=8,
                leading=9,
                leftIndent=0,
                rightIndent=0,
                spaceBefore=0,
                spaceAfter=0
            )

            # 把每条发票行转成 PDF 表格行（列顺序：INV NO/HID/COMPANY/INV.DATE/CLIENT/DP.DATE/ITIN/CCY/AMOUNT/BAL）
            collected = []  # list of dict: {company, amount, balance, row}
            for r in invoice_rows:
                company_name = r['corporate_name'] or ''
                row_data = [
                    Paragraph(r['invoice_number'] or '', batch_wrap_style),
                    Paragraph(r['hid'] or '', batch_wrap_style),
                    Paragraph(company_name, batch_wrap_style),
                    Paragraph(r['invoice_date'] or '', batch_wrap_style),
                    Paragraph(r['client_name'] or '', batch_wrap_style),
                    Paragraph(r['dep_date'] or '', batch_wrap_style),
                    Paragraph(r['itin_desc'] or '', batch_wrap_style),
                    Paragraph(r['currency'] or 'SGD', batch_wrap_style),
                    Paragraph(f"${r['amount']:,.2f}", batch_wrap_style),
                    Paragraph(f"${r['unpaid_amount']:,.2f}", batch_wrap_style)
                ]
                collected.append({
                    'company': company_name,
                    'amount': r['amount'],
                    'balance': r['unpaid_amount'],
                    'row': row_data
                })

            # 按公司稳定排序，公司内部保留发票日期倒序
            collected.sort(key=lambda r: r['company'] or '')

            # 小计/总计行样式
            subtotal_wrap_style = ParagraphStyle(
                'SubtotalWrapStyle',
                parent=getSampleStyleSheet()['Normal'],
                fontSize=8,
                leading=9,
                leftIndent=0,
                rightIndent=0,
                spaceBefore=0,
                spaceAfter=0
            )
            total_wrap_style = ParagraphStyle(
                'TotalWrapStyle',
                parent=getSampleStyleSheet()['Normal'],
                fontSize=8,
                leading=9,
                leftIndent=0,
                rightIndent=0,
                spaceBefore=0,
                spaceAfter=0
            )

            def _make_subtotal_row(company, amount_sub, bal_sub, count):
                # 列：INV NO / HID / COMPANY / INV.DATE / CLIENT / DP.DATE / ITIN / CCY / AMOUNT / BAL
                return [
                    Paragraph('', subtotal_wrap_style),
                    Paragraph('', subtotal_wrap_style),
                    Paragraph(f"<b>{company or ''}</b>", subtotal_wrap_style),
                    Paragraph('', subtotal_wrap_style),
                    Paragraph('', subtotal_wrap_style),
                    Paragraph('', subtotal_wrap_style),
                    Paragraph(f"<b>SUBTOTAL ({count} item{'s' if count > 1 else ''})</b>", subtotal_wrap_style),
                    Paragraph('SGD', subtotal_wrap_style),
                    Paragraph(f"<b>${amount_sub:,.2f}</b>", subtotal_wrap_style),
                    Paragraph(f"<b>${bal_sub:,.2f}</b>", subtotal_wrap_style)
                ]

            pdf_data = []
            subtotal_pdf_indices = []
            companies_seen = []
            current_company = None
            company_amount_subtotal = 0.0
            company_bal_subtotal = 0.0
            company_count = 0
            grand_amount = 0.0
            grand_bal = 0.0

            for item in collected:
                if current_company is None:
                    current_company = item['company']
                elif item['company'] != current_company:
                    pdf_data.append(_make_subtotal_row(current_company, company_amount_subtotal, company_bal_subtotal, company_count))
                    subtotal_pdf_indices.append(len(pdf_data) - 1)
                    companies_seen.append(current_company)
                    current_company = item['company']
                    company_amount_subtotal = 0.0
                    company_bal_subtotal = 0.0
                    company_count = 0

                pdf_data.append(item['row'])
                company_amount_subtotal += item['amount']
                company_bal_subtotal += item['balance']
                company_count += 1
                grand_amount += item['amount']
                grand_bal += item['balance']

            # 最后一家公司的小计
            pdf_data.append(_make_subtotal_row(current_company, company_amount_subtotal, company_bal_subtotal, company_count))
            subtotal_pdf_indices.append(len(pdf_data) - 1)
            companies_seen.append(current_company)

            # 仅当有多家公司时才显示总计行
            grand_total_pdf_index = None
            if len(companies_seen) > 1:
                pdf_data.append([
                    Paragraph('', total_wrap_style),
                    Paragraph('', total_wrap_style),
                    Paragraph('', total_wrap_style),
                    Paragraph('', total_wrap_style),
                    Paragraph('', total_wrap_style),
                    Paragraph('', total_wrap_style),
                    Paragraph('<b>GRAND TOTAL</b>', total_wrap_style),
                    Paragraph('<b>SGD</b>', total_wrap_style),
                    Paragraph(f"<b>${grand_amount:,.2f}</b>", total_wrap_style),
                    Paragraph(f"<b>${grand_bal:,.2f}</b>", total_wrap_style)
                ])
                grand_total_pdf_index = len(pdf_data) - 1

            # 生成PDF
            pdf_buffer = io.BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER
            )

            story = []

            # 标题
            story.append(Paragraph("SOA Statement of Account", title_style))
            story.append(Spacer(1, 12))

            # 筛选条件说明（同时展示集团/公司/日期等条件）
            filter_parts = []
            if group:
                filter_parts.append(f"Group: {group}")
            if company:
                filter_parts.append(f"Company: {company}")
            # 日期区间优先于 month
            if start_date or end_date:
                filter_parts.append(f"Inv.Date: {start_date or '...'} ~ {end_date or '...'}")
            elif month:
                filter_parts.append(f"Month: {month}")
            filter_text = " | ".join(filter_parts) if filter_parts else "All Data"

            story.append(Paragraph(filter_text, styles['Normal']))
            story.append(Spacer(1, 12))

            # 创建批量PDF表头换行样式
            batch_header_style = ParagraphStyle(
                'BatchHeaderStyle',
                parent=getSampleStyleSheet()['Normal'],
                fontSize=10,
                leading=12,
                leftIndent=0,
                rightIndent=0,
                spaceBefore=0,
                spaceAfter=0
            )

            # 创建表格（10 列：INV NO/HID/COMPANY/INV.DATE/CLIENT/DP.DATE/ITIN/CCY/AMOUNT/BAL）
            table_data = [
                [Paragraph('INV NO', batch_header_style),
                 Paragraph('HID', batch_header_style),
                 Paragraph('COMPANY', batch_header_style),
                 Paragraph('INV.DATE', batch_header_style),
                 Paragraph('CLIENT NAME', batch_header_style),
                 Paragraph('DP.DATE', batch_header_style),
                 Paragraph('ITIN DESCRIPTION', batch_header_style),
                 Paragraph('CCY', batch_header_style),
                 Paragraph('AMOUNT', batch_header_style),
                 Paragraph('BAL', batch_header_style)]
            ] + pdf_data

            table = Table(table_data, colWidths=[0.7*inch, 0.55*inch, 1.65*inch, 0.7*inch, 1.4*inch, 0.7*inch, 1.65*inch, 0.5*inch, 0.85*inch, 0.85*inch])

            # 基础样式：表头绿、数据行浅灰、整表网格
            style_cmds = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ]

            # 小计行：浅黄底、深棕字、加粗。pdf_data 索引 +1 即表格行索引（第 0 行是表头）
            for idx in subtotal_pdf_indices:
                row = idx + 1
                style_cmds.extend([
                    ('BACKGROUND', (0, row), (-1, row), colors.HexColor('#FFF3CD')),
                    ('TEXTCOLOR', (0, row), (-1, row), colors.HexColor('#664D03')),
                    ('FONTNAME', (0, row), (-1, row), 'Helvetica-Bold'),
                ])

            # 总计行：浅绿底、深蓝字、加粗（仅当存在）
            if grand_total_pdf_index is not None:
                row = grand_total_pdf_index + 1
                style_cmds.extend([
                    ('BACKGROUND', (0, row), (-1, row), colors.HexColor('#e8f5e8')),
                    ('TEXTCOLOR', (0, row), (-1, row), colors.HexColor('#000080')),
                    ('FONTNAME', (0, row), (-1, row), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, row), (-1, row), 10),
                ])

            table.setStyle(TableStyle(style_cmds))
            story.append(table)

            # 银行信息
            batch_bank_style = ParagraphStyle(
                'BatchBankStyle',
                parent=getSampleStyleSheet()['Normal'],
                fontSize=9,
                leading=11,
                leftIndent=0,
                alignment=TA_LEFT
            )

            story.append(Spacer(1, 20))
            story.append(Paragraph("<b>JOYFUL ESCAPES PTE. LTD Bank Details:</b>", batch_bank_style))
            story.append(Paragraph("• OCBC Bank Singapore, AC No.: 5956-7793-1001", batch_bank_style))
            story.append(Paragraph("• PayNow UEN: 202337627W", batch_bank_style))
            story.append(Spacer(1, 20))

            doc.build(story)
            pdf_buffer.seek(0)

            return pdf_buffer.getvalue(), None

        except Exception as e:
            import traceback
            traceback.print_exc()
            return None, f"生成PDF文件时出错: {str(e)}"

    # 以下方法保留用于单个项目预览（如果需要），但现在基于Project数据
    def generate_soa_html(self, header_id):
        """生成HTML格式的SOA（单个项目）"""
        try:
            header = ProjectHeader.query.options(
                db.joinedload(ProjectHeader.company)
            ).get(header_id)

            if not header:
                return None, "项目记录不存在"

            # 获取REF明细
            refs = ProjectRef.query.filter_by(header_id=header_id).all()

            # 计算财务数据
            total_selling = sum(float(r.selling_price or 0) for r in refs)
            total_cost = sum(float(r.cost_price or 0) for r in refs)
            balance = self._calculate_project_balance(header_id, total_selling)

            current_date = datetime.now().strftime('%Y-%m-%d')
            current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # 获取联系人名称：优先从人员名单获取leader，其次使用header的leader_name或contact
            contact_name = header.leader_member_name or header.leader_name or header.contact or 'N/A'

            # 这里可以渲染一个新的模板，或者返回简单的HTML
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>SOA - {header.hid}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    h1 {{ color: #28a745; }}
                    table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #28a745; color: white; }}
                    .total {{ background-color: #e8f5e8; font-weight: bold; }}
                </style>
            </head>
            <body>
                <h1>Statement of Account</h1>
                <p><strong>Project:</strong> {header.hid}</p>
                <p><strong>Company:</strong> {header.company.company_name if header.company else 'N/A'}</p>
                <p><strong>Contact:</strong> {contact_name}</p>
                <p><strong>Date:</strong> {current_date}</p>

                <table>
                    <thead>
                        <tr>
                            <th>REF</th>
                            <th>Description</th>
                            <th>Selling</th>
                            <th>Cost</th>
                        </tr>
                    </thead>
                    <tbody>
            """

            for ref in refs:
                html_content += f"""
                        <tr>
                            <td>{ref.ref_number}</td>
                            <td>{ref.description or ref.detailed_description or ''}</td>
                            <td>${float(ref.selling_price or 0):,.2f}</td>
                            <td>${float(ref.cost_price or 0):,.2f}</td>
                        </tr>
                """

            html_content += f"""
                        <tr class="total">
                            <td colspan="2">Total</td>
                            <td>${total_selling:,.2f}</td>
                            <td>${total_cost:,.2f}</td>
                        </tr>
                    </tbody>
                </table>

                <p style="margin-top: 20px;"><strong>Balance Due:</strong> ${balance:,.2f}</p>

                <hr style="margin-top: 30px;">
                <p><strong>JOYFUL ESCAPES PTE. LTD Bank Details:</strong></p>
                <p>• OCBC Bank Singapore, AC No.: 5956-7793-1001</p>
                <p>• PayNow UEN: 202337627W</p>

                <p style="margin-top: 20px; color: #666; font-size: 12px;">Generated: {current_datetime}</p>
            </body>
            </html>
            """

            return html_content, None

        except Exception as e:
            import traceback
            traceback.print_exc()
            return None, f"生成SOA时出错: {str(e)}"

    def generate_soa_pdf(self, header_id):
        """生成PDF格式的SOA（单个项目）"""
        try:
            header = ProjectHeader.query.options(
                db.joinedload(ProjectHeader.company)
            ).get(header_id)

            if not header:
                return None, "项目记录不存在"

            refs = ProjectRef.query.filter_by(header_id=header_id).all()

            total_selling = sum(float(r.selling_price or 0) for r in refs)
            total_cost = sum(float(r.cost_price or 0) for r in refs)
            balance = self._calculate_project_balance(header_id, total_selling)

            # 创建PDF
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER
            )

            story = []

            story.append(Paragraph("Statement of Account", title_style))
            story.append(Spacer(1, 20))

            # 获取联系人名称：优先从人员名单获取leader
            contact_name = header.leader_member_name or header.leader_name or header.contact or 'N/A'

            # 项目信息
            company_info = [
                ['Project Information', 'Statement Information'],
                [f'Project: {header.hid}', f'Date: {datetime.now().strftime("%Y-%m-%d")}'],
                [f'Company: {header.company.company_name if header.company else "N/A"}', f'Currency: SGD'],
                [f'Contact: {contact_name}', '']
            ]

            company_table = Table(company_info, colWidths=[4.5*inch, 4.5*inch])
            company_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC'))
            ]))

            story.append(company_table)
            story.append(Spacer(1, 20))

            # REF明细
            story.append(Paragraph("Transaction Details", styles['Heading2']))

            wrap_style = ParagraphStyle(
                'WrapStyle',
                parent=styles['Normal'],
                fontSize=8,
                leading=9
            )

            detail_data = [
                [Paragraph('REF', wrap_style),
                 Paragraph('Description', wrap_style),
                 Paragraph('Selling', wrap_style),
                 Paragraph('Cost', wrap_style)]
            ]

            for ref in refs:
                detail_data.append([
                    Paragraph(ref.ref_number or '', wrap_style),
                    Paragraph(ref.description or ref.detailed_description or '', wrap_style),
                    Paragraph(f"${float(ref.selling_price or 0):,.2f}", wrap_style),
                    Paragraph(f"${float(ref.cost_price or 0):,.2f}", wrap_style)
                ])

            # 总计行
            detail_data.append([
                Paragraph('TOTAL', wrap_style),
                Paragraph('', wrap_style),
                Paragraph(f"${total_selling:,.2f}", wrap_style),
                Paragraph(f"${total_cost:,.2f}", wrap_style)
            ])

            detail_table = Table(detail_data, colWidths=[1.5*inch, 4*inch, 1.5*inch, 1.5*inch])
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#f8f9fa')),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f5e8')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC'))
            ]))

            story.append(detail_table)
            story.append(Spacer(1, 20))

            # 余额
            story.append(Paragraph(f"<b>Balance Due: ${balance:,.2f}</b>", styles['Normal']))
            story.append(Spacer(1, 20))

            # 银行信息
            bank_style = ParagraphStyle(
                'BankStyle',
                parent=styles['Normal'],
                fontSize=10,
                leading=12,
                alignment=TA_LEFT
            )

            story.append(Paragraph("<b>JOYFUL ESCAPES PTE. LTD Bank Details:</b>", bank_style))
            story.append(Paragraph("• OCBC Bank Singapore, AC No.: 5956-7793-1001", bank_style))
            story.append(Paragraph("• PayNow UEN: 202337627W", bank_style))
            story.append(Spacer(1, 20))

            story.append(Paragraph(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))

            doc.build(story)
            buffer.seek(0)

            return buffer.getvalue(), None

        except Exception as e:
            import traceback
            traceback.print_exc()
            return None, f"生成PDF时出错: {str(e)}"
