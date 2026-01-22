from flask import Blueprint, render_template, request, redirect, url_for, jsonify, flash, current_app
from flask_login import login_required, current_user
from App_new.business.projects.models.ref import ProjectRef
from App_new.business.projects.models.eo import ProjectEO
from App_new.exts import csrf, db
from App_new.business.projects.forms.eo_forms import ProjectEOForm
from App_new.utils.decorators import staff_only, admin_only
from App_new.utils.permissions import can_access_project
import json

project_eo = Blueprint('project_eo', __name__)

@project_eo.route('/create/<int:ref_id>', methods=['GET', 'POST'])
@login_required
@staff_only
def create_eo(ref_id):
    """创建EO"""
    ref = ProjectRef.query.get_or_404(ref_id)
    
    # 员工等级权限检查
    from App_new.business.projects.models.project import ProjectHeader
    header = ProjectHeader.query.get(ref.header_id)
    if header and not can_access_project(header, current_user):
        flash('您没有权限访问此项目', 'error')
        return redirect(url_for('business_projects.list.list_projects'))
    form = ProjectEOForm()
    form.ref_id.data = ref_id
    
    if form.validate_on_submit():
        try:
            eo_number = ProjectEO.generate_eo_number()
            eo = ProjectEO(
                ref_id=ref.id,
                eo_number=eo_number,
                external_system=form.external_system.data,
                external_status=form.external_status.data,
                external_reference=form.external_reference.data,
                status=form.status.data
            )
            db.session.add(eo)
            db.session.commit()
            flash('EO子明细创建成功！', 'success')
            return redirect(url_for('project_ref.ref_detail', ref_id=ref.id))
        except Exception as e:
            db.session.rollback()
            flash(f'创建失败：{str(e)}', 'error')
    elif form.errors:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}', 'error')
    
    # 预填充EO编号
    eo_number = ProjectEO.generate_eo_number()
    form.eo_number.data = eo_number
    
    # 从REF中自动获取和预填充数据
    if request.method == 'GET':
        # 预填充状态（默认为已确认）
        if not form.status.data:
            form.status.data = 'confirmed'
    
    return render_template('business/projects/project_eo/create_eo.html',
                           form=form, ref=ref, eo_number=eo_number)

@project_eo.route('/<int:eo_id>')
@login_required
@staff_only
def eo_detail(eo_id):
    """EO详情页面"""
    eo = ProjectEO.query.get_or_404(eo_id)
    
    # 员工等级权限检查
    from App_new.business.projects.models.project import ProjectHeader
    header = ProjectHeader.query.get(eo.ref.header_id)
    if header and not can_access_project(header, current_user):
        flash('您没有权限访问此项目', 'error')
        return redirect(url_for('business_projects.project_eo.eo_list'))
    return render_template('business/projects/project_eo/eo_detail.html', eo=eo)


@project_eo.route('/<int:eo_id>/update', methods=['POST'])
@login_required
@staff_only
@csrf.exempt
def update_eo(eo_id):
    """更新EO信息"""
    try:
        eo = ProjectEO.query.get_or_404(eo_id)

        # 员工等级权限检查
        from App_new.business.projects.models.project import ProjectHeader
        header = ProjectHeader.query.get(eo.ref.header_id)
        if header and not can_access_project(header, current_user):
            return jsonify({'success': False, 'message': '您没有权限操作此EO'}), 403

        # 已付款的EO不能编辑
        if eo.status == 'paid':
            return jsonify({'success': False, 'message': '已付款的EO不能编辑'}), 400

        # 获取JSON数据
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '无效的请求数据'}), 400
        
        # 更新字段
        from datetime import datetime
        
        # EO日期使用创建日期，不需要更新
        # if 'eo_date' in data:
        #     eo.eo_date = datetime.strptime(data['eo_date'], '%Y-%m-%d').date() if data['eo_date'] else None
        
        if 'conf_code' in data:
            eo.conf_code = data['conf_code'] if data['conf_code'] else None
        
        if 'discount' in data:
            eo.discount = float(data['discount']) if data['discount'] else 0
        
        if 'tax' in data:
            eo.tax = float(data['tax']) if data['tax'] else 0
        
        if 'payment_no' in data:
            eo.payment_no = data['payment_no'] if data['payment_no'] else None
        
        if 'paid_date' in data:
            eo.paid_date = datetime.strptime(data['paid_date'], '%Y-%m-%d').date() if data['paid_date'] else None
        
        if 'pay_amount' in data:
            eo.pay_amount = float(data['pay_amount']) if data['pay_amount'] else None
        
        if 'external_system' in data:
            eo.external_system = data['external_system'] if data['external_system'] else None
        
        if 'external_status' in data:
            eo.external_status = data['external_status'] if data['external_status'] else None
        
        if 'external_reference' in data:
            eo.external_reference = data['external_reference'] if data['external_reference'] else None
        
        if 'status' in data:
            eo.status = data['status']
        
        eo.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'EO更新成功'})
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"更新EO失败: {e}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'更新失败：{str(e)}'}), 500


@project_eo.route('/exchange-order')
@login_required
@staff_only
def exchange_order():
    """Exchange Order 付款页面 - 类似旧系统的付款管理界面"""
    try:
        from sqlalchemy import and_, or_, desc
        from sqlalchemy.orm import aliased
        from App_new.business.projects.models.project import ProjectHeader, CustomerCompany
        from App_new.shared.models.business_types import BusinessType
        from datetime import date

        # 创建别名：SupplierCompany 用于供应商，CustomerCompany 用于客户
        SupplierCompany = aliased(CustomerCompany, name='supplier_company')

        # 获取筛选参数
        supplier_id = request.args.get('supplier', None, type=int)
        date_filter_by = request.args.get('date_filter_by', 'eo_date')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')

        # 分页参数
        page = request.args.get('page', 1, type=int)
        per_page = 25  # 每页显示25条

        # 构建查询 - 查询未付款的EO（confirmed状态）
        query = db.session.query(
            ProjectEO,
            SupplierCompany.company_name.label('supplier_name'),
            BusinessType.name.label('ref_type_name'),
            ProjectRef.ref_number.label('ref_number'),
            ProjectRef.description.label('description'),
            ProjectRef.header_id.label('project_id'),
            ProjectHeader.hid.label('project_hid'),
            CustomerCompany.company_name.label('company_name')
        ).join(
            ProjectRef, ProjectEO.ref_id == ProjectRef.id, isouter=True
        ).join(
            BusinessType, ProjectRef.ref_type_id == BusinessType.id, isouter=True
        ).join(
            SupplierCompany, ProjectRef.supplier_id == SupplierCompany.id, isouter=True
        ).join(
            ProjectHeader, ProjectRef.header_id == ProjectHeader.id, isouter=True
        ).join(
            CustomerCompany, ProjectHeader.company_id == CustomerCompany.id, isouter=True
        ).filter(
            ProjectEO.status == 'confirmed'
        )
        
        # 应用供应商筛选
        if supplier_id:
            query = query.filter(ProjectRef.supplier_id == supplier_id)
        
        # 应用日期筛选
        if date_from:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            if date_filter_by == 'eo_date':
                query = query.filter(ProjectEO.created_at >= date_from_obj)
        
        if date_to:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            if date_filter_by == 'eo_date':
                query = query.filter(ProjectEO.created_at <= date_to_obj)
        
        # 排序
        query = query.order_by(desc(ProjectEO.created_at))
        
        # 使用分页获取结果
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        results = pagination.items
        
        # 批量获取所有 ref_id 对应的乘客信息，避免 N+1 问题
        from App_new.business.flight.models.flight import ProjectFlightPassenger
        ref_ids = [eo.ref_id for eo, *_ in results if eo.ref_id]
        passengers_by_ref = {}
        if ref_ids:
            all_passengers = ProjectFlightPassenger.query.filter(ProjectFlightPassenger.ref_id.in_(ref_ids)).all()
            for p in all_passengers:
                if p.ref_id not in passengers_by_ref:
                    passengers_by_ref[p.ref_id] = []
                if p.name:
                    passengers_by_ref[p.ref_id].append(p.name)
        
        # 处理数据
        eos = []
        for eo, supplier_name, ref_type_name, ref_number, description, project_id, project_hid, company_name in results:
            # 从批量查询结果获取乘客姓名
            pax_names = ', '.join(passengers_by_ref.get(eo.ref_id, []))
            
            cost_price = float(eo.ref.cost_price) if eo.ref and eo.ref.cost_price else 0
            paid_amount = float(eo.pay_amount) if eo.pay_amount else 0
            balance = cost_price - paid_amount
            
            eos.append({
                'id': eo.id,
                'eo_number': eo.eo_number,
                'ref_id': eo.ref_id,
                'ref_number': ref_number or '',
                'description': description or '',
                'supplier_name': supplier_name or '',
                'ref_type_name': ref_type_name or '',
                'project_id': project_id,
                'project_hid': project_hid or '',
                'company_name': company_name or '',
                'pax_names': pax_names,
                'cost_price': cost_price,
                'balance': balance,
                'pay_amount': paid_amount,
                'currency': eo.ref.currency if eo.ref and eo.ref.currency else 'SGD',
                'created_at': eo.created_at,
                'dep_date': None,  # 可从extra_info获取
                'due_date': None,
                'inv_number': None,
                'inv_amount': None,
                'conf_by': None,
                'rate': 1.00
            })
        
        # 获取供应商列表（按名称排序，不区分大小写）
        from sqlalchemy import func
        suppliers = CustomerCompany.query.filter(CustomerCompany.is_supplier == True).order_by(func.lower(CustomerCompany.company_name)).all()

        return render_template('business/projects/project_eo/exchange_order.html',
                             eos=eos,
                             suppliers=suppliers,
                             today=date.today().isoformat(),
                             pagination=pagination,
                             current_filters={
                                 'supplier': supplier_id,
                                 'date_filter_by': date_filter_by,
                                 'date_from': date_from,
                                 'date_to': date_to
                             })
    except Exception as e:
        import traceback
        print(f"Exchange Order页面加载失败: {str(e)}")
        traceback.print_exc()
        flash(f'页面加载失败：{str(e)}', 'error')
        return redirect(url_for('business_projects.project_eo.eo_list'))


@project_eo.route('/exchange-order/generate-payment-no', methods=['GET'])
@csrf.exempt
@login_required
@staff_only
def generate_payment_no():
    """生成付款编号"""
    try:
        from datetime import datetime
        from App_new.business.projects.models.supplier_payment import SupplierPayment

        # 格式：PAY-YYYYMMDD-XXX
        today = datetime.now().strftime('%Y%m%d')
        prefix = f'PAY-{today}-'

        max_num = 0

        # 检查 ProjectEO 表中今天最后一个付款编号
        last_eo = ProjectEO.query.filter(
            ProjectEO.payment_no.like(f'{prefix}%')
        ).order_by(ProjectEO.payment_no.desc()).first()

        if last_eo and last_eo.payment_no:
            try:
                eo_num = int(last_eo.payment_no.split('-')[-1])
                max_num = max(max_num, eo_num)
            except:
                pass

        # 检查 SupplierPayment 表中今天最后一个付款编号
        last_payment = SupplierPayment.query.filter(
            SupplierPayment.payment_no.like(f'{prefix}%')
        ).order_by(SupplierPayment.payment_no.desc()).first()

        if last_payment and last_payment.payment_no:
            try:
                payment_num = int(last_payment.payment_no.split('-')[-1])
                max_num = max(max_num, payment_num)
            except:
                pass

        new_num = max_num + 1
        payment_no = f'{prefix}{str(new_num).zfill(3)}'

        return jsonify({
            'success': True,
            'payment_no': payment_no
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e),
            'payment_no': f'PAY-{datetime.now().strftime("%Y%m%d")}-001'
        })


@project_eo.route('/exchange-order/pay', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def exchange_order_pay():
    """Exchange Order 付款提交"""
    try:
        data = request.get_json()
        eo_ids = data.get('eo_ids', [])
        payment_no = data.get('payment_no')
        paid_date_str = data.get('paid_date')
        total_pay_amount = data.get('pay_amount')
        remarks = data.get('remarks', '')
        
        if not eo_ids:
            return jsonify({'success': False, 'message': 'Please select at least one EO'}), 400
        
        if not payment_no or not paid_date_str or total_pay_amount is None:
            return jsonify({'success': False, 'message': 'Please fill in all required fields'}), 400
        
        from datetime import datetime
        paid_date = datetime.strptime(paid_date_str, '%Y-%m-%d').date()
        
        eos = ProjectEO.query.filter(ProjectEO.id.in_(eo_ids)).all()
        
        if len(eos) != len(eo_ids):
            return jsonify({'success': False, 'message': 'Some EOs not found'}), 400
        
        total_cost = sum(float(eo.ref.cost_price) if eo.ref and eo.ref.cost_price else 0 for eo in eos)
        
        success_count = 0
        for eo in eos:
            if eo.status in ['void', 'paid']:
                continue
            
            eo_cost = float(eo.ref.cost_price) if eo.ref and eo.ref.cost_price else 0
            if total_cost > 0:
                eo_pay_amount = (eo_cost / total_cost) * total_pay_amount
            else:
                eo_pay_amount = total_pay_amount / len(eos)
            
            eo.payment_no = payment_no
            eo.paid_date = paid_date
            eo.pay_amount = round(eo_pay_amount, 2)
            eo.status = 'paid'
            success_count += 1

            # 自动创建日记账分录（借：成本费用，贷：银行存款）
            try:
                from App_new.finance.models.journal_entry import JournalEntry
                journal_entry = JournalEntry.create_from_eo(
                    eo,
                    user=current_user.username if current_user else None
                )
                if journal_entry and journal_entry.lines:
                    db.session.add(journal_entry)
                    # 自动过账
                    journal_entry.post(user=current_user.username if current_user else None)
            except Exception as je_error:
                # 日记账创建失败不影响EO付款
                import logging
                logging.getLogger(__name__).warning(f"创建Exchange Order日记账失败: {str(je_error)}")

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Successfully paid {success_count} EO(s), Payment No: {payment_no}'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@project_eo.route('/exchange-order/cancel', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def exchange_order_cancel():
    """Exchange Order 批量取消/作废EO"""
    try:
        data = request.get_json()
        eo_ids = data.get('eo_ids', [])
        
        if not eo_ids:
            return jsonify({'success': False, 'message': 'Please select at least one EO'}), 400
        
        eos = ProjectEO.query.filter(ProjectEO.id.in_(eo_ids)).all()
        
        success_count = 0
        for eo in eos:
            if eo.status == 'void':
                continue
            eo.status = 'void'
            success_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully voided {success_count} EO(s)'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@project_eo.route('/batch-pay')
@login_required
@staff_only
def batch_pay():
    """批量付款页面 - 显示待付款的EO列表"""
    try:
        from sqlalchemy import and_, or_, desc
        from sqlalchemy.orm import aliased
        from App_new.business.projects.models.project import ProjectHeader, CustomerCompany
        from App_new.shared.models.business_types import BusinessType
        from datetime import date

        # 创建别名：SupplierCompany 用于供应商，CustomerCompany 用于客户
        SupplierCompany = aliased(CustomerCompany, name='supplier_company')

        # 获取筛选参数
        supplier_type = request.args.get('supplier_type', '')
        supplier_id = request.args.get('supplier', None, type=int)
        keyword = request.args.get('keyword', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        sort = request.args.get('sort', 'date_desc')  # 合并的排序参数

        # 解析排序参数
        sort_map = {
            'date_desc': ('created_at', 'desc'),
            'date_asc': ('created_at', 'asc'),
            'eo_desc': ('eo_number', 'desc'),
            'eo_asc': ('eo_number', 'asc'),
        }
        sort_by, sort_order = sort_map.get(sort, ('created_at', 'desc'))

        # 分页参数
        page = request.args.get('page', 1, type=int)
        per_page = 25  # 每页显示25条

        # 构建查询 - 只查询未付款的EO（confirmed状态）
        query = db.session.query(
            ProjectEO,
            SupplierCompany.company_name.label('supplier_name'),
            BusinessType.name.label('ref_type_name'),
            ProjectRef.ref_number.label('ref_number'),
            ProjectRef.header_id.label('project_id'),
            ProjectHeader.hid.label('project_hid'),
            CustomerCompany.company_name.label('company_name'),
            ProjectRef.supplier_id.label('supplier_id')
        ).join(
            ProjectRef, ProjectEO.ref_id == ProjectRef.id, isouter=True
        ).join(
            BusinessType, ProjectRef.ref_type_id == BusinessType.id, isouter=True
        ).join(
            SupplierCompany, ProjectRef.supplier_id == SupplierCompany.id, isouter=True
        ).join(
            ProjectHeader, ProjectRef.header_id == ProjectHeader.id, isouter=True
        ).join(
            CustomerCompany, ProjectHeader.company_id == CustomerCompany.id, isouter=True
        ).filter(
            ProjectEO.status == 'confirmed'  # 只显示已确认但未付款的
        )
        
        # 应用筛选条件
        if supplier_type:
            supplier_type_map = {
                'visa': '签证', 'flight': '机票', 'hotel': '酒店',
                'transport': '用车', 'local_operator': '地接', 'other': '其他'
            }
            type_name = supplier_type_map.get(supplier_type)
            if type_name:
                query = query.filter(BusinessType.name == type_name)
        
        if supplier_id:
            query = query.filter(ProjectRef.supplier_id == supplier_id)
        
        if keyword:
            keyword_filter = or_(
                ProjectEO.eo_number.ilike(f'%{keyword}%'),
                ProjectRef.ref_number.ilike(f'%{keyword}%'),
                ProjectHeader.hid.ilike(f'%{keyword}%')
            )
            query = query.filter(keyword_filter)
        
        # 日期筛选
        if start_date:
            from datetime import datetime
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(ProjectEO.created_at >= start_dt)
            except ValueError:
                pass
        
        if end_date:
            from datetime import datetime, timedelta
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(ProjectEO.created_at < end_dt)
            except ValueError:
                pass

        # 排序
        from sqlalchemy import asc
        if sort_by == 'eo_number':
            sort_column = ProjectEO.eo_number
        else:  # 默认按创建日期
            sort_column = ProjectEO.created_at

        if sort_order == 'asc':
            query = query.order_by(asc(sort_column))
        else:
            query = query.order_by(desc(sort_column))

        # 使用分页获取结果
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        results = pagination.items
        
        # 批量获取乘客信息，避免 N+1 问题
        import json
        from App_new.business.projects.models.project_member import ProjectMember
        from App_new.business.flight.models.flight import ProjectFlightPassenger
        
        ref_ids = [eo.ref_id for eo, *_ in results if eo.ref_id]
        passengers_by_ref = {}
        if ref_ids:
            all_passengers = ProjectFlightPassenger.query.filter(ProjectFlightPassenger.ref_id.in_(ref_ids)).all()
            for p in all_passengers:
                if p.ref_id not in passengers_by_ref:
                    passengers_by_ref[p.ref_id] = []
                if p.name:
                    passengers_by_ref[p.ref_id].append(p.name)
        
        # 处理数据
        eos = []
        for eo, supplier_name, ref_type_name, ref_number, project_id, project_hid, company_name, ref_supplier_id in results:
            # 获取乘客姓名 - 支持所有类型的REF
            pax_names = ''
            if eo.ref:
                # 方法1: 优先从extra_info获取（机票REF可能已有pax_names_display）
                if eo.ref.extra_info:
                    try:
                        extra_data = json.loads(eo.ref.extra_info)
                        # 如果已有pax_names_display，直接使用
                        if extra_data.get('pax_names_display'):
                            pax_names = extra_data.get('pax_names_display')
                        elif extra_data.get('pax_name'):
                            pax_names = extra_data.get('pax_name', '')
                    except (json.JSONDecodeError, TypeError):
                        pass

                # 方法2: 从批量查询结果获取乘客姓名
                if not pax_names and eo.ref_id in passengers_by_ref:
                    pax_names = ', '.join(passengers_by_ref[eo.ref_id])

            eos.append({
                'id': eo.id,
                'eo_number': eo.eo_number,
                'ref_id': eo.ref_id,
                'ref_number': ref_number or '',
                'supplier_id': ref_supplier_id,
                'supplier_name': supplier_name or '',
                'ref_type_name': ref_type_name or '',
                'project_id': project_id,
                'project_hid': project_hid or '',
                'company_name': company_name or '',
                'pax_names': pax_names,
                'cost_price': float(eo.ref.cost_price) if eo.ref and eo.ref.cost_price else 0,
                'currency': eo.ref.currency if eo.ref and eo.ref.currency else 'SGD',
                'created_at': eo.created_at
            })
        
        # 获取所有供应商列表（按名称排序，不区分大小写）
        # 前端使用JavaScript按类型筛选
        from sqlalchemy import func
        suppliers_with_type = db.session.query(
            CustomerCompany,
            BusinessType.code.label('supplier_type_code')
        ).outerjoin(
            BusinessType, CustomerCompany.supplier_type_id == BusinessType.id
        ).filter(
            CustomerCompany.is_supplier == True
        ).order_by(func.lower(CustomerCompany.company_name)).all()

        suppliers = [{
            'id': s.CustomerCompany.id,
            'company_name': s.CustomerCompany.company_name,
            'supplier_type_code': s.supplier_type_code or ''
        } for s in suppliers_with_type]

        # 获取供应商类型列表（用于前端筛选下拉框）
        supplier_types = BusinessType.query.filter_by(is_active=True).order_by(BusinessType.sort_order).all()

        return render_template('business/projects/project_eo/batch_pay.html',
                             eos=eos,
                             suppliers=suppliers,
                             supplier_types=supplier_types,
                             today=date.today().isoformat(),
                             pagination=pagination,
                             current_filters={
                                 'supplier_type': supplier_type,
                                 'supplier': supplier_id,
                                 'keyword': keyword,
                                 'start_date': start_date,
                                 'end_date': end_date,
                                 'sort': sort
                             })
    except Exception as e:
        import traceback
        print(f"批量付款页面加载失败: {str(e)}")
        traceback.print_exc()
        flash(f'页面加载失败：{str(e)}', 'error')
        return redirect(url_for('business_projects.project_eo.eo_list'))


@project_eo.route('/batch-pay/export')
@login_required
@staff_only
def batch_pay_export():
    """导出批量付款EO列表为Excel"""
    try:
        from sqlalchemy import and_, or_, desc, asc
        from sqlalchemy.orm import aliased
        from App_new.business.projects.models.project import ProjectHeader, CustomerCompany
        from App_new.shared.models.business_types import BusinessType
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from flask import send_file

        # 创建别名
        SupplierCompany = aliased(CustomerCompany, name='supplier_company')

        # 获取筛选参数（与batch_pay相同）
        supplier_type = request.args.get('supplier_type', '')
        supplier_id = request.args.get('supplier', None, type=int)
        keyword = request.args.get('keyword', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        sort = request.args.get('sort', 'date_desc')

        # 解析排序参数
        sort_map = {
            'date_desc': ('created_at', 'desc'),
            'date_asc': ('created_at', 'asc'),
            'eo_desc': ('eo_number', 'desc'),
            'eo_asc': ('eo_number', 'asc'),
        }
        sort_by, sort_order = sort_map.get(sort, ('created_at', 'desc'))

        # 构建查询
        query = db.session.query(
            ProjectEO,
            SupplierCompany.company_name.label('supplier_name'),
            BusinessType.name.label('ref_type_name'),
            ProjectRef.ref_number.label('ref_number'),
            ProjectHeader.hid.label('project_hid'),
            ProjectRef.cost_price.label('cost_price'),
            ProjectRef.currency.label('currency')
        ).join(
            ProjectRef, ProjectEO.ref_id == ProjectRef.id, isouter=True
        ).join(
            BusinessType, ProjectRef.ref_type_id == BusinessType.id, isouter=True
        ).join(
            SupplierCompany, ProjectRef.supplier_id == SupplierCompany.id, isouter=True
        ).join(
            ProjectHeader, ProjectRef.header_id == ProjectHeader.id, isouter=True
        ).filter(
            ProjectEO.status == 'confirmed'
        )

        # 应用筛选条件
        if supplier_type:
            supplier_type_map = {
                'visa': '签证', 'flight': '机票', 'hotel': '酒店',
                'transport': '用车', 'local_operator': '地接', 'other': '其他'
            }
            type_name = supplier_type_map.get(supplier_type)
            if type_name:
                query = query.filter(BusinessType.name == type_name)

        if supplier_id:
            query = query.filter(ProjectRef.supplier_id == supplier_id)

        if keyword:
            keyword_filter = or_(
                ProjectEO.eo_number.ilike(f'%{keyword}%'),
                ProjectRef.ref_number.ilike(f'%{keyword}%'),
                ProjectHeader.hid.ilike(f'%{keyword}%')
            )
            query = query.filter(keyword_filter)

        if start_date:
            from datetime import datetime
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(ProjectEO.created_at >= start_dt)
            except ValueError:
                pass

        if end_date:
            from datetime import datetime, timedelta
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(ProjectEO.created_at < end_dt)
            except ValueError:
                pass

        # 排序
        if sort_by == 'eo_number':
            sort_column = ProjectEO.eo_number
        else:
            sort_column = ProjectEO.created_at

        if sort_order == 'asc':
            query = query.order_by(asc(sort_column))
        else:
            query = query.order_by(desc(sort_column))

        # 获取全部结果（不分页）
        results = query.all()

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "待付款EO列表"

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头
        headers = ['Date', 'EO No', 'Currency', 'Cost', 'HID', 'REF', 'Type', 'Supplier']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据行
        for row_idx, (eo, supplier_name, ref_type_name, ref_number, project_hid, cost_price, currency) in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=eo.created_at.strftime('%Y-%m-%d') if eo.created_at else '')
            ws.cell(row=row_idx, column=2, value=eo.eo_number or '')
            ws.cell(row=row_idx, column=3, value=currency or 'SGD')
            ws.cell(row=row_idx, column=4, value=float(cost_price) if cost_price else 0)
            ws.cell(row=row_idx, column=5, value=project_hid or '')
            ws.cell(row=row_idx, column=6, value=ref_number or '')
            ws.cell(row=row_idx, column=7, value=ref_type_name or '')
            ws.cell(row=row_idx, column=8, value=supplier_name or '')

            # 设置边框
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).border = thin_border

        # 调整列宽
        column_widths = [12, 15, 10, 12, 10, 12, 12, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 生成文件名
        from datetime import datetime
        filename = f"batch_pay_eo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        print(f"导出Excel失败: {str(e)}")
        traceback.print_exc()
        flash(f'导出失败：{str(e)}', 'error')
        return redirect(url_for('business_projects.project_eo.batch_pay'))


@project_eo.route('/batch-pay/submit', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def batch_pay_submit():
    """批量付款提交 - 支持按供应商自动选择预付账款（FIFO），创建付款记录"""
    try:
        data = request.get_json()
        eo_ids = data.get('eo_ids', [])
        payment_no = data.get('payment_no')
        payment_voucher_no = data.get('payment_voucher_no', '')
        paid_date_str = data.get('paid_date')
        total_pay_amount = data.get('pay_amount')
        remarks = data.get('remarks', '')
        payment_source = data.get('payment_source', 'bank')  # bank 或 prepayment
        supplier_id = data.get('supplier_id')  # 供应商ID（使用预付账款时需要）

        if not eo_ids:
            return jsonify({'success': False, 'message': '请选择要付款的EO'}), 400

        if not payment_no or not paid_date_str or total_pay_amount is None:
            return jsonify({'success': False, 'message': '请填写完整的付款信息'}), 400

        # 解析日期
        from datetime import datetime
        from decimal import Decimal
        paid_date = datetime.strptime(paid_date_str, '%Y-%m-%d').date()
        total_pay_amount_decimal = Decimal(str(total_pay_amount))

        # 查询所有选中的EO
        eos = ProjectEO.query.filter(ProjectEO.id.in_(eo_ids)).all()

        if len(eos) != len(eo_ids):
            return jsonify({'success': False, 'message': '部分EO不存在'}), 400

        from App_new.business.projects.models.supplier_prepayment import SupplierPrepayment, PrepaymentUsage
        from App_new.business.projects.models.supplier_payment import SupplierPayment

        # 统计所有EO已有的预付使用记录（confirmed或pending状态）
        existing_usage_total = Decimal('0')
        for eo in eos:
            if eo.status not in ['void', 'paid']:
                existing_usage = PrepaymentUsage.query.filter(
                    PrepaymentUsage.eo_id == eo.id,
                    PrepaymentUsage.status.in_(['pending', 'confirmed'])
                ).first()
                if existing_usage:
                    existing_usage_total += existing_usage.amount

        # 获取供应商的可用预付账款列表（FIFO排序）
        available_prepayments = []
        prepayment_amount_used = Decimal('0')
        if payment_source == 'prepayment':
            if not supplier_id:
                return jsonify({'success': False, 'message': '请选择同一供应商的EO'}), 400

            # 按创建时间升序获取可用预付账款（FIFO）
            available_prepayments = SupplierPrepayment.query.filter(
                SupplierPrepayment.supplier_id == supplier_id,
                SupplierPrepayment.status.in_(['confirmed', 'partial_used']),
                SupplierPrepayment.balance_amount > 0
            ).order_by(SupplierPrepayment.created_at.asc()).all()

            if not available_prepayments:
                return jsonify({'success': False, 'message': '该供应商无可用预付账款'}), 400

            # 计算总可用余额（已有使用记录的金额不需要重新扣减）
            total_available = sum(p.balance_amount for p in available_prepayments) + existing_usage_total
            if total_pay_amount_decimal > total_available:
                return jsonify({'success': False, 'message': f'付款金额 {total_pay_amount} 超过可用预付余额 {total_available}'}), 400

        # 如果没有指定 supplier_id，尝试从 EO 获取
        if not supplier_id:
            # 尝试从第一个 EO 的 REF 获取供应商
            for eo in eos:
                if eo.ref and eo.ref.supplier_id:
                    supplier_id = eo.ref.supplier_id
                    break

        # 创建付款记录
        # 如果使用预付账款，prepayment_amount = total_amount
        prepayment_amount = total_pay_amount_decimal if payment_source == 'prepayment' else Decimal('0')
        payment_record = SupplierPayment(
            payment_no=payment_no,
            supplier_id=supplier_id,
            payment_date=paid_date,
            total_amount=total_pay_amount_decimal,
            currency='SGD',
            payment_source=payment_source,
            payment_voucher_no=payment_voucher_no if payment_voucher_no else None,
            prepayment_amount=prepayment_amount,
            eo_count=len([eo for eo in eos if eo.status not in ['void', 'paid']]),
            status='confirmed',
            remarks=remarks,
            created_by=current_user.username if current_user else None
        )
        db.session.add(payment_record)
        db.session.flush()  # 获取 payment_record.id

        # 计算每个EO的付款金额（按比例分配）
        total_cost = sum(float(eo.ref.cost_price) if eo.ref and eo.ref.cost_price else 0 for eo in eos)

        success_count = 0
        remaining_to_pay = total_pay_amount_decimal  # 剩余需要扣减的金额
        prepayment_index = 0  # 当前使用的预付账款索引

        for eo in eos:
            # 检查状态
            if eo.status in ['void', 'paid']:
                continue

            # 计算该EO的付款金额（按比例）
            eo_cost = float(eo.ref.cost_price) if eo.ref and eo.ref.cost_price else 0
            if total_cost > 0:
                eo_pay_amount = (eo_cost / total_cost) * float(total_pay_amount)
            else:
                eo_pay_amount = float(total_pay_amount) / len(eos)

            eo_pay_amount = round(eo_pay_amount, 2)
            eo_pay_amount_decimal = Decimal(str(eo_pay_amount))

            # 更新付款信息
            eo.payment_record_id = payment_record.id  # 关联付款记录
            eo.payment_no = payment_no
            eo.payment_voucher_no = payment_voucher_no
            eo.paid_date = paid_date
            eo.pay_amount = eo_pay_amount
            eo.payment_remarks = remarks
            eo.status = 'paid'
            success_count += 1

            # 检查是否有已存在的预付使用记录（pending或confirmed）
            existing_usage = PrepaymentUsage.query.filter(
                PrepaymentUsage.eo_id == eo.id,
                PrepaymentUsage.status.in_(['pending', 'confirmed'])
            ).first()

            if payment_source == 'prepayment' and available_prepayments:
                # 如果已有confirmed的使用记录，不需要再创建，只需更新描述
                if existing_usage and existing_usage.status == 'confirmed':
                    existing_usage.description = f'EO付款确认 {eo.eo_number} ({payment_no})'
                    prepayment_amount_used += existing_usage.amount
                    remaining_to_pay -= eo_pay_amount_decimal
                    continue  # 跳过后续的扣减逻辑

                # 处理已有的pending记录（兼容旧数据）
                if existing_usage and existing_usage.status == 'pending':
                    # 退还原预付账款余额
                    old_prepayment = SupplierPrepayment.query.get(existing_usage.prepayment_id)
                    if old_prepayment:
                        old_prepayment.balance_amount += existing_usage.amount
                        old_prepayment.update_status()
                    existing_usage.status = 'reversed'
                    existing_usage.description = f'批量付款重新分配'

                # 按FIFO从预付账款中扣减
                amount_to_deduct = eo_pay_amount_decimal
                while amount_to_deduct > 0 and prepayment_index < len(available_prepayments):
                    current_prepayment = available_prepayments[prepayment_index]

                    # 计算可从当前预付账款扣减的金额
                    deduct_amount = min(amount_to_deduct, current_prepayment.balance_amount)

                    if deduct_amount > 0:
                        # 创建预付使用记录
                        new_usage = PrepaymentUsage(
                            prepayment_id=current_prepayment.id,
                            eo_id=eo.id,
                            ref_id=eo.ref_id,
                            amount=deduct_amount,
                            usage_date=paid_date,
                            description=f'批量EO付款 {eo.eo_number} ({payment_no})',
                            status='confirmed',
                            created_by=current_user.username if current_user else None
                        )
                        db.session.add(new_usage)

                        # 扣减预付余额
                        current_prepayment.balance_amount -= deduct_amount
                        current_prepayment.update_status()

                        prepayment_amount_used += deduct_amount
                        amount_to_deduct -= deduct_amount

                    # 如果当前预付账款余额用完，移到下一个
                    if current_prepayment.balance_amount <= 0:
                        prepayment_index += 1

                remaining_to_pay -= eo_pay_amount_decimal

            else:
                # 使用银行付款，如果有预付使用记录需要退还（pending或confirmed）
                if existing_usage:
                    old_prepayment = SupplierPrepayment.query.get(existing_usage.prepayment_id)
                    if old_prepayment:
                        old_prepayment.balance_amount += existing_usage.amount
                        old_prepayment.update_status()
                    existing_usage.status = 'reversed'
                    existing_usage.description = f'批量付款改用银行付款'

            # 自动创建日记账分录
            try:
                from App_new.finance.models.journal_entry import JournalEntry
                if payment_source == 'prepayment' and available_prepayments:
                    # 使用预付款：借-成本费用 贷-预付账款
                    # 使用第一个预付账款的科目ID
                    prepayment_account_id = available_prepayments[0].prepayment_account_id
                    journal_entry = JournalEntry.create_from_eo_with_prepayment(
                        eo,
                        prepayment_account_id=prepayment_account_id,
                        user=current_user.username if current_user else None
                    )
                else:
                    # 银行付款：借-成本费用 贷-银行存款
                    journal_entry = JournalEntry.create_from_eo(
                        eo,
                        user=current_user.username if current_user else None
                    )
                if journal_entry and journal_entry.lines:
                    db.session.add(journal_entry)
                    # 自动过账
                    journal_entry.post(user=current_user.username if current_user else None)
            except Exception as je_error:
                # 日记账创建失败不影响EO付款
                import logging
                logging.getLogger(__name__).warning(f"创建批量EO付款日记账失败: {str(je_error)}")

        # 更新付款记录的预付账款使用金额和实际 EO 数量
        payment_record.prepayment_amount = prepayment_amount_used
        payment_record.eo_count = success_count

        db.session.commit()

        # 构建返回消息
        message = f'成功付款 {success_count} 个EO，付款编号：{payment_no}'
        if payment_source == 'prepayment' and available_prepayments:
            # 计算剩余总余额
            total_remaining = sum(p.balance_amount for p in available_prepayments)
            message += f'（使用预付账款，剩余总余额：{total_remaining}）'

        return jsonify({
            'success': True,
            'message': message,
            'payment_id': payment_record.id
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@project_eo.route('/<int:eo_id>/pay', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def pay_eo(eo_id):
    """付款给供应商"""
    try:
        eo = ProjectEO.query.get_or_404(eo_id)

        # 员工等级权限检查
        from App_new.business.projects.models.project import ProjectHeader
        header = ProjectHeader.query.get(eo.ref.header_id)
        if header and not can_access_project(header, current_user):
            return jsonify({'success': False, 'message': '您没有权限操作此EO'}), 403

        # 检查状态
        if eo.status == 'void':
            return jsonify({'success': False, 'message': '此EO已作废，无法付款'}), 400

        if eo.status == 'paid':
            return jsonify({'success': False, 'message': '此EO已付款'}), 400

        # 获取请求数据
        data = request.get_json()
        payment_no = data.get('payment_no')
        payment_voucher_no = data.get('payment_voucher_no', '')
        paid_date_str = data.get('paid_date')
        pay_amount = data.get('pay_amount')
        payment_remarks = data.get('payment_remarks', '')
        payment_source = data.get('payment_source', 'bank')  # bank 或 prepayment
        prepayment_id = data.get('prepayment_id')

        if not payment_no or not paid_date_str or pay_amount is None:
            return jsonify({'success': False, 'message': '请填写完整的付款信息'}), 400

        # 解析日期
        from datetime import datetime
        from decimal import Decimal
        paid_date = datetime.strptime(paid_date_str, '%Y-%m-%d').date()
        pay_amount_decimal = Decimal(str(pay_amount))

        # 处理预付账款
        from App_new.business.projects.models.supplier_prepayment import SupplierPrepayment, PrepaymentUsage

        # 检查是否有已存在的预付使用记录（EO创建时自动生成的，现在状态为confirmed）
        existing_usage = PrepaymentUsage.query.filter(
            PrepaymentUsage.eo_id == eo.id,
            PrepaymentUsage.status.in_(['pending', 'confirmed'])
        ).first()

        prepayment_usage = None
        if payment_source == 'prepayment':
            if not prepayment_id:
                return jsonify({'success': False, 'message': '请选择预付账款'}), 400

            prepayment = SupplierPrepayment.query.get(prepayment_id)

            if not prepayment:
                return jsonify({'success': False, 'message': '预付账款不存在'}), 400

            if prepayment.status not in ('confirmed', 'partial_used'):
                return jsonify({'success': False, 'message': '该预付账款状态不允许使用'}), 400

            if existing_usage and existing_usage.prepayment_id == prepayment_id:
                # 使用同一个预付账款，更新现有记录
                if existing_usage.status == 'confirmed':
                    # 已经是confirmed状态，只需更新描述
                    existing_usage.description = f'EO付款确认 {eo.eo_number}'
                    prepayment_usage = existing_usage
                else:
                    # pending状态，调整金额差额
                    amount_diff = pay_amount_decimal - existing_usage.amount
                    if amount_diff > prepayment.balance_amount:
                        return jsonify({'success': False, 'message': f'付款金额超过预付余额 {prepayment.balance_amount + existing_usage.amount}'}), 400

                    # 调整预付余额（差额）
                    prepayment.balance_amount -= amount_diff
                    prepayment.update_status()

                    # 更新使用记录为confirmed
                    existing_usage.amount = pay_amount_decimal
                    existing_usage.usage_date = paid_date
                    existing_usage.status = 'confirmed'
                    existing_usage.description = f'EO付款确认 {eo.eo_number}'
                    prepayment_usage = existing_usage
            else:
                # 使用不同的预付账款
                if pay_amount_decimal > prepayment.balance_amount:
                    return jsonify({'success': False, 'message': f'付款金额超过预付余额 {prepayment.balance_amount}'}), 400

                # 如果有已存在的使用记录，先退还（reversed）
                if existing_usage:
                    old_prepayment = SupplierPrepayment.query.get(existing_usage.prepayment_id)
                    if old_prepayment:
                        old_prepayment.balance_amount += existing_usage.amount
                        old_prepayment.update_status()
                    existing_usage.status = 'reversed'
                    existing_usage.description = f'已切换预付账款'

                # 创建新的预付款使用记录
                prepayment_usage = PrepaymentUsage(
                    prepayment_id=prepayment_id,
                    eo_id=eo.id,
                    amount=pay_amount_decimal,
                    usage_date=paid_date,
                    description=f'EO付款 {eo.eo_number}',
                    status='confirmed',
                    created_by=current_user.username if current_user else None
                )
                db.session.add(prepayment_usage)

                # 扣减预付余额
                prepayment.balance_amount -= pay_amount_decimal
                prepayment.update_status()
        else:
            # 使用银行付款，如果有预付使用记录需要退还
            if existing_usage:
                old_prepayment = SupplierPrepayment.query.get(existing_usage.prepayment_id)
                if old_prepayment:
                    old_prepayment.balance_amount += existing_usage.amount
                    old_prepayment.update_status()
                existing_usage.status = 'reversed'
                existing_usage.description = f'改用银行付款'

        # 更新付款信息
        eo.payment_no = payment_no
        eo.payment_voucher_no = payment_voucher_no if payment_voucher_no else None
        eo.paid_date = paid_date
        eo.pay_amount = pay_amount
        eo.payment_remarks = payment_remarks
        eo.status = 'paid'

        # 自动创建日记账分录
        journal_entry = None
        try:
            from App_new.finance.models.journal_entry import JournalEntry
            if payment_source == 'prepayment' and prepayment:
                # 使用预付款：借-成本费用 贷-预付账款
                journal_entry = JournalEntry.create_from_eo_with_prepayment(
                    eo,
                    prepayment_account_id=prepayment.prepayment_account_id,
                    user=current_user.username if current_user else None
                )
            else:
                # 银行付款：借-成本费用 贷-银行存款
                journal_entry = JournalEntry.create_from_eo(
                    eo,
                    user=current_user.username if current_user else None
                )
            if journal_entry and journal_entry.lines:
                db.session.add(journal_entry)
                # 自动过账
                journal_entry.post(user=current_user.username if current_user else None)
                # 关联日记账到预付款使用记录
                if prepayment_usage:
                    prepayment_usage.journal_entry_id = journal_entry.id
        except Exception as je_error:
            # 日记账创建失败不影响EO付款
            import logging
            logging.getLogger(__name__).warning(f"创建EO付款日记账失败: {str(je_error)}")

        db.session.commit()

        message = f'EO {eo.eo_number} 付款成功'
        if payment_source == 'prepayment':
            message += f'（使用预付账款）'
        if journal_entry:
            message += f'，已生成日记账 {journal_entry.entry_number}'

        return jsonify({
            'success': True,
            'message': message
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@project_eo.route('/<int:eo_id>/void', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def void_eo(eo_id):
    """作废EO"""
    try:
        eo = ProjectEO.query.get_or_404(eo_id)

        # 员工等级权限检查
        from App_new.business.projects.models.project import ProjectHeader
        header = ProjectHeader.query.get(eo.ref.header_id)
        if header and not can_access_project(header, current_user):
            return jsonify({'success': False, 'message': '您没有权限操作此EO'}), 403

        # 检查是否已经作废
        if eo.status == 'void':
            return jsonify({'success': False, 'message': '此EO已经作废'}), 400

        # 恢复预付账款余额（如果有使用，包括pending和confirmed状态）
        from App_new.business.projects.models.supplier_prepayment import SupplierPrepayment, PrepaymentUsage
        prepayment_usages = PrepaymentUsage.query.filter(
            PrepaymentUsage.eo_id == eo.id,
            PrepaymentUsage.status.in_(['confirmed', 'pending'])
        ).all()

        for usage in prepayment_usages:
            # 恢复预付账款余额
            prepayment = SupplierPrepayment.query.get(usage.prepayment_id)
            if prepayment:
                prepayment.balance_amount += usage.amount
                prepayment.update_status()

            # 标记使用记录为已冲销
            usage.status = 'reversed'
            usage.description = f'EO作废冲销 - {eo.eo_number}'

        # 更新状态为作废
        eo.status = 'void'
        db.session.commit()

        return jsonify({'success': True, 'message': f'EO {eo.eo_number} 已作废'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@project_eo.route('/<int:eo_id>/cancel_payment', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def cancel_payment(eo_id):
    """取消EO付款"""
    try:
        eo = ProjectEO.query.get_or_404(eo_id)

        # 员工等级权限检查
        from App_new.business.projects.models.project import ProjectHeader
        header = ProjectHeader.query.get(eo.ref.header_id)
        if header and not can_access_project(header, current_user):
            return jsonify({'success': False, 'message': '您没有权限操作此EO'}), 403

        # 检查是否已付款
        if eo.status != 'paid' and not eo.pay_amount:
            return jsonify({'success': False, 'message': '此EO未付款，无需取消'}), 400

        # 检查是否已作废
        if eo.status == 'void':
            return jsonify({'success': False, 'message': '此EO已作废'}), 400

        from App_new.business.projects.models.supplier_prepayment import SupplierPrepayment, PrepaymentUsage
        from decimal import Decimal

        # 注意：取消付款时，预付使用记录保持 confirmed 状态
        # 因为 EO 还存在，预付账款仍然被占用
        # 只有作废 EO 时才会返还预付账款

        # 处理日记账（如有）
        journal_reversed = False
        prepayment_usages = PrepaymentUsage.query.filter(
            PrepaymentUsage.eo_id == eo.id,
            PrepaymentUsage.status == 'confirmed'
        ).all()
        for usage in prepayment_usages:
            if usage.journal_entry_id:
                from App_new.finance.models.journal_entry import JournalEntry
                journal = JournalEntry.query.get(usage.journal_entry_id)
                if journal and journal.status == 'posted':
                    # 创建冲销分录
                    try:
                        reversal = journal.reverse(
                            reason=f'取消EO付款 {eo.eo_number}',
                            user=current_user.username if current_user else None
                        )
                        if reversal:
                            db.session.add(reversal)
                            journal_reversed = True
                    except Exception as je_err:
                        import logging
                        logging.getLogger(__name__).warning(f"冲销日记账失败: {str(je_err)}")

        # 清除付款信息
        old_payment_no = eo.payment_no
        eo.payment_no = None
        eo.payment_voucher_no = None
        eo.paid_date = None
        eo.pay_amount = None
        eo.payment_remarks = None
        eo.status = 'confirmed'

        db.session.commit()

        message = f'EO {eo.eo_number} 付款已取消'
        if prepayment_usages:
            message += '（预付账款仍被占用，作废EO时才会释放）'
        if journal_reversed:
            message += '，日记账已冲销'

        return jsonify({
            'success': True,
            'message': message
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@project_eo.route('/quick_create/<int:ref_id>', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def quick_create_eo(ref_id):
    """一键生成EO - API接口"""
    try:
        ref = ProjectRef.query.get_or_404(ref_id)
        
        # 员工等级权限检查
        from App_new.business.projects.models.project import ProjectHeader
        header = ProjectHeader.query.get(ref.header_id)
        if header and not can_access_project(header, current_user):
            return jsonify({
                'success': False,
                'message': '您没有权限访问此项目'
            }), 403
        
        # 检查REF是否已经有EO
        db.session.expire_all()  # 刷新会话，确保获取最新数据
        existing_eo = ProjectEO.query.filter_by(ref_id=ref.id).first()

        is_reactivated = False
        if existing_eo:
            if existing_eo.status == 'void':
                # 已作废的EO，重新激活
                print(f"DEBUG: REF {ref.id} 存在已作废的EO {existing_eo.eo_number}，重新激活")
                existing_eo.status = 'confirmed'
                existing_eo.payment_no = None
                existing_eo.payment_voucher_no = None
                existing_eo.paid_date = None
                existing_eo.pay_amount = None
                existing_eo.payment_remarks = None
                eo = existing_eo
                is_reactivated = True
            else:
                # 有效的EO，不能重复创建
                print(f"DEBUG: REF {ref.id} (ref_number: {ref.ref_number}) 已经存在EO {existing_eo.eo_number} (id: {existing_eo.id})")
                return jsonify({
                    'success': False,
                    'message': f'此REF已存在EO编号 {existing_eo.eo_number}，无法重复创建'
                }), 400
        else:
            print(f"DEBUG: REF {ref.id} 没有EO记录，准备创建...")

            # 创建EO
            eo = ProjectEO(
                ref_id=ref.id,
                external_system=None,
                external_status=None,
                external_reference=None,
                status='confirmed'
            )

            # 先保存获取ID
            db.session.add(eo)
            db.session.flush()  # 获取ID但不提交

            # 使用ID生成EO编号
            eo.eo_number = f'E{str(eo.id).zfill(3)}'

        # 提交事务
        db.session.commit()

        action = '重新激活' if is_reactivated else '创建'
        return jsonify({
            'success': True,
            'message': f'EO {eo.eo_number} {action}成功！',
            'eo_number': eo.eo_number,
            'eo_id': eo.id
        })
        
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        
        # 特殊处理重复键错误
        if 'Duplicate entry' in error_msg and 'unique_ref_eo' in error_msg:
            print(f"ERROR: REF {ref_id} 存在重复的EO记录")
            # 尝试查找已存在的EO
            existing_eo = ProjectEO.query.filter_by(ref_id=ref_id).first()
            if existing_eo:
                return jsonify({
                    'success': False,
                    'message': f'此REF已存在EO编号 {existing_eo.eo_number}，请勿重复创建'
                }), 400
            else:
                return jsonify({
                    'success': False,
                    'message': '此REF已存在EO记录，请刷新页面查看'
                }), 400
        
        # 其他错误
        print(f"ERROR: 快速创建EO失败 [ref_id={ref_id}]: {error_msg}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'message': f'创建失败：{error_msg}'
        }), 500


@project_eo.route('/list')
@login_required
@staff_only
def eo_list():
    """EO列表页面 - 支持筛选、搜索和分页"""
    try:
        from sqlalchemy import and_, or_, desc, asc
        from sqlalchemy.orm import aliased
        from datetime import datetime, timedelta
        from App_new.business.projects.models.ref import ProjectRef
        from App_new.business.projects.models.project import ProjectHeader, CustomerCompany

        # 创建别名：SupplierCompany 用于供应商，CustomerCompany 用于客户
        SupplierCompany = aliased(CustomerCompany, name='supplier_company')

        # 获取筛选参数
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 25, type=int)
        supplier_type = request.args.get('supplier_type', '')
        status = request.args.get('status', '')
        supplier_id = request.args.get('supplier', None, type=int)
        external_system = request.args.get('external_system', '')
        date_range = request.args.get('date_range', '')
        min_amount = request.args.get('min_amount', None, type=float)
        max_amount = request.args.get('max_amount', None, type=float)
        keyword = request.args.get('keyword', '')
        sort_by = request.args.get('sort_by', 'created_at')
        sort_order = request.args.get('sort_order', 'desc')
        has_invoice = request.args.get('has_invoice', '')  # 筛选有无发票

        # 构建查询
        from App_new.shared.models.business_types import BusinessType
        query = db.session.query(
            ProjectEO,
            SupplierCompany.company_name.label('supplier_name'),
            BusinessType.name.label('ref_type_name'),
            ProjectRef.description.label('ref_description'),
            ProjectRef.detailed_description.label('ref_detailed_description'),
            ProjectRef.ref_number.label('ref_number'),
            ProjectRef.header_id.label('project_id'),
            ProjectHeader.hid.label('project_hid'),
            ProjectHeader.desc.label('project_name'),
            CustomerCompany.company_name.label('company_name')
        ).join(
            ProjectRef, ProjectEO.ref_id == ProjectRef.id, isouter=True
        ).join(
            BusinessType, ProjectRef.ref_type_id == BusinessType.id, isouter=True
        ).join(
            SupplierCompany, ProjectRef.supplier_id == SupplierCompany.id, isouter=True
        ).join(
            ProjectHeader, ProjectRef.header_id == ProjectHeader.id, isouter=True
        ).join(
            CustomerCompany, ProjectHeader.company_id == CustomerCompany.id, isouter=True
        )
        
        # 应用筛选条件
        filters = []
        
        # 根据员工等级过滤EO
        if current_user.role and current_user.role.name == 'staff':
            # 检查用户资料中的员工等级
            staff_level = 1  # 默认等级
            if current_user.profile:
                staff_level = current_user.profile.staff_level or 1
            
            if staff_level == 1:
                # 1级员工只能看到自己创建的项目的EO
                filters.append(ProjectHeader.staff_name == current_user.username)
            # 2级员工可以看到所有EO，不需要额外过滤
        
        if status:
            filters.append(ProjectEO.status == status)
        
        if supplier_id:
            filters.append(ProjectRef.supplier_id == supplier_id)
        
        if supplier_type:
            # 通过REF的ref_type_id关联到business_types来筛选
            # 根据供应商类型枚举值匹配业务类型名称
            supplier_type_map = {
                'visa': '签证',
                'flight': '机票',
                'hotel': '酒店',
                'transport': '用车',
                'local_operator': '地接',
                'other': '其他'
            }
            type_name = supplier_type_map.get(supplier_type)
            if type_name:
                filters.append(BusinessType.name == type_name)
        
        if external_system:
            filters.append(ProjectEO.external_system.ilike(f'%{external_system}%'))
        
        if date_range:
            today = datetime.now().date()
            if date_range == 'today':
                start_date = today
                end_date = today + timedelta(days=1)
            elif date_range == 'week':
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=7)
            elif date_range == 'month':
                start_date = today.replace(day=1)
                if today.month == 12:
                    end_date = today.replace(year=today.year + 1, month=1, day=1)
                else:
                    end_date = today.replace(month=today.month + 1, day=1)
            elif date_range == 'quarter':
                quarter = (today.month - 1) // 3
                start_date = today.replace(month=quarter * 3 + 1, day=1)
                if quarter == 3:
                    end_date = today.replace(year=today.year + 1, month=1, day=1)
                else:
                    end_date = today.replace(month=quarter * 3 + 4, day=1)
            elif date_range == 'year':
                start_date = today.replace(month=1, day=1)
                end_date = today.replace(year=today.year + 1, month=1, day=1)
            
            filters.append(and_(
                ProjectEO.created_at >= start_date,
                ProjectEO.created_at < end_date
            ))
        
        if min_amount is not None and min_amount > 0:
            filters.append(ProjectRef.cost_price >= float(min_amount))
        
        if max_amount is not None and max_amount > 0:
            filters.append(ProjectRef.cost_price <= float(max_amount))
        
        if keyword:
            keyword_filter = or_(
                ProjectEO.eo_number.ilike(f'%{keyword}%'),
                ProjectEO.external_system.ilike(f'%{keyword}%'),
                ProjectEO.external_reference.ilike(f'%{keyword}%'),
                ProjectRef.description.ilike(f'%{keyword}%'),
                ProjectRef.detailed_description.ilike(f'%{keyword}%'),
                ProjectRef.ref_number.ilike(f'%{keyword}%'),
                ProjectHeader.desc.ilike(f'%{keyword}%'),
                CustomerCompany.company_name.ilike(f'%{keyword}%')
            )
            filters.append(keyword_filter)

        # 筛选有无发票
        if has_invoice:
            from App_new.business.projects.models.invoice import InvoiceItem
            # 获取有发票的 ref_id 列表
            refs_with_invoice = db.session.query(InvoiceItem.ref_id).filter(
                InvoiceItem.ref_id.isnot(None)
            ).distinct().subquery()

            if has_invoice == 'yes':
                filters.append(ProjectEO.ref_id.in_(db.session.query(refs_with_invoice.c.ref_id)))
            elif has_invoice == 'no':
                filters.append(~ProjectEO.ref_id.in_(db.session.query(refs_with_invoice.c.ref_id)))
        
        # 应用筛选条件
        if filters:
            query = query.filter(and_(*filters))
        
        # 排序
        if sort_by == 'created_at':
            order_column = ProjectEO.created_at
        elif sort_by == 'amount':
            order_column = ProjectRef.cost_price
        elif sort_by == 'status':
            order_column = ProjectEO.status
        else:
            order_column = ProjectEO.created_at
        
        if sort_order == 'asc':
            query = query.order_by(asc(order_column))
        else:
            query = query.order_by(desc(order_column))
        
        # 分页
        try:
            pagination = query.paginate(
                page=page, 
                per_page=per_page, 
                error_out=False
            )
        except AttributeError:
            # 如果新版本方法不存在，使用旧版本
            pagination = query.paginate(
                page=page, 
                per_page=per_page, 
                error_out=False
            )
        
        # 确保分页对象有必要的属性
        if not hasattr(pagination, 'has_prev'):
            pagination.has_prev = pagination.page > 1
            pagination.has_next = pagination.page < pagination.pages
            pagination.prev_num = pagination.page - 1 if pagination.page > 1 else None
            pagination.next_num = pagination.page + 1 if pagination.page < pagination.pages else None
        
        # 确保分页对象有iter_pages方法
        if not hasattr(pagination, 'iter_pages'):
            def iter_pages(left_edge=2, left_current=2, right_current=5, right_edge=2):
                last = 0
                for num in range(1, pagination.pages + 1):
                    if num <= left_edge or \
                       (num > pagination.page - left_current - 1 and \
                        num < pagination.page + right_current) or \
                       num > pagination.pages - right_edge:
                        if last + 1 != num:
                            yield None
                        yield num
                        last = num
            pagination.iter_pages = iter_pages
        
        # 处理EO数据，添加显示属性
        from App_new.business.projects.models.invoice import InvoiceItem, ProjectInvoice
        eos = []
        for eo, supplier_name, ref_type_name, ref_description, ref_detailed_description, ref_number, project_id, project_hid, project_name, company_name in pagination.items:
            # 获取乘客/出行人姓名
            pax_names = ''
            if eo.ref:
                # 1. 先尝试从机票乘客表获取
                from App_new.business.flight.models.flight import ProjectFlightPassenger
                passengers = ProjectFlightPassenger.query.filter_by(ref_id=eo.ref_id).all()
                if passengers:
                    pax_names = ', '.join([p.name for p in passengers if p.name])
                # 2. 如果没有乘客记录，从项目成员表获取
                if not pax_names and eo.ref.header:
                    from App_new.business.projects.models.project_member import ProjectMember
                    members = ProjectMember.query.filter_by(header_id=eo.ref.header.id).all()
                    if members:
                        # 优先显示 leader，否则显示所有成员
                        leaders = [m.member_name for m in members if m.is_leader and m.member_name]
                        if leaders:
                            pax_names = ', '.join(leaders)
                        else:
                            pax_names = ', '.join([m.member_name for m in members if m.member_name])

            # 获取关联的发票信息
            invoice_info = None
            if eo.ref_id:
                invoice_item = InvoiceItem.query.filter_by(ref_id=eo.ref_id).first()
                if invoice_item and invoice_item.invoice:
                    inv = invoice_item.invoice
                    invoice_info = {
                        'id': inv.id,
                        'invoice_number': inv.invoice_number,
                        'status': inv.status
                    }
            
            # 获取匹配状态和匹配详情
            matched_transactions = eo.bank_transactions if hasattr(eo, 'bank_transactions') else []
            is_matched = len(matched_transactions) > 0
            matched_tx_count = len(matched_transactions)
            # 获取匹配的银行流水信息（日期、金额、对方）
            matched_tx_info = []
            for tx in matched_transactions[:3]:  # 最多显示3条
                matched_tx_info.append({
                    'date': tx.transaction_date.strftime('%m-%d') if tx.transaction_date else '',
                    'amount': float(tx.amount) if tx.amount else 0,
                    'counterparty': tx.counterparty_name or tx.description[:20] if tx.description else ''
                })

            eo_dict = {
                'id': eo.id,
                'eo_number': str(eo.eo_number) if eo.eo_number else '',
                'ref_id': int(eo.ref_id) if eo.ref_id else None,
                'ref_description': str(ref_description) if ref_description else '',
                'ref_detailed_description': str(ref_detailed_description) if ref_detailed_description else '',
                'ref_number': str(ref_number) if ref_number else '',
                'supplier_name': str(supplier_name) if supplier_name else '',
                'ref_type_name': str(ref_type_name) if ref_type_name else '',
                'project_id': int(project_id) if project_id else None,
                'project_hid': str(project_hid) if project_hid else '',
                'project_name': str(project_name) if project_name else '',
                'company_name': str(company_name) if company_name else '',
                'pax_names': pax_names,
                # EO新字段
                'conf_code': str(eo.conf_code) if eo.conf_code else '',
                'discount': float(eo.discount) if eo.discount else 0,
                'tax': float(eo.tax) if eo.tax else 0,
                'payment_no': str(eo.payment_no) if eo.payment_no else '',
                'paid_date': eo.paid_date,
                'pay_amount': float(eo.pay_amount) if eo.pay_amount else None,
                # REF价格信息
                'selling_price': float(eo.ref.selling_price) if eo.ref and eo.ref.selling_price else None,
                'cost_price': float(eo.ref.cost_price) if eo.ref and eo.ref.cost_price else 0,
                'currency': str(eo.ref.currency) if eo.ref and eo.ref.currency else 'SGD',
                # 外部系统
                'external_system': str(eo.external_system) if eo.external_system else '',
                'external_status': str(eo.external_status) if eo.external_status else '',
                'external_reference': str(eo.external_reference) if eo.external_reference else '',
                # 状态
                'status': str(eo.status) if eo.status else 'draft',
                'status_display': get_status_display(eo.status),
                'status_color': get_status_color(eo.status),
                'created_at': eo.created_at,
                'updated_at': eo.updated_at,
                # 发票信息
                'invoice': invoice_info,
                # 匹配状态
                'is_reconciled': eo.is_reconciled,
                'is_matched': is_matched,
                'matched_tx_count': matched_tx_count,
                'matched_tx_info': matched_tx_info
            }
            eos.append(eo_dict)
        
        # 获取筛选选项数据（供应商列表带类型代码，用于前端JavaScript筛选）
        from sqlalchemy import func
        suppliers_with_type = db.session.query(
            CustomerCompany,
            BusinessType.code.label('supplier_type_code')
        ).outerjoin(
            BusinessType, CustomerCompany.supplier_type_id == BusinessType.id
        ).filter(
            CustomerCompany.is_supplier == True
        ).order_by(func.lower(CustomerCompany.company_name)).all()

        suppliers = [{
            'id': s.CustomerCompany.id,
            'company_name': s.CustomerCompany.company_name,
            'supplier_type_code': s.supplier_type_code or ''
        } for s in suppliers_with_type]

        # 获取供应商类型列表
        supplier_types = BusinessType.query.filter_by(is_active=True).order_by(BusinessType.sort_order).all()
        
        # 计算筛选结果数量
        filtered_count = pagination.total if any([supplier_type, status, supplier_id, external_system, date_range, min_amount, max_amount, keyword, has_invoice]) else None

        return render_template('business/projects/project_eo/eo_list.html',
                             eos=eos,
                             pagination=pagination,
                             suppliers=suppliers,
                             supplier_types=supplier_types,
                             filtered_count=filtered_count,
                             current_filters={
                                 'supplier_type': supplier_type,
                                 'status': status,
                                 'supplier': supplier_id,
                                 'external_system': external_system,
                                 'date_range': date_range,
                                 'min_amount': min_amount,
                                 'max_amount': max_amount,
                                 'keyword': keyword,
                                 'sort_by': sort_by,
                                 'sort_order': sort_order,
                                 'has_invoice': has_invoice
                             })
                             
    except Exception as e:
        import traceback
        print(f"EO列表加载失败: {str(e)}")
        print(f"错误详情: {traceback.format_exc()}")
        flash(f'EO列表加载失败：{str(e)}', 'error')
        return redirect(url_for('business_projects.list.list_projects'))


def get_status_display(status):
    """获取状态的中文显示名称"""
    if not status or not isinstance(status, str):
        return '未知'
    
    status_map = {
        'draft': '草稿',
        'confirmed': '已确认',
        'paid': '已付款',
        'cancelled': '已取消',
        'void': '已作废'
    }
    return status_map.get(status, status)


def get_status_color(status):
    """获取状态对应的Bootstrap颜色类"""
    if not status or not isinstance(status, str):
        return 'secondary'
    
    color_map = {
        'draft': 'secondary',
        'confirmed': 'info',
        'paid': 'success',
        'cancelled': 'danger',
        'void': 'dark'
    }
    return color_map.get(status, 'secondary')


def get_supplier_type_display(supplier_type):
    """获取供应商类型的中文显示名称"""
    if not supplier_type or not isinstance(supplier_type, str):
        return '未知'
    
    type_map = {
        'visa': '签证',
        'flight': '机票',
        'hotel': '酒店',
        'transport': '交通',
        'local_operator': '地接社',
        'other': '其他'
    }
    return type_map.get(supplier_type, supplier_type)


def get_supplier_type_color(supplier_type):
    """获取供应商类型对应的Bootstrap颜色类"""
    if not supplier_type or not isinstance(supplier_type, str):
        return 'secondary'

    color_map = {
        'visa': 'warning',
        'flight': 'primary',
        'hotel': 'success',
        'transport': 'info',
        'local_operator': 'dark',
        'other': 'secondary'
    }
    return color_map.get(supplier_type, 'secondary')


@project_eo.route('/export')
@login_required
@staff_only
def eo_export():
    """导出EO列表到Excel - 导出所有符合筛选条件的数据"""
    try:
        import io
        import pandas as pd
        from flask import make_response
        from sqlalchemy import and_, or_, desc, asc
        from sqlalchemy.orm import aliased
        from datetime import datetime, timedelta
        from App_new.business.projects.models.ref import ProjectRef
        from App_new.business.projects.models.project import ProjectHeader, CustomerCompany
        from App_new.shared.models.business_types import BusinessType
        from App_new.business.projects.models.invoice import InvoiceItem

        # 创建别名
        SupplierCompany = aliased(CustomerCompany, name='supplier_company')

        # 获取筛选参数（与 eo_list 相同）
        supplier_type = request.args.get('supplier_type', '')
        status = request.args.get('status', '')
        supplier_id = request.args.get('supplier', None, type=int)
        external_system = request.args.get('external_system', '')
        date_range = request.args.get('date_range', '')
        min_amount = request.args.get('min_amount', None, type=float)
        max_amount = request.args.get('max_amount', None, type=float)
        keyword = request.args.get('keyword', '')
        has_invoice = request.args.get('has_invoice', '')

        # 构建查询
        query = db.session.query(
            ProjectEO,
            SupplierCompany.company_name.label('supplier_name'),
            BusinessType.name.label('ref_type_name'),
            ProjectRef.ref_number.label('ref_number'),
            ProjectHeader.hid.label('project_hid'),
            CustomerCompany.company_name.label('company_name')
        ).join(
            ProjectRef, ProjectEO.ref_id == ProjectRef.id, isouter=True
        ).join(
            BusinessType, ProjectRef.ref_type_id == BusinessType.id, isouter=True
        ).join(
            SupplierCompany, ProjectRef.supplier_id == SupplierCompany.id, isouter=True
        ).join(
            ProjectHeader, ProjectRef.header_id == ProjectHeader.id, isouter=True
        ).join(
            CustomerCompany, ProjectHeader.company_id == CustomerCompany.id, isouter=True
        )

        # 应用筛选条件
        filters = []

        # 根据员工等级过滤EO
        if current_user.role and current_user.role.name == 'staff':
            staff_level = 1
            if current_user.profile:
                staff_level = current_user.profile.staff_level or 1
            if staff_level == 1:
                filters.append(ProjectHeader.staff_name == current_user.username)

        if status:
            filters.append(ProjectEO.status == status)

        if supplier_id:
            filters.append(ProjectRef.supplier_id == supplier_id)

        if supplier_type:
            supplier_type_map = {
                'visa': '签证',
                'flight': '机票',
                'hotel': '酒店',
                'transport': '用车',
                'local_operator': '地接',
                'other': '其他'
            }
            type_name = supplier_type_map.get(supplier_type)
            if type_name:
                filters.append(BusinessType.name == type_name)

        if external_system:
            filters.append(ProjectEO.external_system.ilike(f'%{external_system}%'))

        if date_range:
            today = datetime.now().date()
            if date_range == 'today':
                start_date = today
                end_date = today + timedelta(days=1)
            elif date_range == 'week':
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=7)
            elif date_range == 'month':
                start_date = today.replace(day=1)
                if today.month == 12:
                    end_date = today.replace(year=today.year + 1, month=1, day=1)
                else:
                    end_date = today.replace(month=today.month + 1, day=1)
            elif date_range == 'quarter':
                quarter = (today.month - 1) // 3
                start_date = today.replace(month=quarter * 3 + 1, day=1)
                if quarter == 3:
                    end_date = today.replace(year=today.year + 1, month=1, day=1)
                else:
                    end_date = today.replace(month=quarter * 3 + 4, day=1)
            elif date_range == 'year':
                start_date = today.replace(month=1, day=1)
                end_date = today.replace(year=today.year + 1, month=1, day=1)

            filters.append(and_(
                ProjectEO.created_at >= start_date,
                ProjectEO.created_at < end_date
            ))

        if min_amount is not None and min_amount > 0:
            filters.append(ProjectRef.cost_price >= float(min_amount))

        if max_amount is not None and max_amount > 0:
            filters.append(ProjectRef.cost_price <= float(max_amount))

        if keyword:
            keyword_filter = or_(
                ProjectEO.eo_number.ilike(f'%{keyword}%'),
                ProjectEO.external_system.ilike(f'%{keyword}%'),
                ProjectEO.external_reference.ilike(f'%{keyword}%'),
                ProjectRef.description.ilike(f'%{keyword}%'),
                ProjectRef.ref_number.ilike(f'%{keyword}%'),
                ProjectHeader.desc.ilike(f'%{keyword}%'),
                CustomerCompany.company_name.ilike(f'%{keyword}%')
            )
            filters.append(keyword_filter)

        if has_invoice:
            refs_with_invoice = db.session.query(InvoiceItem.ref_id).filter(
                InvoiceItem.ref_id.isnot(None)
            ).distinct().subquery()

            if has_invoice == 'yes':
                filters.append(ProjectEO.ref_id.in_(db.session.query(refs_with_invoice.c.ref_id)))
            elif has_invoice == 'no':
                filters.append(~ProjectEO.ref_id.in_(db.session.query(refs_with_invoice.c.ref_id)))

        # 应用筛选条件
        if filters:
            query = query.filter(and_(*filters))

        # 排序
        query = query.order_by(desc(ProjectEO.created_at))

        # 获取所有数据（不分页）
        results = query.all()

        # 构建导出数据
        export_data = []
        for eo, supplier_name, ref_type_name, ref_number, project_hid, company_name in results:
            # 获取乘客姓名
            pax_names = ''
            if eo.ref:
                from App_new.business.flight.models.flight import ProjectFlightPassenger
                passengers = ProjectFlightPassenger.query.filter_by(ref_id=eo.ref_id).all()
                if passengers:
                    pax_names = ', '.join([p.name for p in passengers if p.name])
                if not pax_names and eo.ref.header:
                    from App_new.business.projects.models.project_member import ProjectMember
                    members = ProjectMember.query.filter_by(header_id=eo.ref.header.id).all()
                    if members:
                        leaders = [m.member_name for m in members if m.is_leader and m.member_name]
                        if leaders:
                            pax_names = ', '.join(leaders)
                        else:
                            pax_names = ', '.join([m.member_name for m in members if m.member_name])

            # 获取发票号
            invoice_number = ''
            if eo.ref_id:
                invoice_item = InvoiceItem.query.filter_by(ref_id=eo.ref_id).first()
                if invoice_item and invoice_item.invoice:
                    invoice_number = invoice_item.invoice.invoice_number

            export_data.append({
                'EO No': eo.eo_number or '',
                'EO Date': eo.created_at.strftime('%Y-%m-%d') if eo.created_at else '',
                'HID': project_hid or '',
                'REF': ref_number or '',
                'Invoice': invoice_number,
                'Type': ref_type_name or '',
                'Supplier': supplier_name or '',
                'Company': company_name or '',
                'Pax': pax_names,
                'Curr': eo.ref.currency if eo.ref and eo.ref.currency else 'SGD',
                'Cost': float(eo.ref.cost_price) if eo.ref and eo.ref.cost_price else 0,
                'Tax': float(eo.tax) if eo.tax else 0,
                'Disc': float(eo.discount) if eo.discount else 0,
                'Payment No': eo.payment_no or '',
                'Paid Date': eo.paid_date.strftime('%Y-%m-%d') if eo.paid_date else '',
                'Pay Amt': float(eo.pay_amount) if eo.pay_amount else '',
                'Status': get_status_display(eo.status)
            })

        # 创建 DataFrame
        df = pd.DataFrame(export_data)

        # 生成 Excel 文件
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='EO List', index=False)

            # 设置列宽
            worksheet = writer.sheets['EO List']
            column_widths = {
                'A': 12, 'B': 12, 'C': 10, 'D': 10, 'E': 15, 'F': 10,
                'G': 20, 'H': 20, 'I': 25, 'J': 8, 'K': 12, 'L': 10,
                'M': 10, 'N': 15, 'O': 12, 'P': 12, 'Q': 10
            }
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

        excel_buffer.seek(0)

        # 生成文件名
        filename = f"EO_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = make_response(excel_buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        import traceback
        current_app.logger.error(f"EO导出失败: {str(e)}")
        current_app.logger.error(f"错误详情: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500
