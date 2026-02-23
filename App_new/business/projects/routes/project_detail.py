# -*- coding: utf-8 -*-
"""
项目详情路由
包含项目详情显示、REF记录、财务统计等功能
"""

from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify
from flask_login import login_required, current_user
from ..services.project_service import ProjectService
from ..services.project_stats import ProjectStatsService
from App_new.utils.decorators import staff_only
from App_new.utils.permissions import can_access_project

from App_new.business.projects.models.ref import ProjectRef
from App_new.exts import db, csrf
from datetime import datetime
import traceback

# 创建蓝图
bp = Blueprint('detail', __name__)

@bp.route('/test')
def test_detail():
    """测试路由 - 验证基本功能"""
    return "项目详情路由工作正常！"

@bp.route('/<int:project_id>')
@login_required
@staff_only
def project_detail(project_id):
    """项目详情页面"""
    try:
        print(f"DEBUG: 访问项目详情页面，project_id: {project_id}")
        
        from App_new.business.projects.models.project import ProjectHeader
        from App_new.business.projects.models.project import CustomerCompany
        
        # 使用project_id查询ProjectHeader
        header = ProjectHeader.query.get_or_404(project_id)
        print(f"DEBUG: 找到项目 HID: {header.hid}, 描述: {header.desc}")
        
        # 员工等级权限检查
        if not can_access_project(header, current_user):
            flash('您没有权限访问此项目', 'error')
            return redirect(url_for('business_projects.list.list_projects'))
        
        # 手动加载相关的REF数据
        refs = ProjectRef.query.filter_by(header_id=project_id).all()
        header.refs = refs
        print(f"DEBUG: 加载了 {len(refs)} 个REF记录")

        # 获取上一个和下一个项目（优化查询）
        prev_header = ProjectHeader.query.filter(
            ProjectHeader.id < project_id
        ).order_by(ProjectHeader.id.desc()).limit(1).first()

        next_header = ProjectHeader.query.filter(
            ProjectHeader.id > project_id
        ).order_by(ProjectHeader.id.asc()).limit(1).first()

        # 获取公司信息（通过backref自动关联）
        company = header.company
        print(f"DEBUG: 公司信息: {company.company_name if company else 'None'}")

        # 获取所有活跃的公司列表供选择
        companies = CustomerCompany.query.filter_by(status='active').order_by(CustomerCompany.company_name).all()
        print(f"DEBUG: 加载了 {len(companies)} 个活跃公司")

        # 获取项目的 leader 成员名称
        from App_new.business.projects.models.project_member import ProjectMember
        leader_member = ProjectMember.query.filter_by(
            header_id=project_id,
            is_leader=True
        ).first()
        if not leader_member:
            # 没有指定 leader，取第一个成员
            leader_member = ProjectMember.query.filter_by(header_id=project_id).first()

        leader_member_name = ''
        if leader_member:
            leader_member_name = leader_member.member_name
            if leader_member.title:
                leader_member_name = f"{leader_member.title} {leader_member.member_name}"

        # 获取供应商预付余额（用于在REF列表中显示）
        from App_new.business.projects.models.supplier_prepayment import SupplierPrepayment, PrepaymentUsage
        from App_new.business.projects.models.eo import ProjectEO
        from sqlalchemy import func
        supplier_balances = {}

        # 查询所有有余额的供应商预付记录
        prepayment_balances = db.session.query(
            SupplierPrepayment.supplier_id,
            func.sum(SupplierPrepayment.balance_amount).label('total_balance'),
            SupplierPrepayment.currency
        ).filter(
            SupplierPrepayment.status.in_(['confirmed', 'partial_used']),
            SupplierPrepayment.balance_amount > 0
        ).group_by(
            SupplierPrepayment.supplier_id,
            SupplierPrepayment.currency
        ).all()

        for supplier_id, total_balance, currency in prepayment_balances:
            if supplier_id:
                supplier_balances[supplier_id] = {
                    'balance': float(total_balance),
                    'currency': currency,
                    'pending_eo': 0,
                    'expected_balance': float(total_balance)
                }

        # 获取已经从预付款扣减过的EO ID列表
        used_eo_subquery = db.session.query(PrepaymentUsage.eo_id).filter(
            PrepaymentUsage.eo_id.isnot(None)
        ).distinct().subquery()

        # 查询每个供应商的待付款EO总额（排除已从预付款扣减的）
        pending_eo_amounts = db.session.query(
            ProjectRef.supplier_id,
            func.sum(ProjectRef.cost_price).label('pending_total')
        ).join(
            ProjectEO, ProjectEO.ref_id == ProjectRef.id
        ).filter(
            ProjectRef.supplier_id.isnot(None),
            ProjectEO.status == 'confirmed',  # 待付款状态
            ~ProjectEO.id.in_(used_eo_subquery)  # 排除已扣减的
        ).group_by(
            ProjectRef.supplier_id
        ).all()

        # 更新预计余额
        for supplier_id, pending_total in pending_eo_amounts:
            if supplier_id in supplier_balances:
                pending_amount = float(pending_total) if pending_total else 0
                supplier_balances[supplier_id]['pending_eo'] = pending_amount
                supplier_balances[supplier_id]['expected_balance'] = supplier_balances[supplier_id]['balance'] - pending_amount

        print(f"DEBUG: 准备渲染模板 business/projects/project_detail.html")
        return render_template('business/projects/project_detail.html',
                               header=header,
                               company=company,
                               companies=companies,
                               prev_header=prev_header,
                               next_header=next_header,
                               leader_member_name=leader_member_name,
                               supplier_balances=supplier_balances)
                               
    except Exception as e:
        error_msg = str(e)
        print(f"ERROR: 加载项目详情失败")
        print(f"ERROR: 错误类型: {type(e).__name__}")
        print(f"ERROR: 错误信息: {error_msg}")
        import traceback
        error_trace = traceback.format_exc()
        print(f"ERROR: 完整堆栈:\n{error_trace}")
        
        # 记录到日志文件
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"项目详情页面加载失败 [project_id={project_id}]: {error_msg}")
        logger.error(f"完整堆栈: {error_trace}")
        
        flash(f'加载项目详情失败：{error_msg}', 'error')
        return redirect(url_for('business_projects.list.list_projects'))

@bp.route('/<int:project_id>/refs')
@login_required
@staff_only
def project_refs(project_id):
    """项目REF记录列表"""
    try:
        project_service = ProjectService()
        
        # 获取项目信息
        project = project_service.get_project_by_id(project_id)
        if not project:
            return jsonify({'success': False, 'error': '项目不存在'}), 404
        
        # 获取REF记录
        refs = project_service.get_project_refs(project_id)
        
        return jsonify({
            'success': True,
            'data': [ref.to_dict() for ref in refs]
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@bp.route('/<int:project_id>/stats')
@login_required
@staff_only
def project_stats(project_id):
    """项目统计信息"""
    try:
        stats_service = ProjectStatsService()
        
        # 获取项目统计
        stats = stats_service.get_project_stats(project_id)
        
        return jsonify({
            'success': True,
            'data': stats
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/header/<int:header_id>/reminder', methods=['POST', 'PUT'])
@login_required
@staff_only
def manage_reminder(header_id):
    """添加或更新项目提醒"""
    try:
        from App_new.business.projects.models.project import ProjectHeader
        
        # 获取项目
        header = ProjectHeader.query.get_or_404(header_id)
        
        # 获取请求数据
        data = request.get_json()
        reminder_event = data.get('reminder_event', '').strip()
        reminder_date_str = data.get('reminder_date', '')
        
        # 验证数据
        if not reminder_event:
            return jsonify({
                'success': False,
                'message': '提醒事件不能为空'
            }), 400
        
        if not reminder_date_str:
            return jsonify({
                'success': False,
                'message': '提醒日期不能为空'
            }), 400
        
        # 解析日期
        try:
            reminder_date = datetime.strptime(reminder_date_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({
                'success': False,
                'message': '提醒日期格式不正确'
            }), 400
        
        # 检查日期不能早于今天
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        if reminder_date < today:
            return jsonify({
                'success': False,
                'message': '提醒日期不能早于今天'
            }), 400
        
        # 更新提醒信息
        header.reminder_event = reminder_event
        header.reminder_date = reminder_date
        header.reminder_sent = False  # 重置发送状态
        
        db.session.commit()
        
        # 如果有提醒信息，同步到待办事项
        if header.reminder_event and header.reminder_date:
            try:
                from App_new.utils.reminder_utils import create_reminder_todo
                create_reminder_todo(header)
            except Exception as e:
                print(f"DEBUG: Failed to create reminder todo: {str(e)}")
        
        action = '添加' if request.method == 'POST' else '更新'
        return jsonify({
            'success': True,
            'message': f'提醒{action}成功'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"DEBUG: Error managing reminder: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'操作失败：{str(e)}'
        }), 500


@bp.route('/header/<int:header_id>/reminder', methods=['DELETE'])
@login_required
@staff_only
def delete_reminder(header_id):
    """删除项目提醒"""
    try:
        from App_new.business.projects.models.project import ProjectHeader
        
        # 获取项目
        header = ProjectHeader.query.get_or_404(header_id)
        
        # 清除提醒信息
        header.reminder_event = None
        header.reminder_date = None
        header.reminder_sent = False
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '提醒删除成功'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"DEBUG: Error deleting reminder: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'删除失败：{str(e)}'
        }), 500

@bp.route('/<int:project_id>/receipts')
@login_required
@staff_only
def project_receipts(project_id):
    """项目收款记录列表"""
    try:
        project_service = ProjectService()
        
        # 获取项目信息
        project = project_service.get_project_by_id(project_id)
        if not project:
            return jsonify({'success': False, 'error': '项目不存在'}), 404
        
        # 获取收款记录
        receipts = project_service.get_project_receipts(project_id)
        
        return jsonify({
            'success': True,
            'data': [receipt.to_dict() for receipt in receipts]
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@bp.route('/<int:project_id>/payments')
@login_required
@staff_only
def project_payments(project_id):
    """项目付款记录列表"""
    try:
        project_service = ProjectService()
        
        # 获取项目信息
        project = project_service.get_project_by_id(project_id)
        if not project:
            return jsonify({'success': False, 'error': '项目不存在'}), 404
        
        # 获取付款记录
        payments = project_service.get_project_payments(project_id)
        
        return jsonify({
            'success': True,
            'data': [payment.to_dict() for payment in payments]
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@bp.route('/<int:project_id>/documents')
@login_required
@staff_only
def project_documents(project_id):
    """项目文档列表"""
    try:
        project_service = ProjectService()
        
        # 获取项目信息
        project = project_service.get_project_by_id(project_id)
        if not project:
            return jsonify({'success': False, 'error': '项目不存在'}), 404
        
        # 获取文档记录
        documents = project_service.get_project_documents(project_id)
        
        return jsonify({
            'success': True,
            'data': [doc.to_dict() for doc in documents]
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== 打印REF价格单 ====================

@bp.route('/<int:project_id>/print-statement')
@login_required
@staff_only
def print_statement(project_id):
    """打印REF价格单（Price Statement）"""
    try:
        from App_new.business.projects.models.project import ProjectHeader
        from App_new.business.projects.models.project_member import ProjectMember
        import json

        header = ProjectHeader.query.get_or_404(project_id)

        # 权限检查
        if not can_access_project(header, current_user):
            flash('您没有权限访问此项目', 'error')
            return redirect(url_for('business_projects.list.list_projects'))

        # 加载REF数据
        refs = ProjectRef.query.filter_by(header_id=project_id).all()

        # 为每个REF构建extra_data（复用invoice的逻辑）
        def build_ref_extra_data(ref):
            extra_data = {}
            if ref.extra_info:
                try:
                    extra_data = json.loads(ref.extra_info)
                except (json.JSONDecodeError, TypeError):
                    extra_data = {}

            # 机票类型：从flight_passengers表获取乘客信息
            if hasattr(ref, 'flight_passengers') and ref.flight_passengers:
                pax_names = [p.name for p in ref.flight_passengers if p.name]
                extra_data['pax_names_display'] = ', '.join(pax_names)
                if hasattr(ref, 'flight_segments') and ref.flight_segments:
                    seg = ref.flight_segments[0]
                    if seg.departure_time:
                        extra_data['departure_date'] = seg.departure_time.strftime('%Y-%m-%d')
            elif not extra_data.get('pax_names_display'):
                # 非机票类型：从项目成员表获取
                pax_names_ids = extra_data.get('pax_names', [])
                if pax_names_ids:
                    members = ProjectMember.query.filter(ProjectMember.id.in_(pax_names_ids)).all()
                    pax_names_list = [f"{m.title} {m.member_name}" if m.title else m.member_name for m in members]
                    extra_data['pax_names_display'] = ', '.join(pax_names_list)
                else:
                    extra_data['pax_names_display'] = extra_data.get('pax_name', '')

            return extra_data

        for ref in refs:
            ref.extra_data = build_ref_extra_data(ref)

        # 获取项目leader
        leader_member = ProjectMember.query.filter_by(
            header_id=project_id, is_leader=True
        ).first()
        if not leader_member:
            leader_member = ProjectMember.query.filter_by(header_id=project_id).first()

        leader_name = ''
        if leader_member:
            leader_name = leader_member.member_name
            if leader_member.title:
                leader_name = f"{leader_member.title} {leader_member.member_name}"

        # 客户公司信息
        customer_company_display = None
        customer_company_address = None
        if header.company:
            company_name = header.company.company_name
            if company_name and company_name.lower() in ['个人', 'cash', '现金']:
                customer_company_display = header.contact if header.contact else leader_name
            else:
                customer_company_display = company_name
            customer_company_address = header.company.address

        # 计算总金额
        total_amount = sum(float(ref.selling_price or 0) for ref in refs)

        return render_template('business/projects/ref_statement_print.html',
                               header=header,
                               refs=refs,
                               leader_name=leader_name,
                               customer_company_display=customer_company_display,
                               customer_company_address=customer_company_address,
                               project_contact=header.contact,
                               total_amount=total_amount,
                               consultant=current_user.username if current_user.is_authenticated else '-',
                               print_date=datetime.now())

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'生成价格单失败：{str(e)}', 'error')
        return redirect(url_for('business_projects.detail.project_detail', project_id=project_id))


# ==================== 导出REF Excel ====================

@bp.route('/<int:project_id>/export-refs')
@login_required
@staff_only
def export_refs(project_id):
    """导出项目REF明细为Excel"""
    try:
        from App_new.business.projects.models.project import ProjectHeader
        from io import BytesIO
        from flask import send_file
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        header = ProjectHeader.query.get_or_404(project_id)

        # 权限检查
        if not can_access_project(header, current_user):
            flash('您没有权限导出此项目', 'error')
            return redirect(url_for('business_projects.list.list_projects'))

        # 加载REF数据
        refs = ProjectRef.query.filter_by(header_id=project_id).all()

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "REF List"

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

        # 项目信息行
        ws.merge_cells('A1:K1')
        title_cell = ws.cell(row=1, column=1, value=f"Project: {header.hid} - {header.desc or ''}")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells('A2:K2')
        info_cell = ws.cell(row=2, column=1, value=f"Company: {header.company.company_name if header.company else '-'}    Currency: {header.currency or 'SGD'}    Contact: {header.contact or '-'}")
        info_cell.font = Font(size=10, color="666666")
        ws.row_dimensions[2].height = 20

        # 表头（第3行）
        headers_list = ['REF No', 'Type', 'Description', 'Invoice No', 'EO No',
                        'Selling Price', 'Cost Price', 'Profit', 'Unpaid Amount',
                        'Supplier', 'Currency']
        for col, h in enumerate(headers_list, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 数据行
        total_selling = 0
        total_cost = 0
        total_profit = 0
        total_unpaid = 0

        for row_idx, ref in enumerate(refs, 4):
            selling = float(ref.selling_price or 0)
            cost = float(ref.cost_price or 0)
            profit = float(ref.ref_profit) if hasattr(ref, 'ref_profit') else selling - cost
            unpaid = float(ref.unpaid_amount) if hasattr(ref, 'unpaid_amount') else selling

            total_selling += selling
            total_cost += cost
            total_profit += profit
            total_unpaid += unpaid

            # Invoice编号
            invoice_str = ''
            if hasattr(ref, 'get_invoices') and callable(ref.get_invoices):
                invoices = ref.get_invoices()
                if invoices:
                    invoice_str = ', '.join(inv.invoice_number for inv in invoices)

            # EO编号
            eo_str = ''
            if ref.eos and hasattr(ref.eos, 'eo_number') and ref.eos.status != 'void':
                eo_str = ref.eos.eo_number

            row_data = [
                ref.ref_number,
                ref.ref_type.name if ref.ref_type else '',
                ref.description or ref.detailed_description or '',
                invoice_str,
                eo_str,
                selling,
                cost,
                profit,
                unpaid,
                ref.supplier.company_name if ref.supplier else '',
                ref.currency or header.currency or 'SGD'
            ]

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                # 数字列右对齐并格式化
                if col in (6, 7, 8, 9):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'

        # 总计行
        total_row = 4 + len(refs)
        ws.cell(row=total_row, column=5, value='TOTAL').font = total_font
        ws.cell(row=total_row, column=5).fill = total_fill
        ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=total_row, column=5).border = thin_border

        for col, val in [(6, total_selling), (7, total_cost), (8, total_profit), (9, total_unpaid)]:
            cell = ws.cell(row=total_row, column=col, value=val)
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = '#,##0.00'
            cell.border = thin_border

        # 空单元格也加样式
        for col in [1, 2, 3, 4, 10, 11]:
            cell = ws.cell(row=total_row, column=col)
            cell.fill = total_fill
            cell.border = thin_border

        # 调整列宽
        column_widths = [15, 10, 30, 15, 15, 14, 14, 14, 14, 25, 8]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        hid = header.hid or f'P{project_id}'
        filename = f"{hid}_REF_List_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'导出失败：{str(e)}', 'error')
        return redirect(url_for('business_projects.detail.project_detail', project_id=project_id))


# ==================== 邮件功能 ====================

@bp.route('/<int:project_id>/email/templates')
@login_required
@staff_only
def get_email_templates(project_id):
    """获取可用的邮件模板列表"""
    try:
        from App_new.business.projects.models.project import EmailTemplate
        templates = EmailTemplate.query.filter_by(is_active=True).order_by(EmailTemplate.category, EmailTemplate.name).all()
        return jsonify({
            'success': True,
            'templates': [t.to_dict() for t in templates]
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/<int:project_id>/email/contacts')
@login_required
@staff_only
def get_email_contacts(project_id):
    """获取项目关联公司的联系人列表"""
    try:
        from App_new.business.projects.models.project import ProjectHeader, CompanyContact
        
        header = ProjectHeader.query.get_or_404(project_id)
        contacts = []
        
        # 如果有关联公司，获取公司联系人
        if header.company_id:
            company_contacts = CompanyContact.query.filter_by(company_id=header.company_id).all()
            for c in company_contacts:
                if c.email:  # 只返回有邮箱的联系人
                    contacts.append({
                        'id': c.id,
                        'name': c.name,
                        'email': c.email,
                        'position': c.position,
                        'is_primary': c.is_primary
                    })
        
        # 添加项目头部的联系人信息（如果有邮箱）
        if header.contact:
            # 尝试从公司信息获取邮箱
            if header.company and header.company.contact_email:
                contacts.insert(0, {
                    'id': 0,
                    'name': header.contact,
                    'email': header.company.contact_email,
                    'position': '项目联系人',
                    'is_primary': True
                })
        
        return jsonify({
            'success': True,
            'contacts': contacts,
            'company_name': header.company.company_name if header.company else None
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/<int:project_id>/email/send', methods=['POST'])
@login_required
@staff_only
def send_project_email(project_id):
    """发送项目邮件"""
    try:
        from flask import current_app
        from flask_mail import Mail, Message
        from App_new.business.projects.models.project import ProjectHeader, ProjectEmail, EmailTemplate
        from werkzeug.utils import secure_filename
        import json
        import os
        import tempfile
        
        header = ProjectHeader.query.get_or_404(project_id)
        
        # 处理FormData（支持文件上传）
        if request.content_type and 'multipart/form-data' in request.content_type:
            recipients = json.loads(request.form.get('recipients', '[]'))
            cc = json.loads(request.form.get('cc', '[]'))
            subject = request.form.get('subject', '')
            body = request.form.get('body', '')
            template_id = request.form.get('template_id')
            attachment_count = int(request.form.get('attachment_count', 0))
        else:
            # 兼容JSON格式
            data = request.get_json()
            recipients = data.get('recipients', [])
            cc = data.get('cc', [])
            subject = data.get('subject', '')
            body = data.get('body', '')
            template_id = data.get('template_id')
            attachment_count = 0
        
        # 验证必填字段
        if not recipients:
            return jsonify({'success': False, 'message': '请选择收件人'}), 400
        if not subject:
            return jsonify({'success': False, 'message': '请输入邮件主题'}), 400
        if not body:
            return jsonify({'success': False, 'message': '请输入邮件内容'}), 400
        
        # 处理附件
        attachments_info = []
        temp_files = []  # 保存临时文件路径，发送后删除
        
        if attachment_count > 0:
            upload_folder = os.path.join(current_app.static_folder, 'uploads', 'email_attachments')
            os.makedirs(upload_folder, exist_ok=True)
            
            for i in range(attachment_count):
                file_key = f'attachment_{i}'
                if file_key in request.files:
                    file = request.files[file_key]
                    if file and file.filename:
                        # 保存临时文件
                        filename = secure_filename(file.filename)
                        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S_')
                        unique_filename = timestamp + filename
                        file_path = os.path.join(upload_folder, unique_filename)
                        file.save(file_path)
                        temp_files.append(file_path)
                        
                        # 获取文件MIME类型
                        import mimetypes
                        mime_type, _ = mimetypes.guess_type(filename)
                        if not mime_type:
                            mime_type = 'application/octet-stream'
                        
                        attachments_info.append({
                            'name': filename,
                            'path': file_path,
                            'type': mime_type
                        })
        
        # 创建邮件记录
        email_record = ProjectEmail(
            header_id=project_id,
            template_id=int(template_id) if template_id else None,
            subject=subject,
            body=body,
            recipients=json.dumps(recipients),
            cc=json.dumps(cc),
            attachments=json.dumps(attachments_info),
            status='draft',
            created_by=current_user.username if current_user.is_authenticated else 'system'
        )
        db.session.add(email_record)
        
        # 发送邮件
        try:
            # 检查邮件配置
            mail_server = current_app.config.get('MAIL_SERVER')
            mail_username = current_app.config.get('MAIL_USERNAME')
            mail_password = current_app.config.get('MAIL_PASSWORD')
            mail_port = current_app.config.get('MAIL_PORT')
            mail_use_ssl = current_app.config.get('MAIL_USE_SSL', False)
            mail_use_tls = current_app.config.get('MAIL_USE_TLS', False)
            
            # 记录配置信息（用于调试，不记录密码）
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"邮件发送配置检查 - 服务器: {mail_server}, 端口: {mail_port}, SSL: {mail_use_ssl}, TLS: {mail_use_tls}, 用户名: {mail_username}")
            
            if not mail_server:
                raise ValueError('邮件服务器(MAIL_SERVER)未配置，请检查环境变量或配置文件')
            if not mail_username:
                raise ValueError('邮件用户名(MAIL_USERNAME)未配置，请检查环境变量或配置文件')
            if not mail_password:
                raise ValueError('邮件密码(MAIL_PASSWORD)未配置，请检查环境变量或配置文件')
            
            mail = Mail(current_app)
            sender_email = current_app.config.get('MAIL_DEFAULT_SENDER') or mail_username
            
            # 处理邮件正文：将换行符转换为HTML格式
            # 检查是否包含HTML标签（简单判断）
            has_html_tags = '<' in body and '>' in body and any(tag in body.lower() for tag in ['<br', '<p', '<div', '<span', '<h'])
            
            if not has_html_tags:
                # 纯文本格式，将换行符转换为<br>，并包装在HTML中
                # 先转义HTML特殊字符
                import html
                escaped_body = html.escape(body)
                html_body = escaped_body.replace('\n', '<br>')
                html_body = f'<div style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">{html_body}</div>'
            else:
                # 已经是HTML格式，但可能还有未转换的换行符，补充转换
                # 只转换不在HTML标签内的换行符（简单处理）
                html_body = body.replace('\n', '<br>')
            
            msg = Message(
                subject=subject,
                sender=sender_email,
                recipients=recipients,
                cc=cc if cc else None,
                html=html_body
            )
            
            # 添加附件
            for att_info in attachments_info:
                with open(att_info['path'], 'rb') as f:
                    msg.attach(
                        att_info['name'],
                        att_info['type'],
                        f.read()
                    )
            
            logger.info(f"准备发送邮件 - 主题: {subject}, 收件人: {recipients}, 抄送: {cc}")
            mail.send(msg)
            logger.info("邮件发送成功")
            
            # 更新发送状态
            email_record.status = 'sent'
            email_record.sent_at = datetime.utcnow()
            db.session.commit()
            
            # 清理临时文件
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except Exception as e:
                    print(f"清理临时文件失败: {temp_file}, 错误: {e}")
            
            return jsonify({
                'success': True,
                'message': f'邮件发送成功，已发送至 {len(recipients)} 位收件人' + (f'，包含 {len(attachments_info)} 个附件' if attachments_info else ''),
                'email_id': email_record.id
            })
            
        except Exception as mail_error:
            import traceback
            import logging
            logger = logging.getLogger(__name__)
            
            # 记录详细错误信息
            error_detail = str(mail_error)
            error_traceback = traceback.format_exc()
            logger.error(f"邮件发送失败 - 错误: {error_detail}")
            logger.error(f"错误堆栈: {error_traceback}")
            
            # 检查常见错误原因并提供友好提示
            error_message = f'邮件发送失败：{error_detail}'
            
            # 常见错误提示
            if 'timeout' in error_detail.lower() or 'timed out' in error_detail.lower():
                error_message += '。可能原因：网络连接超时，请检查服务器是否能访问SMTP服务器'
            elif 'connection' in error_detail.lower() or 'refused' in error_detail.lower():
                error_message += '。可能原因：无法连接到SMTP服务器，请检查网络连接和防火墙设置'
            elif 'authentication' in error_detail.lower() or 'login' in error_detail.lower() or '535' in error_detail:
                error_message += '。可能原因：邮箱账号或密码错误，请检查MAIL_USERNAME和MAIL_PASSWORD配置'
            elif 'ssl' in error_detail.lower() or 'certificate' in error_detail.lower():
                error_message += '。可能原因：SSL/TLS证书验证失败，请检查MAIL_USE_SSL和MAIL_USE_TLS配置'
            elif '550' in error_detail or '553' in error_detail:
                error_message += '。可能原因：邮件被拒绝，请检查发件人地址和收件人地址是否正确'
            
            email_record.status = 'failed'
            email_record.error_message = error_detail
            db.session.commit()
            
            # 清理临时文件
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except:
                    pass
            
            return jsonify({'success': False, 'message': error_message}), 500
        
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/<int:project_id>/email/history')
@login_required
@staff_only
def get_email_history(project_id):
    """获取项目邮件发送历史"""
    try:
        from App_new.business.projects.models.project import ProjectEmail
        
        emails = ProjectEmail.query.filter_by(header_id=project_id).order_by(ProjectEmail.created_at.desc()).all()
        return jsonify({
            'success': True,
            'emails': [e.to_dict() for e in emails]
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== 邮件模板管理 ====================

@bp.route('/email/templates')
@login_required
@staff_only
def email_templates_list():
    """邮件模板列表页面"""
    from App_new.business.projects.models.project import EmailTemplate
    
    templates = EmailTemplate.query.order_by(EmailTemplate.category, EmailTemplate.name).all()
    categories = db.session.query(EmailTemplate.category).distinct().filter(EmailTemplate.category.isnot(None)).all()
    categories = [c[0] for c in categories]
    
    return render_template('business/projects/email_templates/list.html', 
                         templates=templates, 
                         categories=categories)


@bp.route('/email/templates/create', methods=['GET', 'POST'])
@login_required
@staff_only
def create_email_template():
    """创建邮件模板"""
    from App_new.business.projects.models.project import EmailTemplate
    
    if request.method == 'POST':
        try:
            template = EmailTemplate(
                name=request.form.get('name'),
                subject=request.form.get('subject'),
                body=request.form.get('body'),
                category=request.form.get('category'),
                is_active=request.form.get('is_active') == '1',
                created_by=current_user.username if current_user.is_authenticated else 'system'
            )
            db.session.add(template)
            db.session.commit()
            flash('模板创建成功！', 'success')
            return redirect(url_for('business_projects.detail.email_templates_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'创建失败：{str(e)}', 'error')
    
    categories = ['flight', 'hotel', 'visa', 'invoice', 'general']
    return render_template('business/projects/email_templates/form.html', 
                         template=None, 
                         categories=categories)


@bp.route('/email/templates/<int:template_id>/edit', methods=['GET', 'POST'])
@login_required
@staff_only
def edit_email_template(template_id):
    """编辑邮件模板"""
    from App_new.business.projects.models.project import EmailTemplate
    
    template = EmailTemplate.query.get_or_404(template_id)
    
    if request.method == 'POST':
        try:
            template.name = request.form.get('name')
            template.subject = request.form.get('subject')
            template.body = request.form.get('body')
            template.category = request.form.get('category')
            template.is_active = request.form.get('is_active') == '1'
            db.session.commit()
            pass
            return redirect(url_for('business_projects.detail.email_templates_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败：{str(e)}', 'error')
    
    categories = ['flight', 'hotel', 'visa', 'invoice', 'general']
    return render_template('business/projects/email_templates/form.html', 
                         template=template, 
                         categories=categories)


@bp.route('/email/templates/<int:template_id>/delete', methods=['POST'])
@login_required
@staff_only
def delete_email_template(template_id):
    """删除邮件模板"""
    from App_new.business.projects.models.project import EmailTemplate

    template = EmailTemplate.query.get_or_404(template_id)
    try:
        db.session.delete(template)
        db.session.commit()
        flash('模板删除成功！', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'error')

    return redirect(url_for('business_projects.detail.email_templates_list'))


@bp.route('/email/templates/<int:template_id>/copy', methods=['POST'])
@login_required
@staff_only
def copy_email_template(template_id):
    """复制邮件模板"""
    from App_new.business.projects.models.project import EmailTemplate

    template = EmailTemplate.query.get_or_404(template_id)
    try:
        new_template = EmailTemplate(
            name=template.name + ' (副本)',
            subject=template.subject,
            body=template.body,
            category=template.category,
            is_active=template.is_active,
            created_by=current_user.username if current_user.is_authenticated else 'system'
        )
        db.session.add(new_template)
        db.session.commit()
        pass
    except Exception as e:
        db.session.rollback()
        flash(f'复制失败：{str(e)}', 'error')

    return redirect(url_for('business_projects.detail.email_templates_list'))


# ==================== 复制项目功能 ====================

@bp.route('/<int:project_id>/copy', methods=['POST'])
@login_required
@staff_only
@csrf.exempt
def copy_project(project_id):
    """复制项目 - 复制项目主表、REF记录、人员信息、机票航段和乘客"""
    try:
        from App_new.business.projects.models.project import ProjectHeader
        from App_new.business.projects.models.project_member import ProjectMember
        from App_new.business.flight.models.flight import ProjectFlightSegment, ProjectFlightPassenger

        # 获取原项目
        original = ProjectHeader.query.get_or_404(project_id)

        # 生成新的HID
        new_hid = ProjectHeader.generate_hid()

        # 复制项目主表（日期使用当前时间）
        from datetime import datetime
        now = datetime.utcnow()  # 使用UTC时间，与模型默认值保持一致
        print(f"[DEBUG] 复制项目 - 当前时间(UTC): {now}")

        new_header = ProjectHeader(
            hid=new_hid,
            desc=original.desc,
            company_id=original.company_id,
            limit=original.limit,
            contact=original.contact,
            dept=original.dept,
            staff_id=original.staff_id,
            currency=original.currency,
            leader_name=original.leader_name,
            type=original.type,
            status='active',  # 新项目默认为进行中
            remarks=f'复制自 {original.hid}',
            # 利润分配相关
            order_type=original.order_type,
            operator_ids=original.operator_ids,
            operator_names=original.operator_names,
            salesperson_ids=original.salesperson_ids,
            salesperson_names=original.salesperson_names,
            # 显式设置日期为当前时间
            created_at=now,
            updated_at=now,
        )
        db.session.add(new_header)
        db.session.flush()  # 获取新项目ID

        # 复制REF记录及关联的机票数据
        ref_count = 0
        segment_count = 0
        passenger_count = 0

        for ref in original.refs:
            new_ref = ProjectRef(
                header_id=new_header.id,
                ref_number=ProjectRef.generate_ref_number(),
                description=ref.description,
                detailed_description=ref.detailed_description,
                ref_type_id=ref.ref_type_id,
                supplier_id=ref.supplier_id,
                selling_price=ref.selling_price,
                cost_price=ref.cost_price,
                currency=ref.currency,
                remarks=ref.remarks,
                extra_info=ref.extra_info,
                status='confirmed',
                payment_status='unpaid',
                # 显式设置日期为当前时间
                created_at=now,
                updated_at=now,
            )
            db.session.add(new_ref)
            db.session.flush()  # 获取新REF ID
            ref_count += 1

            # 复制机票航段
            segments = ProjectFlightSegment.query.filter_by(ref_id=ref.id).all()
            for seg in segments:
                new_segment = ProjectFlightSegment(
                    ref_id=new_ref.id,
                    flight_number=seg.flight_number,
                    departure_airport=seg.departure_airport,
                    arrival_airport=seg.arrival_airport,
                    departure_time=seg.departure_time,
                    arrival_time=seg.arrival_time,
                    cabin_class=seg.cabin_class,
                    cabin_code=seg.cabin_code,
                    status=seg.status,
                    # 显式设置日期为当前时间
                    created_at=now,
                    updated_at=now,
                )
                db.session.add(new_segment)
                segment_count += 1

            # 复制机票乘客
            passengers = ProjectFlightPassenger.query.filter_by(ref_id=ref.id).all()
            for pax in passengers:
                new_passenger = ProjectFlightPassenger(
                    ref_id=new_ref.id,
                    name=pax.name,
                    passenger_type=pax.passenger_type,
                    selling_price=pax.selling_price,
                    cost_price=pax.cost_price,
                    ticket_number=None,  # 不复制票号
                    pnr=None,  # 不复制PNR
                    # 显式设置日期为当前时间
                    created_at=now,
                    updated_at=now,
                )
                db.session.add(new_passenger)
                passenger_count += 1

        # 复制项目人员
        member_count = 0
        for member in original.members:
            new_member = ProjectMember(
                header_id=new_header.id,
                title=member.title,
                member_name=member.member_name,
                member_name_en=member.member_name_en,
                member_role=member.member_role,
                member_phone=member.member_phone,
                member_email=member.member_email,
                id_type=member.id_type,
                id_number=member.id_number,
                is_leader=member.is_leader,
                remarks=member.remarks,
                # 显式设置日期为当前时间
                created_at=now,
                updated_at=now,
            )
            db.session.add(new_member)
            member_count += 1

        db.session.commit()

        msg = f'项目复制成功！新项目编号: {new_hid}，包含 {ref_count} 个REF，{member_count} 个人员'
        if segment_count > 0:
            msg += f'，{segment_count} 个航段，{passenger_count} 个乘客'

        return jsonify({
            'success': True,
            'message': msg,
            'new_project_id': new_header.id,
            'new_hid': new_hid
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'复制失败：{str(e)}'
        }), 500