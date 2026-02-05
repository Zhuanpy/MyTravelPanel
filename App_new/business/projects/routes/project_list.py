# -*- coding: utf-8 -*-
"""
项目列表路由
包含项目列表显示、筛选、分页等功能
"""

from flask import Blueprint, render_template, request, flash, jsonify, send_file, redirect
from flask_login import login_required, current_user
from App_new.exts import db, csrf
from App_new.business.projects.models.project import ProjectHeader, CustomerCompany
from App_new.business.projects.models.ref import ProjectRef
from App_new.business.projects.models.receipt import ProjectReceipt, ReceiptInvoiceAllocation
from ..services.project_service import ProjectService
from ..services.project_stats import ProjectStatsService
from App_new.utils.decorators import staff_only
from datetime import datetime
import traceback
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter


# 创建蓝图
bp = Blueprint('list', __name__)


@bp.route('/api/update_company', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def api_update_company():
    try:
        project_id = request.json.get('project_id')
        company_id = request.json.get('company_id') or None

        project = ProjectHeader.query.get_or_404(project_id)
        if company_id:
            company = CustomerCompany.query.get(company_id)
            if not company:
                return jsonify({'success': False, 'message': '客户公司不存在'}), 400
            project.company_id = company.id
        else:
            project.company_id = None

        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/update_company_name', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def api_update_company_name():
    try:
        project_id = request.json.get('project_id')
        company_name = request.json.get('company_name', '').strip()

        project = ProjectHeader.query.get_or_404(project_id)
        
        if company_name:
            # 查找或创建公司
            company = CustomerCompany.query.filter_by(company_name=company_name).first()
            if not company:
                # 创建新公司
                company = CustomerCompany(
                    company_name=company_name,
                    company_code=company_name[:10],  # 使用前10个字符作为代码
                    status='active'
                )
                db.session.add(company)
                db.session.flush()  # 获取新创建的ID
            
            project.company_id = company.id
        else:
            project.company_id = None

        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/search_companies')
@login_required
@staff_only
def api_search_companies():
    try:
        query = request.args.get('q', '').strip()
        if not query:
            return jsonify({'success': True, 'companies': []})
        
        # 搜索匹配的公司名称
        companies = CustomerCompany.query.filter(
            CustomerCompany.company_name.like(f'%{query}%')
        ).limit(10).all()
        
        return jsonify({
            'success': True,
            'companies': [{'id': c.id, 'company_name': c.company_name} for c in companies]
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/')
@login_required
@staff_only
def list_projects():
    """项目列表页面 - 增强版筛选和统计"""
    try:
        # 获取筛选参数
        status = request.args.get('status', '')
        search = request.args.get('search', '')
        company = request.args.get('company', '')
        leader = request.args.get('leader', '')
        contact = request.args.get('contact', '')
        staff_name = request.args.get('staff_name', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        selling_min = request.args.get('selling_min', '')
        selling_max = request.args.get('selling_max', '')
        profit_min = request.args.get('profit_min', '')
        profit_max = request.args.get('profit_max', '')
        profit_status = request.args.get('profit_status', '')  # 盈亏状态：profit/loss/zero
        balance_min = request.args.get('balance_min', '')
        balance_max = request.args.get('balance_max', '')
        payment_status = request.args.get('payment_status', '')
        settlement_status = request.args.get('settlement_status', '')  # 结算状态：settled/unsettled
        sort_by = request.args.get('sort_by', 'hid_desc')

        # 使用服务层处理业务逻辑
        project_service = ProjectService()
        stats_service = ProjectStatsService()
        
        # 构建查询
        base_query = ProjectHeader.query
        
        # 只预加载必要的关联，避免过度加载
        base_query = base_query.options(
            db.joinedload(ProjectHeader.company)
        )
        
        # 根据员工等级过滤项目
        if current_user.role and current_user.role.name == 'staff':
            # 检查用户资料中的员工等级
            staff_level = 1  # 默认等级
            if current_user.profile:
                staff_level = current_user.profile.staff_level or 1
            
            if staff_level == 1:
                # 1级员工只能看到自己创建的订单
                # 支持 staff_id 或 staff_name（用户名/全名）匹配
                user_full_name = current_user.profile.get_full_name() if current_user.profile else None
                if user_full_name and user_full_name != "未设置姓名":
                    # 匹配 staff_id 或 staff_name（用户名或全名）
                    base_query = base_query.filter(
                        db.or_(
                            ProjectHeader.staff_id == current_user.id,
                            ProjectHeader.staff_name == current_user.username,
                            ProjectHeader.staff_name == user_full_name
                        )
                    )
                else:
                    # 匹配 staff_id 或 staff_name（用户名）
                    base_query = base_query.filter(
                        db.or_(
                            ProjectHeader.staff_id == current_user.id,
                            ProjectHeader.staff_name == current_user.username
                        )
                    )
            # 2级员工可以看到所有订单，不需要额外过滤
        
        # 应用筛选条件
        if status:
            base_query = base_query.filter(ProjectHeader.status == status)

        if search:
            search_lower = search.lower()
            # 子查询：查找包含搜索关键词的项目成员所属的项目ID
            from ..models.project_member import ProjectMember
            member_subquery = db.session.query(ProjectMember.header_id).filter(
                db.or_(
                    ProjectMember.member_name.ilike(f'%{search_lower}%'),
                    ProjectMember.member_name_en.ilike(f'%{search_lower}%')
                )
            ).subquery()

            base_query = base_query.filter(
                db.or_(
                    ProjectHeader.hid.ilike(f'%{search_lower}%'),
                    ProjectHeader.desc.ilike(f'%{search_lower}%'),
                    ProjectHeader.contact.ilike(f'%{search_lower}%'),
                    ProjectHeader.leader_name.ilike(f'%{search_lower}%'),
                    ProjectHeader.id.in_(member_subquery)  # 搜索乘客姓名
                )
            )

        if company:
            base_query = base_query.join(CustomerCompany).filter(
                CustomerCompany.company_name.like(f'%{company}%')
            )
        
        if leader:
            base_query = base_query.filter(
                ProjectHeader.leader_name.like(f'%{leader}%')
            )
        
        if contact:
            base_query = base_query.filter(
                ProjectHeader.contact.like(f'%{contact}%')
            )
        
        if staff_name:
            base_query = base_query.filter(
                ProjectHeader.staff_name.like(f'%{staff_name}%')
            )
        
        # 添加项目类型筛选
        project_type = request.args.get('type', '')
        if project_type:
            base_query = base_query.filter(ProjectHeader.type == project_type)

        # 添加结算状态筛选
        if settlement_status:
            if settlement_status == 'settled':
                base_query = base_query.filter(ProjectHeader.is_settled == True)
            elif settlement_status == 'unsettled':
                base_query = base_query.filter(ProjectHeader.is_settled == False)
            elif settlement_status == 'can_settle':
                # 可结算：未结算 且 满足结算条件
                from App_new.business.projects.models.eo import ProjectEO
                from App_new.business.projects.models.invoice import InvoiceItem

                # 1. 必须未结算
                base_query = base_query.filter(ProjectHeader.is_settled == False)

                # 2. 必须有REF且所有REF都有发票
                # 子查询：每个项目的REF数量
                ref_count_subq = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(ProjectRef.id).label('ref_count')
                ).group_by(ProjectRef.header_id).subquery()

                # 子查询：每个项目有发票的REF数量
                invoiced_ref_subq = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(db.distinct(InvoiceItem.ref_id)).label('invoiced_count')
                ).join(InvoiceItem, InvoiceItem.ref_id == ProjectRef.id
                ).group_by(ProjectRef.header_id).subquery()

                # 3. 所有EO已付款（或没有EO）
                # 子查询：每个项目的EO数量
                eo_count_subq = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(ProjectEO.id).label('eo_count')
                ).join(ProjectEO, ProjectEO.ref_id == ProjectRef.id
                ).group_by(ProjectRef.header_id).subquery()

                # 子查询：每个项目已付款的EO数量（is_paid == True 视为已付款）
                paid_eo_subq = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(ProjectEO.id).label('paid_count')
                ).join(ProjectEO, ProjectEO.ref_id == ProjectRef.id
                ).filter(ProjectEO.is_paid == True
                ).group_by(ProjectRef.header_id).subquery()

                # 4. Balance = 0
                # 子查询：每个项目的销售总额
                selling_subq = db.session.query(
                    ProjectRef.header_id,
                    db.func.coalesce(db.func.sum(ProjectRef.selling_price), 0).label('total_selling')
                ).group_by(ProjectRef.header_id).subquery()

                # 子查询：每个项目的收款总额（基于发票分配计算，正确处理跨项目收款）
                from App_new.business.projects.models.invoice import ProjectInvoice as PI
                from sqlalchemy import union_all

                # 方式1: 通过发票分配表统计（仅限项目级别收款，避免与REF级别收款重复计算）
                invoice_alloc_q = db.session.query(
                    PI.header_id.label('header_id'),
                    ReceiptInvoiceAllocation.allocated_amount.label('amount')
                ).join(
                    ReceiptInvoiceAllocation, ReceiptInvoiceAllocation.invoice_id == PI.id
                ).join(
                    ProjectReceipt, ReceiptInvoiceAllocation.receipt_id == ProjectReceipt.id
                ).filter(
                    ProjectReceipt.status == 'confirmed',
                    ProjectReceipt.ref_id == None  # 仅项目级别收款，REF级别收款在方式2统计
                )

                # 方式2: REF级别直接收款
                ref_receipt_q = db.session.query(
                    ProjectReceipt.header_id.label('header_id'),
                    ProjectReceipt.amount.label('amount')
                ).filter(
                    ProjectReceipt.status == 'confirmed',
                    ProjectReceipt.ref_id.isnot(None)
                )

                # 合并两种方式
                combined = union_all(invoice_alloc_q, ref_receipt_q).alias('combined_receipts')
                receipt_subq = db.session.query(
                    combined.c.header_id,
                    db.func.coalesce(db.func.sum(combined.c.amount), 0).label('total_received')
                ).group_by(combined.c.header_id).subquery()

                # 组合查询：找出满足所有条件的项目
                can_settle_subq = db.session.query(ref_count_subq.c.header_id).outerjoin(
                    invoiced_ref_subq, invoiced_ref_subq.c.header_id == ref_count_subq.c.header_id
                ).outerjoin(
                    eo_count_subq, eo_count_subq.c.header_id == ref_count_subq.c.header_id
                ).outerjoin(
                    paid_eo_subq, paid_eo_subq.c.header_id == ref_count_subq.c.header_id
                ).outerjoin(
                    selling_subq, selling_subq.c.header_id == ref_count_subq.c.header_id
                ).outerjoin(
                    receipt_subq, receipt_subq.c.header_id == ref_count_subq.c.header_id
                ).filter(
                    # 条件1：有REF且所有REF都有发票
                    ref_count_subq.c.ref_count > 0,
                    ref_count_subq.c.ref_count == db.func.coalesce(invoiced_ref_subq.c.invoiced_count, 0),
                    # 条件2：所有REF都生成了EO
                    ref_count_subq.c.ref_count == db.func.coalesce(eo_count_subq.c.eo_count, 0),
                    # 条件3：所有EO已付款
                    db.func.coalesce(eo_count_subq.c.eo_count, 0) > 0,
                    db.func.coalesce(eo_count_subq.c.eo_count, 0) == db.func.coalesce(paid_eo_subq.c.paid_count, 0),
                    # 条件4：Balance = 0
                    db.func.abs(
                        db.func.coalesce(selling_subq.c.total_selling, 0) -
                        db.func.coalesce(receipt_subq.c.total_received, 0)
                    ) < 0.01
                )

                base_query = base_query.filter(ProjectHeader.id.in_(can_settle_subq))

            elif settlement_status == 'incomplete':
                # 未完成：未结算 且 不满足可结算条件
                from App_new.business.projects.models.eo import ProjectEO
                from App_new.business.projects.models.invoice import InvoiceItem

                # 1. 必须未结算
                base_query = base_query.filter(ProjectHeader.is_settled == False)

                # 构建可结算项目的子查询（与 can_settle 相同逻辑）
                ref_count_subq = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(ProjectRef.id).label('ref_count')
                ).group_by(ProjectRef.header_id).subquery()

                invoiced_ref_subq = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(db.distinct(InvoiceItem.ref_id)).label('invoiced_count')
                ).join(InvoiceItem, InvoiceItem.ref_id == ProjectRef.id
                ).group_by(ProjectRef.header_id).subquery()

                eo_count_subq = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(ProjectEO.id).label('eo_count')
                ).join(ProjectEO, ProjectEO.ref_id == ProjectRef.id
                ).group_by(ProjectRef.header_id).subquery()

                paid_eo_subq = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(ProjectEO.id).label('paid_count')
                ).join(ProjectEO, ProjectEO.ref_id == ProjectRef.id
                ).filter(ProjectEO.is_paid == True
                ).group_by(ProjectRef.header_id).subquery()

                selling_subq = db.session.query(
                    ProjectRef.header_id,
                    db.func.coalesce(db.func.sum(ProjectRef.selling_price), 0).label('total_selling')
                ).group_by(ProjectRef.header_id).subquery()

                from App_new.business.projects.models.invoice import ProjectInvoice as PI
                from sqlalchemy import union_all

                # 方式1: 通过发票分配表统计（仅限项目级别收款，避免与REF级别收款重复计算）
                invoice_alloc_q = db.session.query(
                    PI.header_id.label('header_id'),
                    ReceiptInvoiceAllocation.allocated_amount.label('amount')
                ).join(
                    ReceiptInvoiceAllocation, ReceiptInvoiceAllocation.invoice_id == PI.id
                ).join(
                    ProjectReceipt, ReceiptInvoiceAllocation.receipt_id == ProjectReceipt.id
                ).filter(
                    ProjectReceipt.status == 'confirmed',
                    ProjectReceipt.ref_id == None
                )

                # 方式2: REF级别直接收款
                ref_receipt_q = db.session.query(
                    ProjectReceipt.header_id.label('header_id'),
                    ProjectReceipt.amount.label('amount')
                ).filter(
                    ProjectReceipt.status == 'confirmed',
                    ProjectReceipt.ref_id.isnot(None)
                )

                combined = union_all(invoice_alloc_q, ref_receipt_q).alias('combined_receipts')
                receipt_subq = db.session.query(
                    combined.c.header_id,
                    db.func.coalesce(db.func.sum(combined.c.amount), 0).label('total_received')
                ).group_by(combined.c.header_id).subquery()

                # 可结算项目ID列表
                can_settle_ids = db.session.query(ref_count_subq.c.header_id).outerjoin(
                    invoiced_ref_subq, invoiced_ref_subq.c.header_id == ref_count_subq.c.header_id
                ).outerjoin(
                    eo_count_subq, eo_count_subq.c.header_id == ref_count_subq.c.header_id
                ).outerjoin(
                    paid_eo_subq, paid_eo_subq.c.header_id == ref_count_subq.c.header_id
                ).outerjoin(
                    selling_subq, selling_subq.c.header_id == ref_count_subq.c.header_id
                ).outerjoin(
                    receipt_subq, receipt_subq.c.header_id == ref_count_subq.c.header_id
                ).filter(
                    # 条件1：有REF且所有REF都有发票
                    ref_count_subq.c.ref_count > 0,
                    ref_count_subq.c.ref_count == db.func.coalesce(invoiced_ref_subq.c.invoiced_count, 0),
                    # 条件2：所有REF都生成了EO
                    ref_count_subq.c.ref_count == db.func.coalesce(eo_count_subq.c.eo_count, 0),
                    # 条件3：所有EO已付款
                    db.func.coalesce(eo_count_subq.c.eo_count, 0) > 0,
                    db.func.coalesce(eo_count_subq.c.eo_count, 0) == db.func.coalesce(paid_eo_subq.c.paid_count, 0),
                    # 条件4：Balance = 0
                    db.func.abs(
                        db.func.coalesce(selling_subq.c.total_selling, 0) -
                        db.func.coalesce(receipt_subq.c.total_received, 0)
                    ) < 0.01
                )

                # 未完成 = 未结算 且 不在可结算列表中
                base_query = base_query.filter(~ProjectHeader.id.in_(can_settle_ids))

        # 添加付款状态筛选 - 基于发票数据判断
        if payment_status:
            from App_new.business.projects.models.invoice import ProjectInvoice
            from sqlalchemy import union_all

            # 按 header 统计发票金额和已付金额
            invoice_sum_subq = db.session.query(
                ProjectInvoice.header_id.label('hid'),
                db.func.coalesce(db.func.sum(ProjectInvoice.amount), 0).label('invoice_total'),
                db.func.coalesce(db.func.sum(ProjectInvoice.paid_amount), 0).label('invoice_paid')
            ).filter(
                ProjectInvoice.status != 'cancelled'
            ).group_by(ProjectInvoice.header_id).subquery()

            # 按 header 统计总销售金额（用于没有发票的项目）
            refs_sum_subq = db.session.query(
                ProjectRef.header_id.label('hid'),
                db.func.coalesce(db.func.sum(ProjectRef.selling_price), 0).label('total_selling')
            ).group_by(ProjectRef.header_id).subquery()

            # 按 header 统计总收款金额（基于发票分配计算，正确处理跨项目收款）
            # 方式1: 通过发票分配表统计（仅限项目级别收款，避免与REF级别收款重复计算）
            invoice_alloc_filter = db.session.query(
                ProjectInvoice.header_id.label('hid'),
                ReceiptInvoiceAllocation.allocated_amount.label('amount')
            ).join(
                ReceiptInvoiceAllocation, ReceiptInvoiceAllocation.invoice_id == ProjectInvoice.id
            ).join(
                ProjectReceipt, ReceiptInvoiceAllocation.receipt_id == ProjectReceipt.id
            ).filter(
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id == None  # 仅项目级别收款，REF级别收款在方式2统计
            )

            # 方式2: REF级别直接收款
            ref_receipt_filter = db.session.query(
                ProjectReceipt.header_id.label('hid'),
                ProjectReceipt.amount.label('amount')
            ).filter(
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id.isnot(None)
            )

            # 方式3: 旧项目级别收款（没有分配记录也没有ref_id）- 向后兼容
            old_receipts_subq = db.session.query(
                ReceiptInvoiceAllocation.receipt_id
            ).distinct().subquery()

            old_project_receipt_filter = db.session.query(
                ProjectReceipt.header_id.label('hid'),
                ProjectReceipt.amount.label('amount')
            ).filter(
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id == None,  # 项目级别收款
                ~ProjectReceipt.id.in_(old_receipts_subq)  # 没有分配记录
            )

            # 合并三种方式
            combined_filter = union_all(invoice_alloc_filter, ref_receipt_filter, old_project_receipt_filter).alias('combined_receipts_filter')
            receipts_sum_subq = db.session.query(
                combined_filter.c.hid,
                db.func.coalesce(db.func.sum(combined_filter.c.amount), 0).label('total_received')
            ).group_by(combined_filter.c.hid).subquery()

            if payment_status == 'unpaid':
                # 未付款项目：销售金额 > 实际收款金额（与列表显示的"未付款"计算方式一致）
                unpaid_projects = db.session.query(refs_sum_subq.c.hid).outerjoin(
                    receipts_sum_subq,
                    receipts_sum_subq.c.hid == refs_sum_subq.c.hid
                ).filter(
                    refs_sum_subq.c.total_selling > 0,
                    db.or_(
                        receipts_sum_subq.c.total_received.is_(None),
                        refs_sum_subq.c.total_selling > db.func.coalesce(receipts_sum_subq.c.total_received, 0)
                    )
                )

                base_query = base_query.filter(
                    ProjectHeader.id.in_(unpaid_projects)
                )

            elif payment_status == 'paid':
                # 已付款项目：销售金额 <= 实际收款金额（与列表显示的"未付款"计算方式一致）
                paid_projects = db.session.query(refs_sum_subq.c.hid).outerjoin(
                    receipts_sum_subq,
                    receipts_sum_subq.c.hid == refs_sum_subq.c.hid
                ).filter(
                    refs_sum_subq.c.total_selling > 0,
                    refs_sum_subq.c.total_selling <= db.func.coalesce(receipts_sum_subq.c.total_received, 0)
                )

                base_query = base_query.filter(
                    ProjectHeader.id.in_(paid_projects)
                )
        
        if date_from:
            base_query = base_query.filter(ProjectHeader.created_at >= date_from)
        
        if date_to:
            base_query = base_query.filter(ProjectHeader.created_at <= date_to + ' 23:59:59')

        # 应用排序
        if sort_by == 'created_at_desc':
            base_query = base_query.order_by(ProjectHeader.created_at.desc())
        elif sort_by == 'created_at_asc':
            base_query = base_query.order_by(ProjectHeader.created_at.asc())
        elif sort_by == 'updated_at_desc':
            base_query = base_query.order_by(ProjectHeader.updated_at.desc())
        elif sort_by == 'updated_at_asc':
            base_query = base_query.order_by(ProjectHeader.updated_at.asc())
        elif sort_by == 'hid_desc':
            # 按项目编号降序（提取数字部分排序）
            base_query = base_query.order_by(
                db.func.cast(db.func.substr(ProjectHeader.hid, 2), db.Integer).desc()
            )
        elif sort_by == 'hid_asc':
            # 按项目编号升序
            base_query = base_query.order_by(
                db.func.cast(db.func.substr(ProjectHeader.hid, 2), db.Integer).asc()
            )
        else:
            # 默认按项目编号降序
            base_query = base_query.order_by(
                db.func.cast(db.func.substr(ProjectHeader.hid, 2), db.Integer).desc()
            )

        # 在分页前计算整个筛选结果的总计
        # 获取筛选后的所有项目ID（用于总计计算）
        all_filtered_ids_query = base_query.with_entities(ProjectHeader.id)
        all_filtered_ids = [r.id for r in all_filtered_ids_query.all()]

        # 计算整个筛选结果的总计
        summary_stats = {'total_cost': 0, 'total_profit': 0, 'total_balance': 0, 'total_selling': 0}
        if all_filtered_ids:
            try:
                from App_new.business.projects.models.invoice import ProjectInvoice
                from sqlalchemy import union_all

                # 汇总所有筛选项目的销售和成本
                all_refs_totals = db.session.query(
                    db.func.coalesce(db.func.sum(ProjectRef.selling_price), 0).label('total_selling'),
                    db.func.coalesce(db.func.sum(ProjectRef.cost_price), 0).label('total_cost')
                ).filter(ProjectRef.header_id.in_(all_filtered_ids)).first()

                total_selling_all = float(all_refs_totals.total_selling or 0)
                total_cost_all = float(all_refs_totals.total_cost or 0)

                # 计算总收款（通过发票分配 + REF直接收款 + 旧项目收款）
                # 方式1: 通过发票分配表统计
                alloc_total = db.session.query(
                    db.func.coalesce(db.func.sum(ReceiptInvoiceAllocation.allocated_amount), 0)
                ).join(
                    ProjectInvoice, ReceiptInvoiceAllocation.invoice_id == ProjectInvoice.id
                ).join(
                    ProjectReceipt, ReceiptInvoiceAllocation.receipt_id == ProjectReceipt.id
                ).filter(
                    ProjectInvoice.header_id.in_(all_filtered_ids),
                    ProjectReceipt.status == 'confirmed',
                    ProjectReceipt.ref_id == None
                ).scalar() or 0

                # 方式2: REF级别直接收款
                ref_receipt_total = db.session.query(
                    db.func.coalesce(db.func.sum(ProjectReceipt.amount), 0)
                ).filter(
                    ProjectReceipt.header_id.in_(all_filtered_ids),
                    ProjectReceipt.status == 'confirmed',
                    ProjectReceipt.ref_id.isnot(None)
                ).scalar() or 0

                # 方式3: 旧项目级别收款（没有分配记录也没有ref_id）
                old_receipts_subq = db.session.query(
                    ReceiptInvoiceAllocation.receipt_id
                ).distinct().subquery()

                old_receipt_total = db.session.query(
                    db.func.coalesce(db.func.sum(ProjectReceipt.amount), 0)
                ).filter(
                    ProjectReceipt.header_id.in_(all_filtered_ids),
                    ProjectReceipt.status == 'confirmed',
                    ProjectReceipt.ref_id == None,
                    ~ProjectReceipt.id.in_(old_receipts_subq)
                ).scalar() or 0

                total_received_all = float(alloc_total) + float(ref_receipt_total) + float(old_receipt_total)

                summary_stats = {
                    'total_selling': total_selling_all,
                    'total_cost': total_cost_all,
                    'total_profit': total_selling_all - total_cost_all,
                    'total_balance': total_selling_all - total_received_all
                }
            except Exception as e:
                print(f"计算总计时出错: {e}")
                import traceback
                traceback.print_exc()

        # 分页
        page = request.args.get('page', 1, type=int)
        per_page = 30  # 增加每页显示数量
        pagination = base_query.paginate(page=page, per_page=per_page, error_out=False)
        projects = pagination.items

        # 批量获取项目统计信息，避免 N+1 查询
        project_ids = [project.id for project in projects]
        project_stats = {}
        
        # 检查是否需要显示财务统计（通过查询参数控制）
        show_financial_stats = request.args.get('show_financial', 'true').lower() == 'true'
        
        if project_ids and show_financial_stats:
            try:
                # 批量查询 REF 数据
                refs_data = db.session.query(
                    ProjectRef.header_id,
                    db.func.sum(ProjectRef.selling_price).label('total_selling'),
                    db.func.sum(ProjectRef.cost_price).label('total_cost')
                ).filter(ProjectRef.header_id.in_(project_ids)).group_by(ProjectRef.header_id).all()
                
                # 批量查询收款数据 - 通过发票分配表计算每个项目的收款金额
                # 这样可以正确处理跨项目收款的情况
                from App_new.business.projects.models.invoice import ProjectInvoice

                # 方式1: 通过发票分配表统计（仅限项目级别收款，避免与REF级别收款重复计算）
                allocation_query = db.session.query(
                    ProjectInvoice.header_id,
                    db.func.sum(ReceiptInvoiceAllocation.allocated_amount).label('total_allocated')
                ).join(
                    ReceiptInvoiceAllocation, ReceiptInvoiceAllocation.invoice_id == ProjectInvoice.id
                ).join(
                    ProjectReceipt, ReceiptInvoiceAllocation.receipt_id == ProjectReceipt.id
                ).filter(
                    ProjectInvoice.header_id.in_(project_ids),
                    ProjectReceipt.status == 'confirmed',
                    ProjectReceipt.ref_id == None  # 仅项目级别收款，REF级别收款在方式2统计
                ).group_by(ProjectInvoice.header_id).all()

                # 方式2: REF级别直接收款（没有通过发票分配的收款）
                direct_ref_receipts = db.session.query(
                    ProjectReceipt.header_id,
                    db.func.sum(ProjectReceipt.amount).label('total_direct')
                ).filter(
                    ProjectReceipt.header_id.in_(project_ids),
                    ProjectReceipt.status == 'confirmed',
                    ProjectReceipt.ref_id.isnot(None)  # REF级别收款
                ).group_by(ProjectReceipt.header_id).all()

                # 方式3: 旧项目级别收款（没有分配记录也没有ref_id）- 向后兼容
                # 这些收款没有通过分配表关联到发票，需要直接按header_id计入
                old_project_receipts_subq = db.session.query(
                    ReceiptInvoiceAllocation.receipt_id
                ).distinct().subquery()

                old_project_receipts = db.session.query(
                    ProjectReceipt.header_id,
                    db.func.sum(ProjectReceipt.amount).label('total_old')
                ).filter(
                    ProjectReceipt.header_id.in_(project_ids),
                    ProjectReceipt.status == 'confirmed',
                    ProjectReceipt.ref_id == None,  # 项目级别收款
                    ~ProjectReceipt.id.in_(old_project_receipts_subq)  # 没有分配记录
                ).group_by(ProjectReceipt.header_id).all()

                # 合并三种收款方式的结果
                receipts_data = {}
                for r in allocation_query:
                    receipts_data[r.header_id] = float(r.total_allocated or 0)
                for r in direct_ref_receipts:
                    if r.header_id in receipts_data:
                        receipts_data[r.header_id] += float(r.total_direct or 0)
                    else:
                        receipts_data[r.header_id] = float(r.total_direct or 0)
                for r in old_project_receipts:
                    if r.header_id in receipts_data:
                        receipts_data[r.header_id] += float(r.total_old or 0)
                    else:
                        receipts_data[r.header_id] = float(r.total_old or 0)
                
                # 检查项目是否有EO或receipt
                from App_new.business.projects.models.eo import ProjectEO
                from App_new.business.projects.models.invoice import InvoiceItem

                # 检查EO
                eo_check = db.session.query(
                    ProjectEO.ref_id,
                    ProjectRef.header_id
                ).join(ProjectRef, ProjectEO.ref_id == ProjectRef.id).filter(
                    ProjectRef.header_id.in_(project_ids)
                ).distinct().all()

                # 检查receipt
                receipt_check = db.session.query(
                    ProjectReceipt.header_id
                ).filter(ProjectReceipt.header_id.in_(project_ids)).distinct().all()

                # === 结算状态计算 ===
                # 1. 每个项目的 REF 总数
                ref_counts = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(ProjectRef.id).label('ref_count')
                ).filter(ProjectRef.header_id.in_(project_ids)).group_by(ProjectRef.header_id).all()
                ref_count_dict = {r.header_id: r.ref_count for r in ref_counts}

                # 2. 每个项目有 InvoiceItem 的 REF 数量
                ref_with_invoice = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(db.distinct(InvoiceItem.ref_id)).label('invoiced_ref_count')
                ).join(InvoiceItem, InvoiceItem.ref_id == ProjectRef.id).filter(
                    ProjectRef.header_id.in_(project_ids)
                ).group_by(ProjectRef.header_id).all()
                invoiced_ref_dict = {r.header_id: r.invoiced_ref_count for r in ref_with_invoice}

                # 3. 每个项目的 EO 总数
                eo_counts = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(ProjectEO.id).label('eo_count')
                ).join(ProjectRef, ProjectEO.ref_id == ProjectRef.id).filter(
                    ProjectRef.header_id.in_(project_ids)
                ).group_by(ProjectRef.header_id).all()
                eo_count_dict = {r.header_id: r.eo_count for r in eo_counts}

                # 4. 每个项目已付款的 EO 数量（is_paid == True 视为已付款）
                eo_paid_counts = db.session.query(
                    ProjectRef.header_id,
                    db.func.count(ProjectEO.id).label('paid_eo_count')
                ).join(ProjectRef, ProjectEO.ref_id == ProjectRef.id).filter(
                    ProjectRef.header_id.in_(project_ids),
                    ProjectEO.is_paid == True
                ).group_by(ProjectRef.header_id).all()
                paid_eo_dict = {r.header_id: r.paid_eo_count for r in eo_paid_counts}

                # 构建统计字典
                for project_id in project_ids:
                    ref_info = next((r for r in refs_data if r.header_id == project_id), None)

                    # 检查是否有EO
                    has_eo = any(eo.header_id == project_id for eo in eo_check)
                    # 检查是否有receipt
                    has_receipt = any(rcpt.header_id == project_id for rcpt in receipt_check)

                    # 安全地获取数值，避免字段访问错误
                    total_selling = 0
                    total_cost = 0
                    total_received = 0

                    if ref_info:
                        try:
                            total_selling = float(ref_info.total_selling or 0)
                            total_cost = float(ref_info.total_cost or 0)
                        except (AttributeError, TypeError, ValueError) as e:
                            print(f"获取项目 {project_id} 价格信息时出错: {e}")
                            total_selling = 0
                            total_cost = 0

                    # 从字典获取收款金额
                    total_received = receipts_data.get(project_id, 0)
                    balance = total_selling - total_received

                    # 计算结算状态
                    ref_count = ref_count_dict.get(project_id, 0)
                    invoiced_ref_count = invoiced_ref_dict.get(project_id, 0)
                    eo_count = eo_count_dict.get(project_id, 0)
                    paid_eo_count = paid_eo_dict.get(project_id, 0)

                    # 可结算条件：所有REF生成EO 且 所有EO已付款 且 所有REF有发票 且 Balance=0
                    all_refs_have_eo = ref_count > 0 and ref_count == eo_count
                    all_eos_paid = eo_count > 0 and eo_count == paid_eo_count
                    all_refs_invoiced = ref_count > 0 and ref_count == invoiced_ref_count
                    balance_cleared = abs(balance) < 0.01  # 考虑浮点精度

                    can_settle = all_refs_have_eo and all_eos_paid and all_refs_invoiced and balance_cleared

                    project_stats[project_id] = {
                        'total_selling_price': total_selling,
                        'total_cost_price': total_cost,
                        'total_profit': total_selling - total_cost,
                        'total_received': total_received,
                        'balance': balance,
                        'has_eo': has_eo,
                        'has_receipt': has_receipt,
                        'has_invoice': invoiced_ref_count > 0,
                        # 结算相关
                        'ref_count': ref_count,
                        'invoiced_ref_count': invoiced_ref_count,
                        'eo_count': eo_count,
                        'paid_eo_count': paid_eo_count,
                        'can_settle': can_settle
                    }
            except Exception as e:
                print(f"构建项目统计时出错: {e}")
                # 如果统计构建失败，提供默认值
                for project_id in project_ids:
                    project_stats[project_id] = {
                        'total_selling_price': 0,
                        'total_cost_price': 0,
                        'total_profit': 0,
                        'total_received': 0,
                        'balance': 0,
                        'has_eo': False,
                        'has_receipt': False,
                        'has_invoice': False,
                        'ref_count': 0,
                        'invoiced_ref_count': 0,
                        'eo_count': 0,
                        'paid_eo_count': 0,
                        'can_settle': False
                    }
        else:
            # 如果不显示财务统计，提供空的统计数据
            for project_id in project_ids:
                try:
                    # 即使不显示财务统计，也需要检查EO和receipt状态
                    from App_new.business.projects.models.eo import ProjectEO
                    
                    # 检查EO
                    has_eo = db.session.query(ProjectEO).join(ProjectRef, ProjectEO.ref_id == ProjectRef.id).filter(
                        ProjectRef.header_id == project_id
                    ).first() is not None
                    
                    # 检查receipt
                    has_receipt = db.session.query(ProjectReceipt).filter(
                        ProjectReceipt.header_id == project_id
                    ).first() is not None
                    
                    # 检查是否有发票
                    from App_new.business.projects.models.invoice import ProjectInvoice
                    has_invoice = db.session.query(ProjectInvoice).filter(
                        ProjectInvoice.header_id == project_id,
                        ProjectInvoice.status != 'cancelled'
                    ).first() is not None

                    project_stats[project_id] = {
                        'total_selling_price': 0,
                        'total_cost_price': 0,
                        'total_profit': 0,
                        'total_received': 0,
                        'balance': 0,
                        'has_eo': has_eo,
                        'has_receipt': has_receipt,
                        'has_invoice': has_invoice,
                        'ref_count': 0,
                        'invoiced_ref_count': 0,
                        'eo_count': 0,
                        'paid_eo_count': 0,
                        'can_settle': False
                    }
                except Exception as e:
                    print(f"检查项目 {project_id} EO/receipt状态时出错: {e}")
                    # 如果检查失败，提供默认值
                    project_stats[project_id] = {
                        'total_selling_price': 0,
                        'total_cost_price': 0,
                        'total_profit': 0,
                        'total_received': 0,
                        'balance': 0,
                        'has_eo': False,
                        'has_receipt': False,
                        'has_invoice': False,
                        'ref_count': 0,
                        'invoiced_ref_count': 0,
                        'eo_count': 0,
                        'paid_eo_count': 0,
                        'can_settle': False
                    }

        # 注意：付款状态筛选已在数据库层面完成，无需在此二次筛选
        
        # 获取总体统计（简化版本，只获取基本计数）
        total_stats = {
            'total_projects': ProjectHeader.query.count(),
            'active_projects': ProjectHeader.query.filter_by(status='active').count(),
            'completed_projects': ProjectHeader.query.filter_by(status='completed').count(),
            'draft_projects': ProjectHeader.query.filter_by(status='draft').count(),
            'completed_this_month': ProjectHeader.query.filter(
                ProjectHeader.status == 'completed',
                ProjectHeader.updated_at >= datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            ).count()
        }
        
        # 应用金额筛选（在数据库层面进行，提高效率）
        if selling_min or selling_max or profit_min or profit_max or balance_min or balance_max or profit_status:
            try:
                # 构建子查询来筛选符合条件的项目
                subquery = db.session.query(ProjectRef.header_id).group_by(ProjectRef.header_id)
                
                if selling_min:
                    subquery = subquery.having(db.func.sum(ProjectRef.selling_price) >= float(selling_min))
                if selling_max:
                    subquery = subquery.having(db.func.sum(ProjectRef.selling_price) <= float(selling_max))
                if profit_min:
                    subquery = subquery.having(
                        db.func.sum(ProjectRef.selling_price) - db.func.sum(ProjectRef.cost_price) >= float(profit_min)
                    )
                if profit_max:
                    subquery = subquery.having(
                        db.func.sum(ProjectRef.selling_price) - db.func.sum(ProjectRef.cost_price) <= float(profit_max)
                    )

                # 应用盈亏状态筛选
                if profit_status:
                    if profit_status == 'loss':
                        # 亏损：利润 < 0
                        subquery = subquery.having(
                            db.func.sum(ProjectRef.selling_price) - db.func.sum(ProjectRef.cost_price) < 0
                        )
                    elif profit_status == 'profit':
                        # 盈利：利润 > 0
                        subquery = subquery.having(
                            db.func.sum(ProjectRef.selling_price) - db.func.sum(ProjectRef.cost_price) > 0
                        )
                    elif profit_status == 'zero':
                        # 持平：利润 = 0
                        subquery = subquery.having(
                            db.func.sum(ProjectRef.selling_price) - db.func.sum(ProjectRef.cost_price) == 0
                        )

                # 应用余额筛选（销售金额 - 已收款金额）
                if balance_min or balance_max:
                    print(f"应用余额筛选: balance_min={balance_min}, balance_max={balance_max}")
                    # 需要JOIN收款表来计算余额
                    subquery = subquery.outerjoin(
                        ProjectReceipt, 
                        ProjectRef.header_id == ProjectReceipt.header_id
                    ).group_by(ProjectRef.header_id)
                    
                    if balance_min:
                        print(f"设置余额最小值筛选: {balance_min}")
                        # 确保余额大于0的项目被筛选出来
                        subquery = subquery.having(
                            (db.func.sum(ProjectRef.selling_price) - db.func.coalesce(db.func.sum(ProjectReceipt.amount), 0)) > 0
                        )
                        # 如果指定了具体的balance_min值，则使用该值
                        if float(balance_min) > 0:
                            subquery = subquery.having(
                                (db.func.sum(ProjectRef.selling_price) - db.func.coalesce(db.func.sum(ProjectReceipt.amount), 0)) >= float(balance_min)
                            )
                    if balance_max:
                        print(f"设置余额最大值筛选: {balance_max}")
                        subquery = subquery.having(
                            (db.func.sum(ProjectRef.selling_price) - db.func.coalesce(db.func.sum(ProjectReceipt.amount), 0)) <= float(balance_max)
                        )
                
                # 应用子查询筛选
                base_query = base_query.filter(ProjectHeader.id.in_(subquery))
                
                # 重新执行分页查询
                pagination = base_query.paginate(page=page, per_page=per_page, error_out=False)
                projects = pagination.items
                
                # 重新获取筛选后的项目统计
                project_ids = [project.id for project in projects]
                
                if project_ids:
                    # 批量查询 REF 数据
                    refs_data = db.session.query(
                        ProjectRef.header_id,
                        db.func.sum(ProjectRef.selling_price).label('total_selling'),
                        db.func.sum(ProjectRef.cost_price).label('total_cost')
                    ).filter(ProjectRef.header_id.in_(project_ids)).group_by(ProjectRef.header_id).all()
                    
                    # 批量查询收款数据 - 通过发票分配表计算每个项目的收款金额
                    from App_new.business.projects.models.invoice import ProjectInvoice

                    # 方式1: 通过发票分配表统计（仅限项目级别收款，避免与REF级别收款重复计算）
                    allocation_query = db.session.query(
                        ProjectInvoice.header_id,
                        db.func.sum(ReceiptInvoiceAllocation.allocated_amount).label('total_allocated')
                    ).join(
                        ReceiptInvoiceAllocation, ReceiptInvoiceAllocation.invoice_id == ProjectInvoice.id
                    ).join(
                        ProjectReceipt, ReceiptInvoiceAllocation.receipt_id == ProjectReceipt.id
                    ).filter(
                        ProjectInvoice.header_id.in_(project_ids),
                        ProjectReceipt.status == 'confirmed',
                        ProjectReceipt.ref_id == None  # 仅项目级别收款
                    ).group_by(ProjectInvoice.header_id).all()

                    # 方式2: REF级别直接收款
                    direct_ref_receipts = db.session.query(
                        ProjectReceipt.header_id,
                        db.func.sum(ProjectReceipt.amount).label('total_direct')
                    ).filter(
                        ProjectReceipt.header_id.in_(project_ids),
                        ProjectReceipt.status == 'confirmed',
                        ProjectReceipt.ref_id.isnot(None)
                    ).group_by(ProjectReceipt.header_id).all()

                    # 方式3: 旧项目级别收款（没有分配记录也没有ref_id）- 向后兼容
                    old_receipts_subq = db.session.query(
                        ReceiptInvoiceAllocation.receipt_id
                    ).distinct().subquery()

                    old_project_receipts = db.session.query(
                        ProjectReceipt.header_id,
                        db.func.sum(ProjectReceipt.amount).label('total_old')
                    ).filter(
                        ProjectReceipt.header_id.in_(project_ids),
                        ProjectReceipt.status == 'confirmed',
                        ProjectReceipt.ref_id == None,
                        ~ProjectReceipt.id.in_(old_receipts_subq)
                    ).group_by(ProjectReceipt.header_id).all()

                    # 合并三种收款方式的结果
                    receipts_data = {}
                    for r in allocation_query:
                        receipts_data[r.header_id] = float(r.total_allocated or 0)
                    for r in direct_ref_receipts:
                        if r.header_id in receipts_data:
                            receipts_data[r.header_id] += float(r.total_direct or 0)
                        else:
                            receipts_data[r.header_id] = float(r.total_direct or 0)
                    for r in old_project_receipts:
                        if r.header_id in receipts_data:
                            receipts_data[r.header_id] += float(r.total_old or 0)
                        else:
                            receipts_data[r.header_id] = float(r.total_old or 0)

                    # 检查项目是否有EO或receipt
                    from App_new.business.projects.models.eo import ProjectEO

                    # 检查EO
                    eo_check = db.session.query(
                        ProjectEO.ref_id,
                        ProjectRef.header_id
                    ).join(ProjectRef, ProjectEO.ref_id == ProjectRef.id).filter(
                        ProjectRef.header_id.in_(project_ids)
                    ).distinct().all()

                    # 检查receipt
                    receipt_check = db.session.query(
                        ProjectReceipt.header_id
                    ).filter(ProjectReceipt.header_id.in_(project_ids)).distinct().all()

                    # === 结算状态计算（筛选后重新计算） ===
                    from App_new.business.projects.models.invoice import InvoiceItem

                    # 1. 每个项目的 REF 总数
                    ref_counts = db.session.query(
                        ProjectRef.header_id,
                        db.func.count(ProjectRef.id).label('ref_count')
                    ).filter(ProjectRef.header_id.in_(project_ids)).group_by(ProjectRef.header_id).all()
                    ref_count_dict = {r.header_id: r.ref_count for r in ref_counts}

                    # 2. 每个项目有 InvoiceItem 的 REF 数量
                    ref_with_invoice = db.session.query(
                        ProjectRef.header_id,
                        db.func.count(db.distinct(InvoiceItem.ref_id)).label('invoiced_ref_count')
                    ).join(InvoiceItem, InvoiceItem.ref_id == ProjectRef.id).filter(
                        ProjectRef.header_id.in_(project_ids)
                    ).group_by(ProjectRef.header_id).all()
                    invoiced_ref_dict = {r.header_id: r.invoiced_ref_count for r in ref_with_invoice}

                    # 3. 每个项目的 EO 总数
                    eo_counts = db.session.query(
                        ProjectRef.header_id,
                        db.func.count(ProjectEO.id).label('eo_count')
                    ).join(ProjectRef, ProjectEO.ref_id == ProjectRef.id).filter(
                        ProjectRef.header_id.in_(project_ids)
                    ).group_by(ProjectRef.header_id).all()
                    eo_count_dict = {r.header_id: r.eo_count for r in eo_counts}

                    # 4. 每个项目已付款的 EO 数量（is_paid == True 视为已付款）
                    eo_paid_counts = db.session.query(
                        ProjectRef.header_id,
                        db.func.count(ProjectEO.id).label('paid_eo_count')
                    ).join(ProjectRef, ProjectEO.ref_id == ProjectRef.id).filter(
                        ProjectRef.header_id.in_(project_ids),
                        ProjectEO.is_paid == True
                    ).group_by(ProjectRef.header_id).all()
                    paid_eo_dict = {r.header_id: r.paid_eo_count for r in eo_paid_counts}

                    # 更新统计字典
                    for project_id in project_ids:
                        ref_info = next((r for r in refs_data if r.header_id == project_id), None)

                        # 检查是否有EO
                        has_eo = any(eo.header_id == project_id for eo in eo_check)
                        # 检查是否有receipt
                        has_receipt = any(rcpt.header_id == project_id for rcpt in receipt_check)

                        # 安全地获取数值
                        total_selling = 0
                        total_cost = 0

                        if ref_info:
                            try:
                                total_selling = float(ref_info.total_selling or 0)
                                total_cost = float(ref_info.total_cost or 0)
                            except (AttributeError, TypeError, ValueError) as e:
                                print(f"获取筛选后项目 {project_id} 价格信息时出错: {e}")
                                total_selling = 0
                                total_cost = 0

                        # 从字典获取收款金额
                        total_received = receipts_data.get(project_id, 0)
                        balance = total_selling - total_received

                        # 计算结算状态
                        ref_count = ref_count_dict.get(project_id, 0)
                        invoiced_ref_count = invoiced_ref_dict.get(project_id, 0)
                        eo_count = eo_count_dict.get(project_id, 0)
                        paid_eo_count = paid_eo_dict.get(project_id, 0)

                        # 可结算条件：所有REF生成EO 且 所有EO已付款 且 所有REF有发票 且 Balance=0
                        all_refs_have_eo = ref_count > 0 and ref_count == eo_count
                        all_eos_paid = eo_count > 0 and eo_count == paid_eo_count
                        all_refs_invoiced = ref_count > 0 and ref_count == invoiced_ref_count
                        balance_cleared = abs(balance) < 0.01

                        can_settle = all_refs_have_eo and all_eos_paid and all_refs_invoiced and balance_cleared

                        # 更新现有的统计数据
                        if project_id in project_stats:
                            project_stats[project_id].update({
                                'total_selling_price': total_selling,
                                'total_cost_price': total_cost,
                                'total_profit': total_selling - total_cost,
                                'total_received': total_received,
                                'balance': balance,
                                'has_eo': has_eo,
                                'has_receipt': has_receipt,
                                'has_invoice': invoiced_ref_count > 0,
                                'ref_count': ref_count,
                                'invoiced_ref_count': invoiced_ref_count,
                                'eo_count': eo_count,
                                'paid_eo_count': paid_eo_count,
                                'can_settle': can_settle
                            })
                        else:
                            project_stats[project_id] = {
                                'total_selling_price': total_selling,
                                'total_cost_price': total_cost,
                                'total_profit': total_selling - total_cost,
                                'total_received': total_received,
                                'balance': balance,
                                'has_eo': has_eo,
                                'has_receipt': has_receipt,
                                'has_invoice': invoiced_ref_count > 0,
                                'ref_count': ref_count,
                                'invoiced_ref_count': invoiced_ref_count,
                                'eo_count': eo_count,
                                'paid_eo_count': paid_eo_count,
                                'can_settle': can_settle
                            }
            except Exception as e:
                print(f"应用金额筛选时出错: {e}")
                # 如果筛选失败，继续使用原始查询结果
                pass

        # 提供客户公司列表用于前端下拉选择
        companies = CustomerCompany.query.order_by(CustomerCompany.company_name).all()

        return render_template(
            'business/projects/project_list.html',
            projects=projects,
            project_stats=project_stats,
            pagination=pagination,
            total_stats=total_stats,
            summary_stats=summary_stats,
            companies=companies,
            filters={
                'status': status,
                'search': search,
                'company': company,
                'leader': leader,
                'contact': contact,
                'staff_name': staff_name,
                'date_from': date_from,
                'date_to': date_to,
                'selling_min': selling_min,
                'selling_max': selling_max,
                'payment_status': payment_status,
                'profit_min': profit_min,
                'profit_max': profit_max,
                'profit_status': profit_status,
                'settlement_status': settlement_status,
                'balance_min': balance_min,
                'balance_max': balance_max,
                'sort_by': sort_by
            }
        )
        
    except Exception as e:
        flash(f'加载项目列表失败：{str(e)}', 'error')
        print(f"项目列表加载错误详情: {str(e)}")  # 添加控制台日志
        print(f"错误类型: {type(e).__name__}")  # 添加错误类型信息
        import traceback
        traceback.print_exc()
        return render_template(
            'business/projects/project_list.html',
            projects=[],
            project_stats={},
            pagination=None,
            total_stats={},
            summary_stats={'total_cost': 0, 'total_profit': 0, 'total_balance': 0, 'total_selling': 0},
            companies=[],
            filters={}
        )

@bp.route('/api/stats')
@login_required
@staff_only
def api_project_stats():
    """获取项目统计信息的API"""
    try:
        stats_service = ProjectStatsService()
        stats = stats_service.get_total_stats()
        
        return jsonify({
            'success': True,
            'data': stats
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@bp.route('/api/quick-filter/<filter_type>')
@login_required
@staff_only
def api_quick_filter(filter_type):
    """快速筛选API"""
    try:
        project_service = ProjectService()
        
        if filter_type == 'active':
            projects = project_service.get_projects_by_status('active')
        elif filter_type == 'completed':
            projects = project_service.get_projects_by_status('completed')
        elif filter_type == 'draft':
            projects = project_service.get_projects_by_status('draft')
        elif filter_type == 'profitable':
            projects = project_service.get_profitable_projects()
        elif filter_type == 'unpaid':
            projects = project_service.get_unpaid_projects()
        else:
            projects = []
        
        return jsonify({
            'success': True,
            'data': [project.to_dict() for project in projects]
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@bp.route('/export/excel')
@login_required
@staff_only
def export_excel():
    """导出项目列表为Excel"""
    try:
        # 获取筛选参数（与list_projects相同）
        status = request.args.get('status', '')
        search = request.args.get('search', '')
        company = request.args.get('company', '')
        leader = request.args.get('leader', '')
        contact = request.args.get('contact', '')
        staff_name = request.args.get('staff_name', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        selling_min = request.args.get('selling_min', '')
        selling_max = request.args.get('selling_max', '')
        profit_min = request.args.get('profit_min', '')
        profit_max = request.args.get('profit_max', '')
        balance_min = request.args.get('balance_min', '')
        balance_max = request.args.get('balance_max', '')
        payment_status = request.args.get('payment_status', '')
        sort_by = request.args.get('sort_by', 'hid_desc')
        
        # 构建查询（与list_projects相同逻辑）
        base_query = ProjectHeader.query
        base_query = base_query.options(db.joinedload(ProjectHeader.company))
        
        # 根据员工等级过滤项目
        if current_user.role and current_user.role.name == 'staff':
            staff_level = 1
            if current_user.profile:
                staff_level = current_user.profile.staff_level or 1
            
            if staff_level == 1:
                user_full_name = current_user.profile.get_full_name() if current_user.profile else None
                if user_full_name and user_full_name != "未设置姓名":
                    base_query = base_query.filter(
                        db.or_(
                            ProjectHeader.staff_id == current_user.id,
                            ProjectHeader.staff_name == current_user.username,
                            ProjectHeader.staff_name == user_full_name
                        )
                    )
                else:
                    base_query = base_query.filter(
                        db.or_(
                            ProjectHeader.staff_id == current_user.id,
                            ProjectHeader.staff_name == current_user.username
                        )
                    )
        
        # 应用筛选条件
        if status:
            base_query = base_query.filter(ProjectHeader.status == status)
        if search:
            search_lower = search.lower()
            # 子查询：查找包含搜索关键词的项目成员所属的项目ID
            from ..models.project_member import ProjectMember
            member_subquery = db.session.query(ProjectMember.header_id).filter(
                db.or_(
                    ProjectMember.member_name.ilike(f'%{search_lower}%'),
                    ProjectMember.member_name_en.ilike(f'%{search_lower}%')
                )
            ).subquery()

            base_query = base_query.filter(
                db.or_(
                    ProjectHeader.hid.ilike(f'%{search_lower}%'),
                    ProjectHeader.desc.ilike(f'%{search_lower}%'),
                    ProjectHeader.contact.ilike(f'%{search_lower}%'),
                    ProjectHeader.leader_name.ilike(f'%{search_lower}%'),
                    ProjectHeader.id.in_(member_subquery)  # 搜索乘客姓名
                )
            )
        if company:
            base_query = base_query.join(CustomerCompany).filter(
                CustomerCompany.company_name.like(f'%{company}%')
            )
        if leader:
            base_query = base_query.filter(ProjectHeader.leader_name.like(f'%{leader}%'))
        if contact:
            base_query = base_query.filter(ProjectHeader.contact.like(f'%{contact}%'))
        if staff_name:
            base_query = base_query.filter(ProjectHeader.staff_name.like(f'%{staff_name}%'))
        
        project_type = request.args.get('type', '')
        if project_type:
            base_query = base_query.filter(ProjectHeader.type == project_type)
        
        # 付款状态筛选 - 基于发票数据判断
        if payment_status:
            from App_new.business.projects.models.invoice import ProjectInvoice
            from sqlalchemy import union_all

            # 按 header 统计发票金额和已付金额
            invoice_sum_subq = db.session.query(
                ProjectInvoice.header_id.label('hid'),
                db.func.coalesce(db.func.sum(ProjectInvoice.amount), 0).label('invoice_total'),
                db.func.coalesce(db.func.sum(ProjectInvoice.paid_amount), 0).label('invoice_paid')
            ).filter(
                ProjectInvoice.status != 'cancelled'
            ).group_by(ProjectInvoice.header_id).subquery()

            # 按 header 统计总销售金额（用于没有发票的项目）
            refs_sum_subq = db.session.query(
                ProjectRef.header_id.label('hid'),
                db.func.coalesce(db.func.sum(ProjectRef.selling_price), 0).label('total_selling')
            ).group_by(ProjectRef.header_id).subquery()

            # 按 header 统计总收款金额（基于发票分配计算，正确处理跨项目收款）
            # 方式1: 通过发票分配表统计（仅限项目级别收款，避免与REF级别收款重复计算）
            invoice_alloc_export = db.session.query(
                ProjectInvoice.header_id.label('hid'),
                ReceiptInvoiceAllocation.allocated_amount.label('amount')
            ).join(
                ReceiptInvoiceAllocation, ReceiptInvoiceAllocation.invoice_id == ProjectInvoice.id
            ).join(
                ProjectReceipt, ReceiptInvoiceAllocation.receipt_id == ProjectReceipt.id
            ).filter(
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id == None  # 仅项目级别收款，REF级别收款在方式2统计
            )

            # 方式2: REF级别直接收款
            ref_receipt_export = db.session.query(
                ProjectReceipt.header_id.label('hid'),
                ProjectReceipt.amount.label('amount')
            ).filter(
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id.isnot(None)
            )

            # 方式3: 旧项目级别收款（没有分配记录也没有ref_id）- 向后兼容
            old_receipts_export_subq = db.session.query(
                ReceiptInvoiceAllocation.receipt_id
            ).distinct().subquery()

            old_project_receipt_export = db.session.query(
                ProjectReceipt.header_id.label('hid'),
                ProjectReceipt.amount.label('amount')
            ).filter(
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id == None,  # 项目级别收款
                ~ProjectReceipt.id.in_(old_receipts_export_subq)  # 没有分配记录
            )

            # 合并三种方式
            combined_export = union_all(invoice_alloc_export, ref_receipt_export, old_project_receipt_export).alias('combined_receipts_export')
            receipts_sum_subq = db.session.query(
                combined_export.c.hid,
                db.func.coalesce(db.func.sum(combined_export.c.amount), 0).label('total_received')
            ).group_by(combined_export.c.hid).subquery()

            if payment_status == 'unpaid':
                # 未付款项目：销售金额 > 实际收款金额（与列表显示的"未付款"计算方式一致）
                unpaid_projects = db.session.query(refs_sum_subq.c.hid).outerjoin(
                    receipts_sum_subq,
                    receipts_sum_subq.c.hid == refs_sum_subq.c.hid
                ).filter(
                    refs_sum_subq.c.total_selling > 0,
                    db.or_(
                        receipts_sum_subq.c.total_received.is_(None),
                        refs_sum_subq.c.total_selling > db.func.coalesce(receipts_sum_subq.c.total_received, 0)
                    )
                )

                base_query = base_query.filter(
                    ProjectHeader.id.in_(unpaid_projects)
                )

            elif payment_status == 'paid':
                # 已付款项目：销售金额 <= 实际收款金额（与列表显示的"未付款"计算方式一致）
                paid_projects = db.session.query(refs_sum_subq.c.hid).outerjoin(
                    receipts_sum_subq,
                    receipts_sum_subq.c.hid == refs_sum_subq.c.hid
                ).filter(
                    refs_sum_subq.c.total_selling > 0,
                    refs_sum_subq.c.total_selling <= db.func.coalesce(receipts_sum_subq.c.total_received, 0)
                )

                base_query = base_query.filter(
                    ProjectHeader.id.in_(paid_projects)
                )

        if date_from:
            base_query = base_query.filter(ProjectHeader.created_at >= date_from)
        if date_to:
            base_query = base_query.filter(ProjectHeader.created_at <= date_to + ' 23:59:59')

        # 应用排序
        if sort_by == 'created_at_desc':
            base_query = base_query.order_by(ProjectHeader.created_at.desc())
        elif sort_by == 'created_at_asc':
            base_query = base_query.order_by(ProjectHeader.created_at.asc())
        elif sort_by == 'updated_at_desc':
            base_query = base_query.order_by(ProjectHeader.updated_at.desc())
        elif sort_by == 'updated_at_asc':
            base_query = base_query.order_by(ProjectHeader.updated_at.asc())
        elif sort_by == 'hid_desc':
            base_query = base_query.order_by(
                db.func.cast(db.func.substr(ProjectHeader.hid, 2), db.Integer).desc()
            )
        elif sort_by == 'hid_asc':
            base_query = base_query.order_by(
                db.func.cast(db.func.substr(ProjectHeader.hid, 2), db.Integer).asc()
            )
        else:
            # 默认按项目编号降序
            base_query = base_query.order_by(
                db.func.cast(db.func.substr(ProjectHeader.hid, 2), db.Integer).desc()
            )

        # 获取所有符合条件的项目（不分页）
        projects = base_query.all()
        
        # 批量获取项目统计信息
        project_ids = [project.id for project in projects]
        project_stats = {}
        
        if project_ids:
            refs_data = db.session.query(
                ProjectRef.header_id,
                db.func.sum(ProjectRef.selling_price).label('total_selling'),
                db.func.sum(ProjectRef.cost_price).label('total_cost')
            ).filter(ProjectRef.header_id.in_(project_ids)).group_by(ProjectRef.header_id).all()
            
            # 批量查询收款数据 - 通过发票分配表计算每个项目的收款金额
            from App_new.business.projects.models.invoice import ProjectInvoice

            # 方式1: 通过发票分配表统计（仅限项目级别收款，避免与REF级别收款重复计算）
            allocation_query = db.session.query(
                ProjectInvoice.header_id,
                db.func.sum(ReceiptInvoiceAllocation.allocated_amount).label('total_allocated')
            ).join(
                ReceiptInvoiceAllocation, ReceiptInvoiceAllocation.invoice_id == ProjectInvoice.id
            ).join(
                ProjectReceipt, ReceiptInvoiceAllocation.receipt_id == ProjectReceipt.id
            ).filter(
                ProjectInvoice.header_id.in_(project_ids),
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id == None  # 仅项目级别收款，REF级别收款在方式2统计
            ).group_by(ProjectInvoice.header_id).all()

            # 方式2: REF级别直接收款
            direct_ref_receipts = db.session.query(
                ProjectReceipt.header_id,
                db.func.sum(ProjectReceipt.amount).label('total_direct')
            ).filter(
                ProjectReceipt.header_id.in_(project_ids),
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id.isnot(None)
            ).group_by(ProjectReceipt.header_id).all()

            # 方式3: 旧项目级别收款（没有分配记录也没有ref_id）- 向后兼容
            old_receipts_subq = db.session.query(
                ReceiptInvoiceAllocation.receipt_id
            ).distinct().subquery()

            old_project_receipts = db.session.query(
                ProjectReceipt.header_id,
                db.func.sum(ProjectReceipt.amount).label('total_old')
            ).filter(
                ProjectReceipt.header_id.in_(project_ids),
                ProjectReceipt.status == 'confirmed',
                ProjectReceipt.ref_id == None,
                ~ProjectReceipt.id.in_(old_receipts_subq)
            ).group_by(ProjectReceipt.header_id).all()

            # 合并三种收款方式的结果
            receipts_data = {}
            for r in allocation_query:
                receipts_data[r.header_id] = float(r.total_allocated or 0)
            for r in direct_ref_receipts:
                if r.header_id in receipts_data:
                    receipts_data[r.header_id] += float(r.total_direct or 0)
                else:
                    receipts_data[r.header_id] = float(r.total_direct or 0)
            for r in old_project_receipts:
                if r.header_id in receipts_data:
                    receipts_data[r.header_id] += float(r.total_old or 0)
                else:
                    receipts_data[r.header_id] = float(r.total_old or 0)

            for project_id in project_ids:
                ref_info = next((r for r in refs_data if r.header_id == project_id), None)

                total_selling = float(ref_info.total_selling or 0) if ref_info else 0
                total_cost = float(ref_info.total_cost or 0) if ref_info else 0
                total_received = receipts_data.get(project_id, 0)

                project_stats[project_id] = {
                    'total_selling_price': total_selling,
                    'total_cost_price': total_cost,
                    'total_profit': total_selling - total_cost,
                    'total_received': total_received,
                    'balance': total_selling - total_received
                }
        
        # 准备Excel数据
        data = []
        for project in projects:
            stats = project_stats.get(project.id, {})
            data.append({
                '项目编号': project.hid or '',
                '描述': project.desc or '',
                '客户公司': project.company.company_name if project.company else '',
                '联系人': project.contact or '',
                '经办人': project.staff_name or '',
                '类型': project.type or '综合',
                '销售金额': stats.get('total_selling_price', 0),
                '成本': stats.get('total_cost_price', 0),
                '利润': stats.get('total_profit', 0),
                '未付款': stats.get('balance', 0),
                '创建时间': project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else '',
                '状态': project.status or ''
            })
        
        # 添加总计行
        if data:
            total_cost = sum(stats.get('total_cost_price', 0) for stats in project_stats.values())
            total_profit = sum(stats.get('total_profit', 0) for stats in project_stats.values())
            total_balance = sum(stats.get('balance', 0) for stats in project_stats.values())
            
            data.append({
                '项目编号': '总计',
                '描述': '',
                '客户公司': '',
                '联系人': '',
                '经办人': '',
                '类型': '',
                '销售金额': '',
                '成本': total_cost,
                '利润': total_profit,
                '未付款': total_balance,
                '创建时间': '',
                '状态': ''
            })
        
        df = pd.DataFrame(data)
        
        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='项目列表')
            worksheet = writer.sheets['项目列表']
            
            # 设置列宽
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max() if len(df) > 0 else 0,
                    len(col)
                ) + 2
                col_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[col_letter].width = min(max_length, 50)
            
            # 设置总计行样式
            if len(df) > 0:
                from openpyxl.styles import Font, PatternFill
                total_row = len(df)
                for idx, cell in enumerate(worksheet[total_row]):
                    cell.font = Font(bold=True)
                    if idx == 0:  # 项目编号列
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        output.seek(0)
        filename = f'项目列表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'导出失败：{str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('business_projects.list.list_projects'))


@bp.route('/api/settle/<int:project_id>', methods=['POST'])
@login_required
@staff_only
def settle_project(project_id):
    """
    结算项目
    条件：所有REF有发票 且 所有EO已付款 且 Balance=0
    """
    try:
        project = ProjectHeader.query.get_or_404(project_id)

        # 检查是否已结算
        if project.is_settled:
            return jsonify({'success': False, 'message': '该项目已结算'})

        # 验证结算条件
        from App_new.business.projects.models.eo import ProjectEO
        from App_new.business.projects.models.invoice import InvoiceItem

        # 1. 获取项目的所有 REF
        refs = ProjectRef.query.filter_by(header_id=project_id).all()
        ref_count = len(refs)

        if ref_count == 0:
            return jsonify({'success': False, 'message': '该项目没有 REF 记录，无法结算'})

        # 2. 检查所有 REF 是否都有发票
        ref_ids = [r.id for r in refs]
        invoiced_ref_count = db.session.query(db.func.count(db.distinct(InvoiceItem.ref_id))).filter(
            InvoiceItem.ref_id.in_(ref_ids)
        ).scalar()

        if invoiced_ref_count != ref_count:
            return jsonify({
                'success': False,
                'message': f'有 {ref_count - invoiced_ref_count} 个 REF 未开发票'
            })

        # 3. 检查所有 EO 是否都已付款
        eos = db.session.query(ProjectEO).join(ProjectRef, ProjectEO.ref_id == ProjectRef.id).filter(
            ProjectRef.header_id == project_id
        ).all()

        unpaid_eos = [eo for eo in eos if not eo.pay_amount or float(eo.pay_amount) <= 0]
        if unpaid_eos:
            return jsonify({
                'success': False,
                'message': f'有 {len(unpaid_eos)} 个 EO 未付款'
            })

        # 4. 检查 Balance 是否为 0
        total_selling = sum(float(r.selling_price or 0) for r in refs)

        # 计算已收款（简化版，与列表页一致）
        receipts = ProjectReceipt.query.filter_by(
            header_id=project_id,
            status='confirmed'
        ).all()
        total_received = sum(float(r.amount or 0) for r in receipts)

        balance = total_selling - total_received
        if abs(balance) > 0.01:  # 考虑浮点精度
            return jsonify({
                'success': False,
                'message': f'项目余额未清零，当前余额: S${balance:.2f}'
            })

        # 执行结算
        project.is_settled = True
        project.settled_at = datetime.now()
        project.settled_by = current_user.display_name if hasattr(current_user, 'display_name') else str(current_user.id)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': '项目结算成功'
        })

    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'结算失败: {str(e)}'})


@bp.route('/api/unsettle/<int:project_id>', methods=['POST'])
@login_required
@staff_only
def unsettle_project(project_id):
    """
    取消结算（撤销结算状态）
    """
    try:
        project = ProjectHeader.query.get_or_404(project_id)

        if not project.is_settled:
            return jsonify({'success': False, 'message': '该项目未结算'})

        project.is_settled = False
        project.settled_at = None
        project.settled_by = None

        db.session.commit()

        return jsonify({
            'success': True,
            'message': '已取消结算'
        })

    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'操作失败: {str(e)}'})
