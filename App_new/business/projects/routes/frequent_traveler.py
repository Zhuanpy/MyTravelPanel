# -*- coding: utf-8 -*-
"""常用旅客管理路由"""

import json
import io
import os
import uuid
import zipfile
from pathlib import Path
from flask import Blueprint, request, jsonify, render_template, send_file, current_app
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from App_new.exts import db, csrf
from App_new.utils.decorators import staff_only
from App_new.business.projects.models.frequent_traveler import FrequentTraveler
from App_new.business.projects.models.traveler_file import TravelerFile
from App_new.business.projects.models.project import CustomerCompany

# 允许上传的文件类型
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png', 'gif', 'webp', 'bmp'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _usage_subquery():
    """常旅客被订单引用的统计（按订单数 + 最近一次引用）

    项目人员和常旅客之间没有外键，全站都是按 member_name 精确匹配来认这个人的
    （见 project_members.get_members），这里沿用同一套规则，否则列表上的引用数
    会和项目详情页的「常客」标识对不上。

    做成子查询而不是逐个统计：既能一次算完，也能直接用来做全局排序 + 分页。
    """
    from App_new.business.projects.models.project_member import ProjectMember
    from App_new.business.projects.models.project import ProjectHeader

    return db.session.query(
        ProjectMember.member_name.label('member_name'),
        db.func.count(db.distinct(ProjectMember.header_id)).label('order_count'),
        db.func.max(ProjectHeader.created_at).label('last_used_at'),
    ).join(
        ProjectHeader, ProjectHeader.id == ProjectMember.header_id
    ).group_by(ProjectMember.member_name).subquery()


def get_traveler_files_path(traveler_id):
    """获取旅客文件存储路径"""
    base_path = Path(os.getcwd()) / '资源' / 'Travelers' / str(traveler_id)
    base_path.mkdir(parents=True, exist_ok=True)
    return base_path

frequent_traveler_bp = Blueprint('frequent_traveler', __name__, url_prefix='/frequent_traveler')

from App_new.auth.models.auth import AuthUser, Role, UserProfile


def _get_staff_list():
    """员工列表（staff + admin 角色，启用中），用于旅客归属选择。"""
    role_ids = [r.id for r in Role.query.filter(Role.name.in_(['staff', 'admin'])).all()]
    if not role_ids:
        return []
    rows = db.session.query(
        AuthUser.id, AuthUser.username, UserProfile.first_name, UserProfile.last_name
    ).outerjoin(UserProfile, AuthUser.id == UserProfile.user_id).filter(
        AuthUser.role_id.in_(role_ids), AuthUser.is_active == True
    ).all()
    result = []
    for r in rows:
        name = f"{r.first_name or ''}{r.last_name or ''}".strip() or r.username
        result.append({'id': r.id, 'name': name})
    result.sort(key=lambda x: x['name'])
    return result


def _current_staff_level():
    """当前用户员工等级；非 staff 角色（管理员等）视为最高级，可见全部。"""
    if not (current_user.is_authenticated and current_user.role and current_user.role.name == 'staff'):
        return 99
    if current_user.profile:
        return current_user.profile.staff_level or 1
    return 1


def _can_access_traveler(traveler):
    """是否有权访问/编辑该旅客：无归属(共享) / 归属本人 / 2级及以上。"""
    if traveler.staff_id is None:
        return True
    if _current_staff_level() >= 2:
        return True
    return current_user.is_authenticated and traveler.staff_id == current_user.id


@frequent_traveler_bp.route('/')
@login_required
@staff_only
def list_page():
    """常用旅客管理页面"""
    companies = CustomerCompany.query.filter_by(is_customer=True).order_by(CustomerCompany.company_name).all()

    # 集团下拉：CustomerCompany.group_name + FrequentTraveler.group_name 的去重并集
    company_groups = {g[0] for g in db.session.query(CustomerCompany.group_name)
                      .filter(CustomerCompany.group_name.isnot(None),
                              CustomerCompany.group_name != '').distinct().all()}
    traveler_groups = {g[0] for g in db.session.query(FrequentTraveler.group_name)
                       .filter(FrequentTraveler.group_name.isnot(None),
                               FrequentTraveler.group_name != '').distinct().all()}
    groups = sorted(company_groups | traveler_groups)

    return render_template('business/projects/frequent_traveler.html',
                           companies=companies, groups=groups,
                           staff_list=_get_staff_list())


@frequent_traveler_bp.route('/api/list', methods=['GET'])
def api_list():
    """获取常用旅客列表（支持搜索和分页）"""
    keyword = request.args.get('keyword', '').strip()
    company_id = request.args.get('company_id', '', type=str)
    group_name = request.args.get('group_name', '', type=str).strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    query = FrequentTraveler.query

    # 归属过滤：1级员工只能看到自己归属 + 无归属(共享)；2级/管理员看全部
    if current_user.is_authenticated and _current_staff_level() < 2:
        query = query.filter(
            db.or_(
                FrequentTraveler.staff_id.is_(None),
                FrequentTraveler.staff_id == current_user.id
            )
        )

    if keyword:
        import re
        like = f'%{keyword}%'
        conds = [
            FrequentTraveler.name.ilike(like),
            FrequentTraveler.name_en.ilike(like),
            FrequentTraveler.phone.ilike(like),
            FrequentTraveler.passport_number.ilike(like),
            FrequentTraveler.email.ilike(like),
        ]
        # 编号搜索：T0042 / t42 / 42 → id=42
        code_match = re.match(r'^[Tt]?(\d{1,9})$', keyword)
        if code_match:
            conds.append(FrequentTraveler.id == int(code_match.group(1)))
        query = query.filter(db.or_(*conds))

    if company_id:
        query = query.filter_by(company_id=int(company_id))

    if group_name:
        # 匹配旅客自身 group_name，或其关联公司的 group_name
        query = query.outerjoin(CustomerCompany, FrequentTraveler.company_id == CustomerCompany.id).filter(
            db.or_(
                FrequentTraveler.group_name == group_name,
                CustomerCompany.group_name == group_name,
            )
        )

    # 引用统计：接到主查询上，这样「按引用数/最近引用排序」是全表排序后再分页，
    # 而不是只在当前页里排。
    usage = _usage_subquery()
    order_count = db.func.coalesce(usage.c.order_count, 0)
    last_used_at = usage.c.last_used_at

    query = query.outerjoin(usage, usage.c.member_name == FrequentTraveler.name)

    # 引用情况筛选：unused = 一次都没被订单引用过（就是可以考虑清理的那批）
    usage_filter = request.args.get('usage', '', type=str).strip()
    if usage_filter == 'unused':
        query = query.filter(order_count == 0)
    elif usage_filter == 'used':
        query = query.filter(order_count > 0)

    # 排序
    sort = request.args.get('sort', 'name', type=str)
    direction = request.args.get('dir', '', type=str).lower()
    sort_columns = {
        'name': FrequentTraveler.name,
        'orders': order_count,
        'last_used': last_used_at,
        'created': FrequentTraveler.created_at,
    }
    if sort not in sort_columns:
        sort = 'name'
    # 姓名默认升序，引用数/最近引用默认降序（最常用、最近用的排前面）
    if direction not in ('asc', 'desc'):
        direction = 'asc' if sort in ('name',) else 'desc'
    sort_col = sort_columns[sort]
    query = query.order_by(
        sort_col.asc() if direction == 'asc' else sort_col.desc(),
        FrequentTraveler.name.asc()
    )

    query = query.add_columns(order_count.label('order_count'), last_used_at.label('last_used_at'))
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    # 批量查询文件数量
    rows_items = list(pagination.items)
    travelers = [row[0] for row in rows_items]
    traveler_ids = [t.id for t in travelers]
    file_counts = {}
    if traveler_ids:
        rows = db.session.query(
            TravelerFile.traveler_id, db.func.count(TravelerFile.id)
        ).filter(TravelerFile.traveler_id.in_(traveler_ids)).group_by(TravelerFile.traveler_id).all()
        file_counts = {r[0]: r[1] for r in rows}

    result = []
    for row in rows_items:
        t, cnt, last_at = row[0], row[1], row[2]
        d = t.to_dict()
        d['file_count'] = file_counts.get(t.id, 0)
        d['order_count'] = int(cnt or 0)
        d['last_used_at'] = last_at.strftime('%Y-%m-%d') if last_at else None
        result.append(d)

    return jsonify({
        'success': True,
        'data': result,
        'total': pagination.total,
        'page': pagination.page,
        'pages': pagination.pages,
        'sort': sort,
        'dir': direction,
    })


@frequent_traveler_bp.route('/api/search', methods=['GET'])
def api_search():
    """搜索常用旅客（供项目人员选取用，轻量接口）。
    支持按编号查询：输入 T0042 / t42 / 42 都会匹配 id=42。
    """
    import re
    keyword = request.args.get('q', '').strip()
    if not keyword:
        return jsonify({'success': True, 'data': []})

    # 解析编号：T 前缀+数字 或 纯数字
    code_match = re.match(r'^[Tt]?(\d{1,9})$', keyword)
    by_id = None
    if code_match:
        by_id = FrequentTraveler.query.get(int(code_match.group(1)))

    like = f'%{keyword}%'
    text_results = FrequentTraveler.query.filter(
        db.or_(
            FrequentTraveler.name.ilike(like),
            FrequentTraveler.name_en.ilike(like),
            FrequentTraveler.phone.ilike(like),
            FrequentTraveler.passport_number.ilike(like),
        )
    ).order_by(FrequentTraveler.name).limit(20).all()

    # 按编号命中的优先放最前，并去重
    travelers = []
    seen_ids = set()
    if by_id:
        travelers.append(by_id)
        seen_ids.add(by_id.id)
    for t in text_results:
        if t.id not in seen_ids:
            travelers.append(t)
            seen_ids.add(t.id)

    return jsonify({
        'success': True,
        'data': [t.to_dict() for t in travelers]
    })


@frequent_traveler_bp.route('/api/check_duplicate', methods=['GET'])
def api_check_duplicate():
    """检查是否存在疑似重复的常用旅客（按姓名或护照号匹配）"""
    name = request.args.get('name', '').strip()
    passport = request.args.get('passport', '').strip()
    exclude_id = request.args.get('exclude_id', '', type=str)  # 编辑时排除自身

    if not name and not passport:
        return jsonify({'success': True, 'duplicates': []})

    conditions = []
    if name:
        conditions.append(FrequentTraveler.name == name)
    if passport:
        conditions.append(FrequentTraveler.passport_number == passport)

    query = FrequentTraveler.query.filter(db.or_(*conditions))
    if exclude_id:
        query = query.filter(FrequentTraveler.id != int(exclude_id))

    matches = query.limit(5).all()
    return jsonify({
        'success': True,
        'duplicates': [{
            'id': t.id,
            'title': t.title,
            'name': t.name,
            'phone': t.phone,
            'passport_number': t.passport_number,
            'company_name': t.company.company_name if t.company else None,
        } for t in matches]
    })


@frequent_traveler_bp.route('/api/create', methods=['POST'])
def api_create():
    """创建常用旅客"""
    data = request.get_json()
    if not data.get('name'):
        return jsonify({'success': False, 'message': '姓名不能为空'}), 400

    traveler = FrequentTraveler(
        title=data.get('title'),
        name=data.get('name'),
        name_en=data.get('name_en'),
        gender=data.get('gender'),
        date_of_birth=data.get('date_of_birth') or None,
        nationality=data.get('nationality'),
        phone=data.get('phone'),
        email=data.get('email'),
        passport_number=data.get('passport_number'),
        passport_issuing_country=data.get('passport_issuing_country'),
        passport_expiry_date=data.get('passport_expiry_date') or None,
        id_card_number=data.get('id_card_number'),
        airline_memberships=json.dumps(data.get('airline_memberships', []), ensure_ascii=False) if data.get('airline_memberships') else None,
        company_id=data.get('company_id') or None,
        group_name=(data.get('group_name') or '').strip() or None,
        remarks=data.get('remarks'),
        staff_id=data.get('staff_id') or None,
    )
    db.session.add(traveler)
    db.session.commit()

    return jsonify({'success': True, 'message': '常用旅客添加成功', 'data': traveler.to_dict()})


@frequent_traveler_bp.route('/api/<int:traveler_id>', methods=['GET'])
def api_get(traveler_id):
    """获取单个常用旅客详情"""
    traveler = FrequentTraveler.query.get_or_404(traveler_id)
    return jsonify({'success': True, 'data': traveler.to_dict()})


@frequent_traveler_bp.route('/api/<int:traveler_id>', methods=['PUT'])
def api_update(traveler_id):
    """更新常用旅客"""
    traveler = FrequentTraveler.query.get_or_404(traveler_id)
    if not _can_access_traveler(traveler):
        return jsonify({'success': False, 'message': '无权编辑该旅客（归属其他员工）'}), 403
    data = request.get_json()

    if 'name' in data and not data['name']:
        return jsonify({'success': False, 'message': '姓名不能为空'}), 400

    for field in ['title', 'name', 'name_en', 'gender', 'nationality',
                  'phone', 'email', 'passport_number', 'passport_issuing_country',
                  'id_card_number', 'remarks']:
        if field in data:
            setattr(traveler, field, data[field])

    if 'date_of_birth' in data:
        traveler.date_of_birth = data['date_of_birth'] or None
    if 'passport_expiry_date' in data:
        traveler.passport_expiry_date = data['passport_expiry_date'] or None
    if 'company_id' in data:
        traveler.company_id = data['company_id'] or None
    if 'group_name' in data:
        traveler.group_name = (data['group_name'] or '').strip() or None
    if 'staff_id' in data:
        traveler.staff_id = data['staff_id'] or None
    if 'airline_memberships' in data:
        traveler.airline_memberships = json.dumps(data['airline_memberships'], ensure_ascii=False) if data['airline_memberships'] else None

    db.session.commit()
    return jsonify({'success': True, 'message': '旅客信息已更新', 'data': traveler.to_dict()})


@frequent_traveler_bp.route('/api/<int:traveler_id>', methods=['DELETE'])
def api_delete(traveler_id):
    """删除常用旅客"""
    traveler = FrequentTraveler.query.get_or_404(traveler_id)
    if not _can_access_traveler(traveler):
        return jsonify({'success': False, 'message': '无权删除该旅客（归属其他员工）'}), 403
    name = traveler.name
    # 删除关联文件（磁盘 + 数据库）
    for f in traveler.files.all():
        try:
            if os.path.exists(f.file_path):
                os.remove(f.file_path)
        except OSError:
            pass
    db.session.delete(traveler)
    db.session.commit()
    return jsonify({'success': True, 'message': f'常用旅客 {name} 已删除'})


@frequent_traveler_bp.route('/api/bulk_delete', methods=['POST'])
@login_required
@staff_only
def api_bulk_delete():
    """批量删除常用旅客（用于清理长期没被订单引用的记录）

    - 无权限的（归属其他员工）跳过，不中断整批
    - 默认拒绝删除仍被订单引用的记录，除非显式传 force=true
    """
    data = request.get_json(silent=True) or {}
    raw_ids = data.get('ids') or []
    force = bool(data.get('force'))

    try:
        ids = [int(i) for i in raw_ids]
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'ids 参数不合法'}), 400
    if not ids:
        return jsonify({'success': False, 'message': '未选择要删除的旅客'}), 400

    travelers = FrequentTraveler.query.filter(FrequentTraveler.id.in_(ids)).all()
    if not travelers:
        return jsonify({'success': False, 'message': '未找到对应的旅客'}), 404

    # 引用数：按姓名匹配，口径与列表页一致
    usage = _usage_subquery()
    names = [t.name for t in travelers if t.name]
    used_counts = {}
    if names:
        rows = db.session.query(usage.c.member_name, usage.c.order_count).filter(
            usage.c.member_name.in_(names)
        ).all()
        used_counts = {r[0]: int(r[1] or 0) for r in rows}

    deleted, skipped_no_perm, skipped_in_use = [], [], []
    for t in travelers:
        if not _can_access_traveler(t):
            skipped_no_perm.append(t.name)
            continue
        if not force and used_counts.get(t.name, 0) > 0:
            skipped_in_use.append(f'{t.name}({used_counts[t.name]}单)')
            continue
        for f in t.files.all():
            try:
                if os.path.exists(f.file_path):
                    os.remove(f.file_path)
            except OSError:
                pass
        db.session.delete(t)
        deleted.append(t.name)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'删除失败: {e}'}), 500

    parts = [f'已删除 {len(deleted)} 条']
    if skipped_in_use:
        parts.append(f'跳过仍被订单引用的 {len(skipped_in_use)} 条: {", ".join(skipped_in_use[:5])}'
                     + ('...' if len(skipped_in_use) > 5 else ''))
    if skipped_no_perm:
        parts.append(f'跳过无权限的 {len(skipped_no_perm)} 条')

    return jsonify({
        'success': True,
        'message': '；'.join(parts),
        'deleted': len(deleted),
        'skipped_in_use': skipped_in_use,
        'skipped_no_perm': skipped_no_perm,
    })


# ==================== Excel 导入导出 ====================

# Excel列定义（与表头对应）
EXCEL_COLUMNS = [
    ('id', 'ID(勿改)', 10),
    ('title', '称谓', 8),
    ('name', '姓名*', 15),
    ('name_en', '英文名', 20),
    ('gender', '性别(male/female)', 12),
    ('date_of_birth', '出生日期(dd-mm-yyyy)', 20),
    ('nationality', '国籍', 12),
    ('phone', '电话', 18),
    ('email', '邮箱', 25),
    ('passport_number', '护照号码', 18),
    ('passport_issuing_country', '护照签发国', 12),
    ('passport_expiry_date', '护照有效期(dd-mm-yyyy)', 22),
    ('id_card_number', '身份证号码', 20),
    ('airline_memberships', '航空会员(航司:号码,多个用;分隔)', 30),
    ('company_name', '关联公司', 20),
    ('group_name', '集团', 15),
    ('remarks', '备注', 25),
]


# 导出/模板里日期列的显示格式（Excel 单元格格式，非字符串）
EXCEL_DATE_FORMAT = 'DD-MM-YYYY'
EXCEL_DATE_FIELDS = ('date_of_birth', 'passport_expiry_date')


@frequent_traveler_bp.route('/api/export', methods=['GET'])
def api_export():
    """导出全部常用旅客为Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    travelers = FrequentTraveler.query.order_by(FrequentTraveler.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '常用旅客'

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 写入表头
    for col_idx, (field, header, width) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # 写入数据
    for row_idx, t in enumerate(travelers, 2):
        d = t.to_dict()
        # 航空会员格式化: "航司:号码;航司:号码"
        memberships_str = ''
        if d['airline_memberships']:
            parts = []
            for m in d['airline_memberships']:
                airline = m.get('airline', '')
                number = m.get('number', '')
                if airline or number:
                    parts.append(f"{airline}:{number}")
            memberships_str = ';'.join(parts)

        row_data = [
            t.id,
            d['title'] or '',
            d['name'] or '',
            d['name_en'] or '',
            d['gender'] or '',
            t.date_of_birth,
            d['nationality'] or '',
            d['phone'] or '',
            d['email'] or '',
            d['passport_number'] or '',
            d['passport_issuing_country'] or '',
            t.passport_expiry_date,
            d['id_card_number'] or '',
            memberships_str,
            d['company_name'] or '',
            d['group_name'] or '',
            d['remarks'] or '',
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if EXCEL_COLUMNS[col_idx - 1][0] in EXCEL_DATE_FIELDS:
                cell.number_format = EXCEL_DATE_FORMAT
                cell.alignment = Alignment(horizontal='center')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name='常用旅客.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@frequent_traveler_bp.route('/api/template', methods=['GET'])
def api_template():
    """下载导入模板（空Excel）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = '常用旅客'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, (field, header, width) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # 称谓下拉
    dv_title = DataValidation(type='list', formula1='"MR,MS,MISS,MASTER"', allow_blank=True)
    dv_title.error = '请选择: MR/MS/MISS/MASTER'
    ws.add_data_validation(dv_title)
    dv_title.add(f'B2:B1000')

    # 性别下拉
    dv_gender = DataValidation(type='list', formula1='"male,female"', allow_blank=True)
    ws.add_data_validation(dv_gender)
    dv_gender.add(f'E2:E1000')

    # 示例行
    # 示例行的日期同样写真正的日期值，套 dd-mm-yyyy 单元格格式
    from datetime import date as _date
    example = ['', 'MR', '张三', 'ZHANG SAN', 'male', _date(1990, 1, 15), 'CHINA',
               '+65 91234567', 'zhangsan@email.com', 'E12345678', 'CHINA',
               _date(2030, 12, 31), '110101199001150011', 'SQ:1234567890;CX:987654',
               '', 'ALIBABA', '示例数据，请删除此行']
    example_font = Font(color='999999', italic=True)
    for col_idx, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = example_font
        cell.border = thin_border
        if EXCEL_COLUMNS[col_idx - 1][0] in EXCEL_DATE_FIELDS:
            cell.number_format = EXCEL_DATE_FORMAT
            cell.alignment = Alignment(horizontal='center')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name='常用旅客_导入模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@frequent_traveler_bp.route('/api/import', methods=['POST'])
@csrf.exempt
def api_import():
    """从Excel导入常用旅客（有ID更新，无ID新增）"""
    from openpyxl import load_workbook

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '请上传 .xlsx 格式的Excel文件'}), 400

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'success': False, 'message': f'无法读取Excel文件: {e}'}), 400

    # 读取表头，确定列映射
    headers = [cell.value for cell in ws[1]]
    col_map = {}  # field_name -> col_index
    for col_idx, (field, header, _) in enumerate(EXCEL_COLUMNS):
        # 通过表头名称匹配
        for h_idx, h_val in enumerate(headers):
            if h_val and header in str(h_val):
                col_map[field] = h_idx
                break

    if 'name' not in col_map:
        return jsonify({'success': False, 'message': 'Excel中未找到"姓名"列，请使用正确的模板'}), 400

    # 预加载公司名称映射
    companies = CustomerCompany.query.filter_by(is_customer=True).all()
    company_map = {c.company_name: c.id for c in companies}

    created = 0
    updated = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        def get_val(field):
            idx = col_map.get(field)
            if idx is not None and idx < len(row):
                v = row[idx]
                return str(v).strip() if v is not None else ''
            return ''

        name = get_val('name')
        if not name:
            skipped += 1
            continue

        # 解析航空会员: "SQ:123;CX:456" -> [{"airline":"SQ","number":"123"}, ...]
        memberships_str = get_val('airline_memberships')
        memberships = []
        if memberships_str:
            for part in memberships_str.split(';'):
                part = part.strip()
                if ':' in part:
                    airline, number = part.split(':', 1)
                    memberships.append({'airline': airline.strip(), 'number': number.strip()})
                elif part:
                    memberships.append({'airline': part, 'number': ''})

        # 解析关联公司
        company_name = get_val('company_name')
        company_id = company_map.get(company_name) if company_name else None

        # 解析日期字段
        def parse_date(val):
            if not val:
                return None
            # 如果是 datetime 对象直接取 date
            if hasattr(val, 'date'):
                return val.date() if hasattr(val, 'date') else val
            val = str(val).strip()
            if not val:
                return None
            from datetime import datetime as dt
            for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
                try:
                    return dt.strptime(val, fmt).date()
                except ValueError:
                    continue
            return None

        # 获取原始日期值（可能是 datetime 对象）
        dob_idx = col_map.get('date_of_birth')
        dob_raw = row[dob_idx] if dob_idx is not None and dob_idx < len(row) else None
        pe_idx = col_map.get('passport_expiry_date')
        pe_raw = row[pe_idx] if pe_idx is not None and pe_idx < len(row) else None

        traveler_data = {
            'title': get_val('title'),
            'name': name,
            'name_en': get_val('name_en'),
            'gender': get_val('gender') if get_val('gender') in ('male', 'female') else None,
            'date_of_birth': parse_date(dob_raw),
            'nationality': get_val('nationality'),
            'phone': get_val('phone'),
            'email': get_val('email'),
            'passport_number': get_val('passport_number'),
            'passport_issuing_country': get_val('passport_issuing_country'),
            'passport_expiry_date': parse_date(pe_raw),
            'id_card_number': get_val('id_card_number'),
            'airline_memberships': json.dumps(memberships, ensure_ascii=False) if memberships else None,
            'company_id': company_id,
            'group_name': get_val('group_name') or None,
            'remarks': get_val('remarks'),
        }

        # 判断更新还是新增
        traveler_id = get_val('id')
        try:
            traveler_id = int(traveler_id) if traveler_id else None
        except (ValueError, TypeError):
            traveler_id = None

        try:
            if traveler_id:
                traveler = FrequentTraveler.query.get(traveler_id)
                if traveler:
                    for k, v in traveler_data.items():
                        setattr(traveler, k, v)
                    updated += 1
                else:
                    # ID不存在，作为新增处理
                    traveler = FrequentTraveler(**traveler_data)
                    db.session.add(traveler)
                    created += 1
            else:
                traveler = FrequentTraveler(**traveler_data)
                db.session.add(traveler)
                created += 1
        except Exception as e:
            errors.append(f'第{row_idx}行: {e}')

    db.session.commit()

    msg = f'导入完成：新增 {created} 条，更新 {updated} 条'
    if skipped:
        msg += f'，跳过 {skipped} 条（无姓名）'
    if errors:
        msg += f'，{len(errors)} 条出错'

    return jsonify({
        'success': True,
        'message': msg,
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:10],
    })


# ==================== 文件管理 ====================

@frequent_traveler_bp.route('/api/<int:traveler_id>/files', methods=['GET'])
def api_files(traveler_id):
    """获取旅客的文件列表"""
    FrequentTraveler.query.get_or_404(traveler_id)
    files = TravelerFile.query.filter_by(traveler_id=traveler_id).order_by(TravelerFile.created_at.desc()).all()
    return jsonify({'success': True, 'data': [f.to_dict() for f in files]})


@frequent_traveler_bp.route('/api/<int:traveler_id>/files/upload', methods=['POST'])
@csrf.exempt
def api_upload_file(traveler_id):
    """上传旅客文件（支持批量，前端可传多个 file 字段）"""
    FrequentTraveler.query.get_or_404(traveler_id)

    files = request.files.getlist('file')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'success': False, 'message': '没有选择文件'}), 400

    file_category = request.form.get('file_category', 'other')
    description = request.form.get('description', '')
    file_dir = get_traveler_files_path(traveler_id)

    saved = []
    skipped = []
    for file in files:
        if file.filename == '':
            continue
        if not allowed_file(file.filename):
            skipped.append(file.filename)
            continue

        original_filename = secure_filename(file.filename) or file.filename
        file_ext = os.path.splitext(original_filename)[1].lower()
        stored_filename = f"{uuid.uuid4().hex}{file_ext}"

        full_path = file_dir / stored_filename
        file.save(str(full_path))
        file_size = os.path.getsize(str(full_path))

        record = TravelerFile(
            traveler_id=traveler_id,
            file_category=file_category,
            filename=file.filename,
            stored_filename=stored_filename,
            file_path=str(full_path),
            file_size=file_size,
            file_type=file.content_type,
            description=description,
        )
        db.session.add(record)
        saved.append(record)

    db.session.commit()

    msg = f'成功上传 {len(saved)} 个文件'
    if skipped:
        msg += f'，{len(skipped)} 个文件类型不支持已跳过'

    return jsonify({
        'success': True,
        'message': msg,
        'data': [r.to_dict() for r in saved],
    })


@frequent_traveler_bp.route('/api/<int:traveler_id>/files/download_all', methods=['GET'])
def api_download_all(traveler_id):
    """打包下载旅客的所有文件为 ZIP"""
    traveler = FrequentTraveler.query.get_or_404(traveler_id)
    files = TravelerFile.query.filter_by(traveler_id=traveler_id).all()

    if not files:
        return jsonify({'success': False, 'message': '该旅客暂无上传文件'}), 404

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        used_names = {}
        for f in files:
            if not os.path.exists(f.file_path):
                continue
            # 按分类建子文件夹，防止重名
            folder = f.category_display
            name = f.filename
            key = f"{folder}/{name}"
            if key in used_names:
                used_names[key] += 1
                base, ext = os.path.splitext(name)
                name = f"{base}_{used_names[key]}{ext}"
            else:
                used_names[key] = 1
            arcname = f"{folder}/{name}"
            zf.write(f.file_path, arcname)

    buf.seek(0)
    zip_name = f"{traveler.name}_资料.zip"
    return send_file(buf, as_attachment=True, download_name=zip_name, mimetype='application/zip')


@frequent_traveler_bp.route('/api/files/<int:file_id>/download', methods=['GET'])
def api_download_file(file_id):
    """下载/查看旅客文件"""
    record = TravelerFile.query.get_or_404(file_id)
    if not os.path.exists(record.file_path):
        return jsonify({'success': False, 'message': '文件不存在'}), 404

    return send_file(
        record.file_path,
        as_attachment=False,
        download_name=record.filename,
    )


@frequent_traveler_bp.route('/api/files/<int:file_id>/delete', methods=['DELETE'])
def api_delete_file(file_id):
    """删除旅客文件"""
    record = TravelerFile.query.get_or_404(file_id)
    # 删除磁盘文件
    try:
        if os.path.exists(record.file_path):
            os.remove(record.file_path)
    except OSError:
        pass
    db.session.delete(record)
    db.session.commit()
    return jsonify({'success': True, 'message': '文件已删除'})
