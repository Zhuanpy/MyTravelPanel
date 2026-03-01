# -*- coding: utf-8 -*-
"""
旅游产品管理蓝图
包含产品的CRUD操作
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import os
import json
from datetime import datetime
import pandas as pd
from io import BytesIO

from sqlalchemy import or_
from openpyxl.utils import get_column_letter

from App_new.exts import db, csrf
from App_new.business.tour.models.Packagemodels import Product, ProductItinerary, ProductPriceVariant, ImageLibrary
from App_new.business.projects.models.project import CustomerCompany
from App_new.utils.decorators import staff_only

# 创建蓝图
tour_products_bp = Blueprint('tour_products', __name__, url_prefix='/tour/products')

# 允许的图片扩展名
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
# 允许的文档扩展名
ALLOWED_DOC_EXTENSIONS = {'pdf', 'doc', 'docx'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_doc_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_DOC_EXTENSIONS


def save_uploaded_file(file, upload_folder='uploads/tour_products', compress=True):
    """
    保存上传的文件，返回相对路径（相对于 App_new/static）

    参数:
        file: 上传的文件对象
        upload_folder: 上传目录
        compress: 是否压缩图片（默认True）
    """
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        name, ext = os.path.splitext(filename)
        # 防止中文文件名被secure_filename清空后多个文件同名覆盖
        import uuid
        unique_id = uuid.uuid4().hex[:8]
        filename = f"{name}_{timestamp}_{unique_id}{ext}"

        upload_path = os.path.join('App_new/static', upload_folder)
        os.makedirs(upload_path, exist_ok=True)

        filepath = os.path.join(upload_path, filename)
        file.save(filepath)

        # 压缩图片
        if compress:
            compress_result = compress_and_resize_image(filepath, max_size=1200, quality=80)
            if compress_result and compress_result.get('new_path'):
                # 如果压缩后文件名改变了（如PNG转JPG），更新文件名
                new_filepath = compress_result.get('new_path')
                filename = os.path.basename(new_filepath)

        # 返回存入数据库使用的相对路径
        return os.path.join(upload_folder, filename).replace('\\', '/')
    return None


def _save_to_image_library(img_path, original_filename, product):
    """将行程图片同步保存到图片库，方便其它产品复用

    参数:
        img_path: 已保存的图片相对路径（相对于static）
        original_filename: 原始文件名（用作图片标题）
        product: Product对象（用于提取国家和城市标签）
    """
    try:
        # 用原始文件名（去扩展名）作为图片标题
        name_part = os.path.splitext(original_filename)[0]
        # 去除secure_filename可能残留的下划线和时间戳
        title = name_part.strip('_') if name_part else '行程图片'

        # 构建标签：国家 + 城市
        tag_parts = []
        if product.country and product.country != '未知':
            tag_parts.append(product.country)
        if product.city_name:
            tag_parts.append(product.city_name)
        tags = ','.join(tag_parts) if tag_parts else None

        # 获取文件信息
        full_path = os.path.join('App_new/static', img_path)
        file_size = os.path.getsize(full_path) if os.path.exists(full_path) else None
        width, height = None, None
        try:
            from PIL import Image
            with Image.open(full_path) as img:
                width, height = img.size
        except Exception:
            pass

        image_record = ImageLibrary(
            title=title,
            image_path=img_path,
            tags=tags,
            category='tour',
            file_size=file_size,
            width=width,
            height=height,
            is_active=True,
            created_by=current_user.username if current_user else None
        )
        db.session.add(image_record)
    except Exception as e:
        # 保存图片库失败不影响行程保存
        print(f"同步图片库失败: {e}")


def compress_and_resize_image(file_path, max_size=1200, quality=80):
    """
    压缩和调整图片尺寸（优化网页加载速度）

    参数:
        file_path: 图片文件路径
        max_size: 最大宽度或高度（像素），默认1200（适合网页显示）
        quality: JPEG压缩质量（1-100），默认80

    返回:
        dict: {'width': 宽度, 'height': 高度, 'file_size': 文件大小, 'new_path': 新路径}
    """
    from PIL import Image
    from flask import current_app

    result = {'width': None, 'height': None, 'file_size': None, 'new_path': file_path}

    try:
        with Image.open(file_path) as img:
            original_format = img.format
            original_size = os.path.getsize(file_path)

            # 处理 RGBA 模式（PNG透明图片）转换为 RGB
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            # 获取原始尺寸
            width, height = img.size

            # 判断是否需要调整尺寸
            needs_resize = width > max_size or height > max_size

            if needs_resize:
                if width > height:
                    new_width = max_size
                    new_height = int(height * (max_size / width))
                else:
                    new_height = max_size
                    new_width = int(width * (max_size / height))

                img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                width, height = new_width, new_height

            # 保存压缩后的图片
            name, ext = os.path.splitext(file_path)

            if original_format == 'PNG' and original_size < 200 * 1024 and not needs_resize:
                img.save(file_path, 'PNG', optimize=True)
            else:
                new_path = name + '.jpg'
                img.save(new_path, 'JPEG', quality=quality, optimize=True)

                if file_path.lower() != new_path.lower() and os.path.exists(file_path):
                    os.remove(file_path)
                file_path = new_path

            result['width'] = width
            result['height'] = height
            result['file_size'] = os.path.getsize(file_path)
            result['new_path'] = file_path

            # 记录压缩效果
            if original_size > 0:
                compression_ratio = (1 - result['file_size'] / original_size) * 100
                current_app.logger.info(
                    f"图片压缩: {os.path.basename(file_path)}, "
                    f"原始: {original_size/1024:.1f}KB, "
                    f"压缩后: {result['file_size']/1024:.1f}KB, "
                    f"压缩率: {compression_ratio:.1f}%"
                )

    except Exception as e:
        current_app.logger.error(f"图片压缩失败: {e}")

    return result


def save_document_file(file, upload_folder='uploads/tour_documents'):
    """保存上传的文档文件，返回相对路径"""
    if file and allowed_doc_file(file.filename):
        original_filename = file.filename
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        name, ext = os.path.splitext(filename)
        filename = f"{name}_{timestamp}{ext}"

        upload_path = os.path.join('App_new/static', upload_folder)
        os.makedirs(upload_path, exist_ok=True)

        filepath = os.path.join(upload_path, filename)
        file.save(filepath)

        # 返回存入数据库使用的相对路径和原始文件名
        return os.path.join(upload_folder, filename).replace('\\', '/'), original_filename
    return None, None


@tour_products_bp.route('/')
@tour_products_bp.route('/list')
@login_required
@staff_only
def product_list():
    """产品列表页面 - 支持筛选、搜索和分页"""
    # 获取筛选参数
    supplier_id = request.args.get('supplier', type=int)
    country = request.args.get('country', '')
    city = request.args.get('city', '')
    status = request.args.get('status', '')
    keyword = request.args.get('keyword', '')
    
    # 获取分页参数
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)  # 每页20个产品

    query = Product.query

    if supplier_id:
        query = query.filter(Product.supplier_id == supplier_id)
    if country:
        # 通过 city 关联筛选国家
        from App_new.business.tour.models.Packagemodels import ProductCity
        query = query.join(ProductCity, Product.city_id == ProductCity.id).filter(ProductCity.country_name == country)
    if city:
        query = query.filter(Product.city_name == city)
    if status:
        if status == 'expired':
            from datetime import date as date_cls
            query = query.filter(Product.valid_until.isnot(None), Product.valid_until < date_cls.today())
        else:
            query = query.filter(Product.product_status == status)
    if keyword:
        query = query.filter(
            or_(
                Product.product_name.like(f'%{keyword}%'),
                Product.product_description.like(f'%{keyword}%')
            )
        )

    # 执行分页查询
    pagination = query.order_by(Product.created_at.desc()).paginate(
        page=page, 
        per_page=per_page, 
        error_out=False
    )
    
    products = pagination.items

    # 全局检查：过期产品自动下架
    from datetime import date
    today = date.today()
    expired_active = Product.query.filter(
        Product.product_status == 'active',
        Product.valid_until.isnot(None),
        Product.valid_until < today
    ).all()
    auto_delisted = []
    for p in expired_active:
        p.product_status = 'inactive'
        auto_delisted.append(p.product_name)
    if auto_delisted:
        try:
            db.session.commit()
            flash(f'以下 {len(auto_delisted)} 个产品已过期，已自动下架：{"、".join(auto_delisted)}', 'warning')
        except Exception:
            db.session.rollback()

    # 添加产品显示属性
    for product in products:
        if product.supplier:
            product.supplier_display_name = product.supplier.company_name
        else:
            product.supplier_display_name = '未指定供应商'

        # 判断是否过期
        product.is_expired = (product.valid_until and product.valid_until < today)
        
        # 确保所有字段都不显示 None
        if product.product_description == 'None':
            product.product_description = None
        if product.country == 'None':
            product.country = '未知'
        if product.city_name == 'None':
            product.city_name = None
        if product.departure_city == 'None':
            product.departure_city = None
        if product.destination_city == 'None':
            product.destination_city = None

    # 通过 business_types 关联查询旅游相关的供应商类型
    from App_new.shared.models.business_types import BusinessType
    tour_type_codes = ['tour', 'tour_package', 'land_tour']
    tour_type_ids = [bt.id for bt in BusinessType.query.filter(
        BusinessType.code.in_(tour_type_codes),
        BusinessType.is_active == True
    ).all()]
    
    suppliers = CustomerCompany.query.filter(
        CustomerCompany.is_supplier == True,
        CustomerCompany.supplier_type_id.in_(tour_type_ids) if tour_type_ids else True
    ).order_by(CustomerCompany.company_name).all()

    # 从 ProductCity 表获取国家列表
    from App_new.business.tour.models.Packagemodels import ProductCity
    countries = db.session.query(ProductCity.country_name).filter(
        ProductCity.country_name.isnot(None)
    ).distinct().order_by(ProductCity.country_name).all()
    countries = [c[0] for c in countries if c[0]]

    cities = db.session.query(Product.city_name).filter(
        Product.city_name.isnot(None)
    ).distinct().order_by(Product.city_name).all()
    cities = [c[0] for c in cities if c[0]]

    # 统计数据
    total_products = Product.query.count()
    active_products = Product.query.filter(Product.product_status == 'active').count()
    draft_products = Product.query.filter(Product.product_status == 'draft').count()
    inactive_products = Product.query.filter(Product.product_status == 'inactive').count()
    disabled_products = Product.query.filter(Product.product_status == 'disabled').count()
    expired_products = Product.query.filter(Product.valid_until.isnot(None), Product.valid_until < today).count()

    return render_template('business/tour/products/product_list.html',
                         products=products,
                         pagination=pagination,
                         suppliers=suppliers,
                         countries=countries,
                         cities=cities,
                         current_filters={
                             'supplier': supplier_id,
                             'country': country,
                             'city': city,
                             'status': status,
                             'keyword': keyword
                         },
                         stats={
                             'total': total_products,
                             'active': active_products,
                             'draft': draft_products,
                             'inactive': inactive_products,
                             'disabled': disabled_products,
                             'expired': expired_products
                         })


@tour_products_bp.route('/add', methods=['GET', 'POST'])
@login_required
@staff_only
def add_product():
    """添加产品"""
    if request.method == 'POST':
        try:
            tags_input = request.form.get('tags', '').strip()
            tags_list = [tag.strip() for tag in tags_input.split(',') if tag.strip()]
            tags_json = json.dumps(tags_list, ensure_ascii=False)

            # 处理封面图：优先使用上传的文件，其次使用图片库选择的
            cover_image_path = None
            cover_file = request.files.get('cover_image')
            if cover_file and cover_file.filename:
                cover_image_path = save_uploaded_file(cover_file)

            # 如果没有上传新文件，检查是否从图片库选择
            if not cover_image_path and request.form.get('selected_cover_image'):
                cover_image_path = request.form.get('selected_cover_image')

            # 处理图片库：合并上传的和从图片库选择的
            gallery_paths = []
            # 处理上传的文件
            if 'gallery_images' in request.files:
                gallery_files = request.files.getlist('gallery_images')
                for file in gallery_files:
                    if file and file.filename:
                        path = save_uploaded_file(file)
                        if path:
                            gallery_paths.append(path)
            # 处理从图片库选择的图片
            if 'selected_gallery_images' in request.form and request.form.get('selected_gallery_images'):
                try:
                    selected_images = json.loads(request.form.get('selected_gallery_images'))
                    gallery_paths.extend(selected_images)
                except:
                    pass
            gallery_json = json.dumps(gallery_paths) if gallery_paths else None

            # 供应商文件上传
            supplier_doc_path = None
            supplier_doc_name = None
            if 'supplier_document' in request.files:
                doc_file = request.files['supplier_document']
                if doc_file and doc_file.filename:
                    supplier_doc_path, supplier_doc_name = save_document_file(doc_file)

            # 解析并转换表单字段（做了类型保护）
            supplier_id = request.form.get('supplier_id') or None
            if supplier_id:
                try:
                    supplier_id = int(supplier_id)
                except ValueError:
                    supplier_id = None

            # 处理 city_id（从 city_name 获取，如果城市不存在则自动创建）
            city_name = request.form.get('city_name')
            city_id = None
            if city_name:
                from App_new.business.tour.models.Packagemodels import ProductCity
                city = ProductCity.query.filter_by(city_name=city_name).first()
                if not city:
                    # 自动创建城市记录
                    country_name = request.form.get('country_name') or '未知'
                    city = ProductCity(
                        city_name=city_name,
                        display_name=city_name,
                        country_name=country_name
                    )
                    db.session.add(city)
                    db.session.flush()  # 获取新创建的城市ID
                city_id = city.id

            # 安全转换数字字段
            duration_days = request.form.get('duration_days', '').strip()
            duration_days_val = int(duration_days) if duration_days else None
            
            min_pax = request.form.get('min_pax', '1').strip()
            min_pax_val = int(min_pax) if min_pax else 1
            
            max_pax = request.form.get('max_pax', '').strip()
            max_pax_val = int(max_pax) if max_pax else None
            
            base_price = request.form.get('base_price', '').strip()
            base_price_val = float(base_price) if base_price else None
            
            child_price = request.form.get('child_price', '').strip()
            child_price_val = float(child_price) if child_price else None
            
            infant_price = request.form.get('infant_price', '').strip()
            infant_price_val = float(infant_price) if infant_price else None
            
            single_supplement = request.form.get('single_room_supplement', '').strip()
            single_supplement_val = float(single_supplement) if single_supplement else None
            
            is_featured = request.form.get('is_featured', '0').strip()
            is_featured_val = bool(int(is_featured)) if is_featured else False
            
            valid_from = request.form.get('valid_from', '').strip()
            valid_from_val = datetime.strptime(valid_from, '%Y-%m-%d').date() if valid_from else None
            
            valid_until = request.form.get('valid_until', '').strip()
            valid_until_val = datetime.strptime(valid_until, '%Y-%m-%d').date() if valid_until else None

            # 自动生成 product_code（使用统一编号规则）
            from App_new.shared.models.business_types import generate_product_code
            product_code = generate_product_code('tour')

            product = Product(
                product_name=request.form['product_name'],
                supplier_id=supplier_id,
                product_code=product_code,
                city_id=city_id,
                city_name=city_name,
                departure_city=request.form.get('departure_city') or None,
                destination_city=request.form.get('destination_city') or None,
                product_type=request.form.get('product_type') or None,
                duration_days=duration_days_val,
                min_pax=min_pax_val,
                max_pax=max_pax_val,
                base_price=base_price_val,
                child_price=child_price_val,
                infant_price=infant_price_val,
                single_room_supplement=single_supplement_val,
                currency=request.form.get('currency', 'SGD'),
                product_description=request.form.get('product_description') or None,
                included_services=request.form.get('included_services') or None,
                excluded_services=request.form.get('excluded_services') or None,
                important_notes=request.form.get('important_notes') or None,
                suitable_season=request.form.get('suitable_season') or None,
                difficulty_level=request.form.get('difficulty_level') or None,
                tags=tags_json,
                cover_image=cover_image_path,
                gallery_images=gallery_json,
                supplier_document=supplier_doc_path,
                supplier_document_name=supplier_doc_name,
                product_status=request.form.get('product_status', 'draft'),
                is_featured=is_featured_val,
                valid_from=valid_from_val,
                valid_until=valid_until_val,
                created_by=current_user.username
            )

            db.session.add(product)
            db.session.flush()  # 获取product.id

            # 同步到统一产品表
            from App_new.business.products.sync_helper import sync_tour_product_to_unified
            sync_tour_product_to_unified(product, created_by=current_user.username)

            db.session.commit()

            flash('产品创建成功！', 'success')
            return redirect(url_for('tour_products.product_detail', product_id=product.id))

        except Exception as e:
            db.session.rollback()
            import traceback
            traceback.print_exc()
            flash(f'创建失败：{str(e)}', 'error')

    # 通过 business_types 关联查询旅游相关的供应商类型
    from App_new.shared.models.business_types import BusinessType
    tour_type_codes = ['tour', 'tour_package', 'land_tour']
    tour_type_ids = [bt.id for bt in BusinessType.query.filter(
        BusinessType.code.in_(tour_type_codes),
        BusinessType.is_active == True
    ).all()]
    
    suppliers = CustomerCompany.query.filter(
        CustomerCompany.is_supplier == True,
        CustomerCompany.supplier_type_id.in_(tour_type_ids) if tour_type_ids else True
    ).order_by(CustomerCompany.company_name).all()

    # 获取城市列表
    from App_new.business.tour.models.Packagemodels import ProductCity
    cities = ProductCity.query.order_by(ProductCity.country_name, ProductCity.city_name).all()

    return render_template('business/tour/products/product_form.html',
                         product=None,
                         suppliers=suppliers,
                         cities=cities,
                         itineraries=[])


@tour_products_bp.route('/<int:product_id>')
@login_required
@staff_only
def product_detail(product_id):
    """产品详情页"""
    product = Product.query.get_or_404(product_id)
    itineraries = ProductItinerary.query.filter_by(product_id=product_id).order_by(ProductItinerary.day_number).all()

    return render_template('business/tour/products/product_detail.html',
                         product=product,
                         itineraries=itineraries)


@tour_products_bp.route('/<int:product_id>/print')
@login_required
@staff_only
def print_itinerary(product_id):
    """打印产品行程单"""
    from App_new.business.tour.models.Packagemodels import CompanyInfo
    
    product = Product.query.get_or_404(product_id)
    itineraries = ProductItinerary.query.filter_by(product_id=product_id).order_by(ProductItinerary.day_number).all()
    company = CompanyInfo.query.first()
    current_time = datetime.now()
    
    return render_template('business/tour/products/product_print_itinerary.html',
                         product=product,
                         itineraries=itineraries,
                         company=company,
                         current_time=current_time)


@tour_products_bp.route('/<int:product_id>/delete-file', methods=['POST'])
@login_required
@staff_only
@csrf.exempt
def delete_product_file(product_id):
    """移除产品关联文件（仅清除引用，不删除物理文件，物理文件通过图片库管理删除）"""
    try:
        product = Product.query.get_or_404(product_id)
        file_type = request.json.get('file_type')
        file_index = request.json.get('file_index')  # 用于图库图片

        if file_type == 'cover_image':
            if product.cover_image:
                product.cover_image = None
                db.session.commit()
                return jsonify({'success': True, 'message': '封面图已移除'})

        elif file_type == 'gallery_image':
            if product.gallery_images:
                images = json.loads(product.gallery_images) if isinstance(product.gallery_images, str) else product.gallery_images
                if file_index is not None and 0 <= file_index < len(images):
                    images.pop(file_index)
                    product.gallery_images = json.dumps(images) if images else None
                    db.session.commit()
                    return jsonify({'success': True, 'message': '图片已移除'})

        elif file_type == 'supplier_document':
            if product.supplier_document:
                product.supplier_document = None
                product.supplier_document_name = None
                db.session.commit()
                return jsonify({'success': True, 'message': '供应商文件已移除'})

        return jsonify({'success': False, 'message': '文件类型无效或文件不存在'}), 400

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/<int:product_id>/clear-gallery', methods=['POST'])
@login_required
@staff_only
@csrf.exempt
def clear_gallery_images(product_id):
    """清除产品的所有图库图片"""
    try:
        product = Product.query.get_or_404(product_id)

        if not product.gallery_images:
            return jsonify({'success': False, 'message': '没有可清除的图片'}), 400

        # 解析图片列表
        images = json.loads(product.gallery_images) if isinstance(product.gallery_images, str) else product.gallery_images

        # 删除所有图片文件
        deleted_count = 0
        for image_path in images:
            if image_path:
                from flask import current_app
                full_path = os.path.join(current_app.static_folder, image_path)
                if os.path.exists(full_path):
                    try:
                        os.remove(full_path)
                        deleted_count += 1
                    except Exception as e:
                        print(f"删除文件失败: {full_path}, 错误: {e}")

        # 清空数据库中的图片列表
        product.gallery_images = None
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'已清除 {deleted_count} 张图片',
            'deleted_count': deleted_count
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/<int:product_id>/itinerary/<int:itinerary_id>')
@login_required
@staff_only
def get_itinerary(product_id, itinerary_id):
    try:
        itinerary = ProductItinerary.query.get_or_404(itinerary_id)
        if itinerary.product_id != product_id:
            return jsonify({'success': False, 'message': '行程不属于该产品'}), 400
        return jsonify({'success': True, 'itinerary': itinerary.to_dict()})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/<int:product_id>/itinerary/add', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def add_itinerary(product_id):
    """添加行程 - API"""
    try:
        # 添加调试日志
        print(f"=== 添加行程 ===")
        print(f"Product ID: {product_id}")
        print(f"Form data: {dict(request.form)}")
        print(f"Files: {list(request.files.keys())}")
        
        product = Product.query.get_or_404(product_id)

        # 获取表单数据，使用 get 方法防止 KeyError
        day_number = request.form.get('day_number')
        day_title = request.form.get('day_title')
        content = request.form.get('content')
        
        if not day_number:
            return jsonify({'success': False, 'message': '天数不能为空'}), 400
        
        itinerary = ProductItinerary(
            product_id=product_id,
            day_number=int(day_number),
            day_title=day_title,
            content=content
        )

        for i in range(1, 4):
            img_field = f'image{i}'
            if img_field in request.files:
                img_file = request.files[img_field]
                if img_file and img_file.filename:
                    original_name = img_file.filename
                    img_path = save_uploaded_file(img_file, upload_folder='uploads/tour_itinerary')
                    if img_path:
                        setattr(itinerary, img_field, img_path)
                        # 同步保存到图片库
                        _save_to_image_library(img_path, original_name, product)

        db.session.add(itinerary)
        db.session.commit()

        print(f"✅ 行程添加成功！ID: {itinerary.id}")
        return jsonify({'success': True, 'message': '行程添加成功！'})

    except Exception as e:
        db.session.rollback()
        print(f"❌ 添加行程失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'保存失败：{str(e)}'}), 500


@tour_products_bp.route('/<int:product_id>/itinerary/<int:itinerary_id>/update', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def update_itinerary(product_id, itinerary_id):
    try:
        # 添加调试日志
        print(f"=== 更新行程 ===")
        print(f"Product ID: {product_id}, Itinerary ID: {itinerary_id}")
        print(f"Form data: {dict(request.form)}")
        print(f"Files: {list(request.files.keys())}")
        
        itinerary = ProductItinerary.query.get_or_404(itinerary_id)
        if itinerary.product_id != product_id:
            return jsonify({'success': False, 'message': '行程不属于该产品'}), 400

        # 获取表单数据，使用 get 方法防止 KeyError
        day_number = request.form.get('day_number')
        day_title = request.form.get('day_title')
        content = request.form.get('content')
        
        if not day_number:
            return jsonify({'success': False, 'message': '天数不能为空'}), 400
        
        itinerary.day_number = int(day_number)
        itinerary.day_title = day_title
        itinerary.content = content

        product = Product.query.get(product_id)
        for i in range(1, 4):
            img_field = f'image{i}'
            if img_field in request.files:
                img_file = request.files[img_field]
                if img_file and img_file.filename:
                    original_name = img_file.filename
                    img_path = save_uploaded_file(img_file, upload_folder='uploads/tour_itinerary')
                    if img_path:
                        setattr(itinerary, img_field, img_path)
                        # 同步保存到图片库
                        _save_to_image_library(img_path, original_name, product)

        itinerary.updated_at = datetime.utcnow()
        db.session.commit()

        print(f"✅ 行程更新成功！ID: {itinerary.id}")
        return jsonify({'success': True, 'message': '行程更新成功！'})

    except Exception as e:
        db.session.rollback()
        print(f"❌ 更新行程失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'保存失败：{str(e)}'}), 500


@tour_products_bp.route('/<int:product_id>/itinerary/<int:itinerary_id>/delete', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def delete_itinerary(product_id, itinerary_id):
    try:
        itinerary = ProductItinerary.query.get_or_404(itinerary_id)
        if itinerary.product_id != product_id:
            return jsonify({'success': False, 'message': '行程不属于该产品'}), 400
        db.session.delete(itinerary)
        db.session.commit()
        return jsonify({'success': True, 'message': '行程删除成功！'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/<int:product_id>/itinerary/auto-assign-images', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def auto_assign_itinerary_images(product_id):
    """自动从产品图片库分配图片给每日行程"""
    try:
        product = Product.query.get_or_404(product_id)
        
        # 获取产品图片库
        gallery_images = []
        if product.gallery_images:
            try:
                gallery_images = json.loads(product.gallery_images) if isinstance(product.gallery_images, str) else product.gallery_images
            except:
                gallery_images = []
        
        if not gallery_images:
            return jsonify({'success': False, 'message': '产品图片库为空，请先上传产品图片'}), 400
        
        # 获取所有行程，按天数排序
        itineraries = ProductItinerary.query.filter_by(product_id=product_id).order_by(ProductItinerary.day_number).all()
        
        if not itineraries:
            return jsonify({'success': False, 'message': '暂无行程，请先添加行程'}), 400
        
        # 自动分配图片：每天最多3张，按顺序分配
        image_index = 0
        updated_count = 0
        
        for itinerary in itineraries:
            images_for_day = []
            # 每天最多分配3张图片
            for i in range(3):
                if image_index < len(gallery_images):
                    images_for_day.append(gallery_images[image_index])
                    image_index += 1
                else:
                    break
            
            # 更新行程图片
            if images_for_day:
                itinerary.image1 = images_for_day[0] if len(images_for_day) > 0 else None
                itinerary.image2 = images_for_day[1] if len(images_for_day) > 1 else None
                itinerary.image3 = images_for_day[2] if len(images_for_day) > 2 else None
                updated_count += 1
        
        db.session.commit()
        
        total_images = len(gallery_images)
        total_days = len(itineraries)
        max_needed = total_days * 3
        
        message = f'已为 {updated_count} 天行程分配图片！'
        if total_images < max_needed:
            message += f'（产品图片库共 {total_images} 张，{total_days} 天行程最多需要 {max_needed} 张）'
        
        return jsonify({'success': True, 'message': message})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ 自动分配图片失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'分配失败：{str(e)}'}), 500


@tour_products_bp.route('/itinerary/template/download')
@login_required
@staff_only
def download_itinerary_template():
    """下载行程导入模板"""
    try:
        # 创建模板数据
        template_data = {
            '天数': [1, 2, 3],
            '标题': ['抵达目的地', '景点游览', '返程'],
            '行程内容': [
                '抵达机场，接机前往酒店办理入住，自由活动',
                '早餐后前往著名景点参观，午餐后继续游览，晚上返回酒店休息',
                '酒店早餐后退房，前往机场，结束愉快旅程'
            ]
        }
        
        df = pd.DataFrame(template_data)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='行程模板')
            worksheet = writer.sheets['行程模板']
            # 设置列宽
            worksheet.column_dimensions['A'].width = 10
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 80
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='行程导入模板.xlsx'
        )
    except Exception as e:
        flash(f'下载模板失败：{str(e)}', 'danger')
        return redirect(request.referrer or url_for('tour_products.product_list'))


@tour_products_bp.route('/<int:product_id>/itinerary/import-excel', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def import_itinerary_excel(product_id):
    """从Excel导入行程数据"""
    try:
        product = Product.query.get_or_404(product_id)
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '请选择要上传的文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '未选择文件'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': '只支持Excel文件'}), 400
        
        df = pd.read_excel(file, engine='openpyxl')
        
        # 检查必需列
        required_columns = ['天数']
        for col in required_columns:
            if col not in df.columns:
                return jsonify({'success': False, 'message': f'缺少必需列：{col}'}), 400
        
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                day_number = row.get('天数')
                if pd.isna(day_number):
                    continue
                
                day_number = int(day_number)
                day_title = str(row.get('标题', '')).strip() if pd.notna(row.get('标题')) else f'第{day_number}天'
                content = str(row.get('行程内容', '')).strip() if pd.notna(row.get('行程内容')) else ''
                
                # 检查是否已存在相同天数的行程
                existing = ProductItinerary.query.filter_by(
                    product_id=product_id,
                    day_number=day_number
                ).first()
                
                if existing:
                    # 更新现有行程
                    existing.day_title = day_title
                    existing.content = content
                else:
                    # 创建新行程
                    itinerary = ProductItinerary(
                        product_id=product_id,
                        day_number=day_number,
                        day_title=day_title,
                        content=content
                    )
                    db.session.add(itinerary)
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f'第{index + 2}行导入失败：{str(e)}')
        
        if imported_count > 0:
            db.session.commit()
        
        message = f'成功导入 {imported_count} 个行程'
        if errors:
            message += f'，{len(errors)} 个失败'
        
        return jsonify({
            'success': True,
            'message': message,
            'imported_count': imported_count,
            'errors': errors
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/<int:product_id>/edit', methods=['GET', 'POST'])
@login_required
@staff_only
def edit_product(product_id):
    product = Product.query.get_or_404(product_id)

    if request.method == 'POST':
        try:
            tags_input = request.form.get('tags', '').strip()
            tags_list = [tag.strip() for tag in tags_input.split(',') if tag.strip()]
            tags_json = json.dumps(tags_list, ensure_ascii=False)

            # 处理封面图：优先使用上传的文件，其次使用图片库选择的
            cover_updated = False
            cover_file = request.files.get('cover_image')
            if cover_file and cover_file.filename:
                cover_image_path = save_uploaded_file(cover_file)
                if cover_image_path:
                    product.cover_image = cover_image_path
                    cover_updated = True

            # 如果没有上传新文件，检查是否从图片库选择
            if not cover_updated and request.form.get('selected_cover_image'):
                product.cover_image = request.form.get('selected_cover_image')

            # 处理图片库：合并上传的和从图片库选择的
            gallery_paths = []
            # 保留现有图片（如果用户没有修改）
            if product.gallery_images:
                try:
                    existing_images = json.loads(product.gallery_images) if isinstance(product.gallery_images, str) else product.gallery_images
                    gallery_paths.extend(existing_images)
                except:
                    pass
            
            # 处理新上传的文件
            if 'gallery_images' in request.files:
                gallery_files = request.files.getlist('gallery_images')
                for file in gallery_files:
                    if file and file.filename:
                        path = save_uploaded_file(file)
                        if path:
                            gallery_paths.append(path)
            
            # 处理从图片库选择的图片
            if 'selected_gallery_images' in request.form and request.form.get('selected_gallery_images'):
                try:
                    selected_images = json.loads(request.form.get('selected_gallery_images'))
                    # 去重并添加新选择的图片
                    for img in selected_images:
                        if img not in gallery_paths:
                            gallery_paths.append(img)
                except:
                    pass
            
            if gallery_paths:
                product.gallery_images = json.dumps(gallery_paths)

            # 供应商文件上传
            if 'supplier_document' in request.files:
                doc_file = request.files['supplier_document']
                if doc_file and doc_file.filename:
                    supplier_doc_path, supplier_doc_name = save_document_file(doc_file)
                    if supplier_doc_path:
                        product.supplier_document = supplier_doc_path
                        product.supplier_document_name = supplier_doc_name

            supplier_id = request.form.get('supplier_id') or None
            if supplier_id:
                try:
                    supplier_id = int(supplier_id)
                except ValueError:
                    supplier_id = None

            # 处理 city_id（从 city_name 获取，如果城市不存在则自动创建）
            city_name = request.form.get('city_name')
            city_id = None
            if city_name:
                from App_new.business.tour.models.Packagemodels import ProductCity
                city = ProductCity.query.filter_by(city_name=city_name).first()
                if not city:
                    # 自动创建城市记录
                    country_name = request.form.get('country_name') or '未知'
                    city = ProductCity(
                        city_name=city_name,
                        display_name=city_name,
                        country_name=country_name
                    )
                    db.session.add(city)
                    db.session.flush()
                city_id = city.id

            product.product_name = request.form['product_name']
            product.supplier_id = supplier_id
            product.product_code = request.form.get('product_code') or None
            product.city_id = city_id
            product.city_name = city_name
            product.departure_city = request.form.get('departure_city') or None
            product.destination_city = request.form.get('destination_city') or None
            product.product_type = request.form.get('product_type') or None
            
            # 安全转换数字字段
            duration_days = request.form.get('duration_days', '').strip()
            product.duration_days = int(duration_days) if duration_days else None
            
            min_pax = request.form.get('min_pax', '1').strip()
            product.min_pax = int(min_pax) if min_pax else 1
            
            max_pax = request.form.get('max_pax', '').strip()
            product.max_pax = int(max_pax) if max_pax else None
            
            base_price = request.form.get('base_price', '').strip()
            product.base_price = float(base_price) if base_price else None
            
            child_price = request.form.get('child_price', '').strip()
            product.child_price = float(child_price) if child_price else None
            
            infant_price = request.form.get('infant_price', '').strip()
            product.infant_price = float(infant_price) if infant_price else None
            
            single_supplement = request.form.get('single_room_supplement', '').strip()
            product.single_room_supplement = float(single_supplement) if single_supplement else None
            
            product.currency = request.form.get('currency', 'SGD')
            product.product_description = request.form.get('product_description') or None
            product.included_services = request.form.get('included_services') or None
            product.excluded_services = request.form.get('excluded_services') or None
            product.important_notes = request.form.get('important_notes') or None
            product.suitable_season = request.form.get('suitable_season') or None
            product.difficulty_level = request.form.get('difficulty_level') or None
            product.tags = tags_json
            product.product_status = request.form.get('product_status', 'draft')
            
            # 安全转换 is_featured
            is_featured = request.form.get('is_featured', '0').strip()
            product.is_featured = bool(int(is_featured)) if is_featured else False
            
            # 安全转换日期字段
            valid_from = request.form.get('valid_from', '').strip()
            product.valid_from = datetime.strptime(valid_from, '%Y-%m-%d').date() if valid_from else None
            
            valid_until = request.form.get('valid_until', '').strip()
            product.valid_until = datetime.strptime(valid_until, '%Y-%m-%d').date() if valid_until else None
            product.updated_at = datetime.utcnow()

            # 同步到统一产品表
            from App_new.business.products.sync_helper import sync_tour_product_to_unified
            sync_tour_product_to_unified(product, created_by=current_user.username)

            db.session.commit()
            return redirect(url_for('tour_products.product_list'))

        except Exception as e:
            db.session.rollback()
            import traceback
            traceback.print_exc()
            flash(f'更新失败：{str(e)}', 'error')

    # 通过 business_types 关联查询旅游相关的供应商类型
    from App_new.shared.models.business_types import BusinessType
    tour_type_codes = ['tour', 'tour_package', 'land_tour']
    tour_type_ids = [bt.id for bt in BusinessType.query.filter(
        BusinessType.code.in_(tour_type_codes),
        BusinessType.is_active == True
    ).all()]
    
    suppliers = CustomerCompany.query.filter(
        CustomerCompany.is_supplier == True,
        CustomerCompany.supplier_type_id.in_(tour_type_ids) if tour_type_ids else True
    ).order_by(CustomerCompany.company_name).all()

    itineraries = ProductItinerary.query.filter_by(product_id=product_id).order_by(ProductItinerary.day_number).all()
    price_variants = ProductPriceVariant.query.filter_by(product_id=product_id).all()

    # 批量查询行程图片对应的图片库信息（标题、标签）
    all_image_paths = []
    for it in itineraries:
        for img in it.images:
            all_image_paths.append(img)
    image_lib_map = {}
    if all_image_paths:
        lib_records = ImageLibrary.query.filter(
            ImageLibrary.image_path.in_(all_image_paths),
            ImageLibrary.is_active == True
        ).all()
        for rec in lib_records:
            image_lib_map[rec.image_path] = {'title': rec.title, 'tags': rec.tags}

    # 获取城市列表
    from App_new.business.tour.models.Packagemodels import ProductCity
    cities = ProductCity.query.order_by(ProductCity.country_name, ProductCity.city_name).all()

    return render_template('business/tour/products/product_form.html',
                         product=product,
                         suppliers=suppliers,
                         cities=cities,
                         itineraries=itineraries,
                         price_variants=price_variants,
                         image_lib_map=image_lib_map)


@tour_products_bp.route('/<int:product_id>/toggle-status', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def toggle_status(product_id):
    """切换产品状态"""
    try:
        product = Product.query.get_or_404(product_id)
        data = request.get_json()
        new_status = data.get('status')

        if new_status not in ['active', 'draft', 'inactive', 'disabled']:
            return jsonify({'success': False, 'message': '无效的状态值'}), 400

        product.product_status = new_status
        product.updated_at = datetime.utcnow()
        db.session.commit()

        # 同步状态到统一产品表（容错，同步失败不影响状态切换）
        try:
            from App_new.business.products.sync_helper import sync_tour_product_to_unified
            sync_tour_product_to_unified(product, created_by=current_user.username)
            db.session.commit()
        except Exception as sync_err:
            db.session.rollback()
            import logging
            logging.getLogger(__name__).warning(f'产品同步失败(不影响状态切换): {sync_err}')

        return jsonify({'success': True, 'message': '状态更新成功', 'status': new_status})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/<int:product_id>/delete', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def delete_product(product_id):
    """删除产品及其关联数据"""
    try:
        product = Product.query.get_or_404(product_id)

        # 删除关联的行程
        ProductItinerary.query.filter_by(product_id=product_id).delete()

        # 删除关联的价格变体
        ProductPriceVariant.query.filter_by(product_id=product_id).delete()

        # 删除统一产品表中的对应记录
        from App_new.business.products.sync_helper import delete_unified_product_by_tour
        delete_unified_product_by_tour(product_id)

        # 删除产品
        db.session.delete(product)
        db.session.commit()

        return jsonify({'success': True, 'message': '产品已删除'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/bulk-delete', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def bulk_delete_products():
    """批量删除产品"""
    try:
        data = request.get_json()
        product_ids = data.get('product_ids', [])

        if not product_ids:
            return jsonify({'success': False, 'message': '请选择要删除的产品'}), 400

        deleted_count = 0
        errors = []

        for product_id in product_ids:
            try:
                product = Product.query.get(product_id)
                if not product:
                    errors.append(f'产品ID {product_id} 不存在')
                    continue

                # 删除关联的行程
                ProductItinerary.query.filter_by(product_id=product_id).delete()

                # 删除关联的价格变体
                ProductPriceVariant.query.filter_by(product_id=product_id).delete()

                # 删除统一产品表中的对应记录
                from App_new.business.products.sync_helper import delete_unified_product_by_tour
                delete_unified_product_by_tour(product_id)

                # 删除产品
                db.session.delete(product)
                deleted_count += 1

            except Exception as e:
                errors.append(f'产品ID {product_id} 删除失败: {str(e)}')

        if deleted_count > 0:
            db.session.commit()

        message = f'成功删除 {deleted_count} 个产品'
        if errors:
            message += f'，{len(errors)} 个失败'

        return jsonify({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'errors': errors
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/<int:product_id>/renew', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def renew_product(product_id):
    """续期产品（更新有效期）"""
    try:
        product = Product.query.get_or_404(product_id)
        data = request.get_json()
        new_valid_until = data.get('valid_until')
        
        if not new_valid_until:
            return jsonify({'success': False, 'message': '请提供新的有效期日期'}), 400
        
        try:
            new_date = datetime.strptime(new_valid_until, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': '日期格式无效'}), 400
        
        product.valid_until = new_date
        product.product_status = 'active'  # 续期后自动激活
        product.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True, 'message': '续期成功', 'valid_until': str(new_date)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ========================================
# 价格变体管理 API
# ========================================

@tour_products_bp.route('/<int:product_id>/price-variants')
@login_required
@staff_only
def get_price_variants(product_id):
    try:
        variants = ProductPriceVariant.query.filter_by(product_id=product_id).all()
        return jsonify({'success': True, 'variants': [v.to_dict() for v in variants]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/<int:product_id>/price-variants/<int:variant_id>')
@login_required
@staff_only
def get_price_variant(product_id, variant_id):
    try:
        variant = ProductPriceVariant.query.get_or_404(variant_id)
        if variant.product_id != product_id:
            return jsonify({'success': False, 'message': '价格变体不属于该产品'}), 400
        return jsonify({'success': True, 'variant': variant.to_dict()})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/<int:product_id>/price-variants/add', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def add_price_variant(product_id):
    try:
        product = Product.query.get_or_404(product_id)

        is_primary = bool(int(request.form.get('is_primary', 0)))

        # 如果设为主要价格，先取消该产品下其他主要价格
        if is_primary:
            ProductPriceVariant.query.filter_by(product_id=product_id, is_primary=True).update({'is_primary': False})

        variant = ProductPriceVariant(
            product_id=product_id,
            variant_name=request.form['variant_name'],
            start_date=datetime.strptime(request.form['start_date'], '%Y-%m-%d').date() if request.form.get('start_date') else None,
            end_date=datetime.strptime(request.form['end_date'], '%Y-%m-%d').date() if request.form.get('end_date') else None,
            min_pax=int(request.form['min_pax']) if request.form.get('min_pax') else None,
            max_pax=int(request.form['max_pax']) if request.form.get('max_pax') else None,
            # 房型价格
            single_price=float(request.form['single_price']) if request.form.get('single_price') else None,
            twin_price=float(request.form['twin_price']) if request.form.get('twin_price') else None,
            third_pax_price=float(request.form['third_pax_price']) if request.form.get('third_pax_price') else None,
            child_no_bed_price=float(request.form['child_no_bed_price']) if request.form.get('child_no_bed_price') else None,
            # 成本价格
            cost_single_price=float(request.form['cost_single_price']) if request.form.get('cost_single_price') else None,
            cost_twin_price=float(request.form['cost_twin_price']) if request.form.get('cost_twin_price') else None,
            cost_third_pax_price=float(request.form['cost_third_pax_price']) if request.form.get('cost_third_pax_price') else None,
            cost_child_no_bed_price=float(request.form['cost_child_no_bed_price']) if request.form.get('cost_child_no_bed_price') else None,
            profit_per_person=float(request.form['profit_per_person']) if request.form.get('profit_per_person') else None,
            currency=request.form.get('currency', 'SGD'),
            is_primary=is_primary,
            is_active=bool(int(request.form.get('is_active', 1)))
        )

        db.session.add(variant)
        db.session.commit()

        return jsonify({'success': True, 'message': '价格变体添加成功！'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/<int:product_id>/price-variants/<int:variant_id>/update', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def update_price_variant(product_id, variant_id):
    try:
        variant = ProductPriceVariant.query.get_or_404(variant_id)
        if variant.product_id != product_id:
            return jsonify({'success': False, 'message': '价格变体不属于该产品'}), 400

        is_primary = bool(int(request.form.get('is_primary', 0)))

        # 如果设为主要价格，先取消该产品下其他主要价格
        if is_primary:
            ProductPriceVariant.query.filter(
                ProductPriceVariant.product_id == product_id,
                ProductPriceVariant.id != variant_id,
                ProductPriceVariant.is_primary == True
            ).update({'is_primary': False})

        variant.variant_name = request.form['variant_name']
        variant.start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date() if request.form.get('start_date') else None
        variant.end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').date() if request.form.get('end_date') else None
        variant.min_pax = int(request.form['min_pax']) if request.form.get('min_pax') else None
        variant.max_pax = int(request.form['max_pax']) if request.form.get('max_pax') else None
        # 房型价格
        variant.single_price = float(request.form['single_price']) if request.form.get('single_price') else None
        variant.twin_price = float(request.form['twin_price']) if request.form.get('twin_price') else None
        variant.third_pax_price = float(request.form['third_pax_price']) if request.form.get('third_pax_price') else None
        variant.child_no_bed_price = float(request.form['child_no_bed_price']) if request.form.get('child_no_bed_price') else None
        # 成本价格
        variant.cost_single_price = float(request.form['cost_single_price']) if request.form.get('cost_single_price') else None
        variant.cost_twin_price = float(request.form['cost_twin_price']) if request.form.get('cost_twin_price') else None
        variant.cost_third_pax_price = float(request.form['cost_third_pax_price']) if request.form.get('cost_third_pax_price') else None
        variant.cost_child_no_bed_price = float(request.form['cost_child_no_bed_price']) if request.form.get('cost_child_no_bed_price') else None
        variant.profit_per_person = float(request.form['profit_per_person']) if request.form.get('profit_per_person') else None
        variant.currency = request.form.get('currency', 'SGD')
        variant.is_primary = is_primary
        variant.is_active = bool(int(request.form.get('is_active', 1)))
        variant.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({'success': True, 'message': '价格变体更新成功！'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@tour_products_bp.route('/<int:product_id>/price-variants/<int:variant_id>/delete', methods=['POST'])
@csrf.exempt
@login_required
@staff_only
def delete_price_variant(product_id, variant_id):
    try:
        variant = ProductPriceVariant.query.get_or_404(variant_id)
        if variant.product_id != product_id:
            return jsonify({'success': False, 'message': '价格变体不属于该产品'}), 400
        db.session.delete(variant)
        db.session.commit()
        return jsonify({'success': True, 'message': '价格变体删除成功！'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ========================================
# Excel 导入导出功能
# ========================================

@tour_products_bp.route('/export/excel')
@login_required
@staff_only
def export_excel():
    """导出产品Excel - 只导出选中的产品"""
    try:
        # 获取选中的产品ID列表
        ids_str = request.args.get('ids', '')
        if not ids_str:
            flash('请先选择要导出的产品', 'warning')
            return redirect(url_for('tour_products.product_list'))

        try:
            product_ids = [int(x) for x in ids_str.split(',') if x.strip()]
        except ValueError:
            flash('产品ID参数无效', 'danger')
            return redirect(url_for('tour_products.product_list'))

        if not product_ids:
            flash('请先选择要导出的产品', 'warning')
            return redirect(url_for('tour_products.product_list'))

        products = Product.query.filter(Product.id.in_(product_ids)).order_by(Product.created_at.desc()).all()
        data = []
        for product in products:
            data.append({
                'ID': product.id,
                '供应商': product.supplier.company_name if product.supplier else '',
                '产品编号': product.product_code or '',
                '产品名称': product.product_name,
                '产品类型': product.product_type or '',
                '国家': product.country or '',
                '城市': product.city_name or '',
                '出发城市': product.departure_city or '',
                '目的地城市': product.destination_city or '',
                '行程天数': product.duration_days or '',
                '最少人数': product.min_pax or '',
                '最多人数': product.max_pax or '',
                '成人价格': product.base_price or '',
                '儿童价格': product.child_price or '',
                '婴儿价格': product.infant_price or '',
                '单房差': product.single_room_supplement or '',
                '货币': product.currency or 'SGD',
                '产品描述': product.product_description or '',
                '包含服务': product.included_services or '',
                '不包含服务': product.excluded_services or '',
                '重要提示': product.important_notes or '',
                '适合季节': product.suitable_season or '',
                '难度等级': product.difficulty_level or '',
                '标签': product.tags or '',
                '产品状态': product.product_status or 'draft',
                '有效期从': product.valid_from.strftime('%Y-%m-%d') if product.valid_from else '',
                '有效期至': product.valid_until.strftime('%Y-%m-%d') if product.valid_until else '',
                '创建时间': product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else '',
                '更新时间': product.updated_at.strftime('%Y-%m-%d %H:%M:%S') if product.updated_at else '',
            })

        df = pd.DataFrame(data)

        # 收集价格方案数据
        price_data = []
        for product in products:
            for pv in product.price_variants:
                price_data.append({
                    '产品编号': product.product_code or '',
                    '方案名称': pv.variant_name or '',
                    '开始日期': pv.start_date.strftime('%Y-%m-%d') if pv.start_date else '',
                    '结束日期': pv.end_date.strftime('%Y-%m-%d') if pv.end_date else '',
                    '最少人数': pv.min_pax or '',
                    '最多人数': pv.max_pax or '',
                    '单人房价格': pv.single_price or '',
                    '双人房价格': pv.twin_price or '',
                    '第三人价格': pv.third_pax_price or '',
                    '儿童不占床价格': pv.child_no_bed_price or '',
                    '成本_单人房': pv.cost_single_price or '',
                    '成本_双人房': pv.cost_twin_price or '',
                    '成本_第三人': pv.cost_third_pax_price or '',
                    '成本_儿童不占床': pv.cost_child_no_bed_price or '',
                    '货币': pv.currency or 'SGD',
                    '每人利润': pv.profit_per_person or '',
                    '是否主要': '是' if pv.is_primary else '否',
                    '是否启用': '是' if pv.is_active else '否',
                })
        df_price = pd.DataFrame(price_data)

        # 收集每日行程数据
        itinerary_data = []
        for product in products:
            for it in product.itineraries:
                itinerary_data.append({
                    '产品编号': product.product_code or '',
                    '天数': it.day_number,
                    '标题': it.day_title or '',
                    '行程内容': it.content or '',
                })
        df_itinerary = pd.DataFrame(itinerary_data)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: 产品数据
            df.to_excel(writer, index=False, sheet_name='产品数据')
            worksheet = writer.sheets['产品数据']
            for idx, col in enumerate(df.columns):
                max_length = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                col_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[col_letter].width = min(max_length, 50)

            # Sheet 2: 价格方案
            if not df_price.empty:
                df_price.to_excel(writer, index=False, sheet_name='价格方案')
            else:
                pd.DataFrame(columns=['产品编号', '方案名称', '开始日期', '结束日期',
                    '最少人数', '最多人数', '单人房价格', '双人房价格', '第三人价格',
                    '儿童不占床价格', '成本_单人房', '成本_双人房', '成本_第三人',
                    '成本_儿童不占床', '货币', '每人利润', '是否主要', '是否启用'
                ]).to_excel(writer, index=False, sheet_name='价格方案')
            ws_price = writer.sheets['价格方案']
            for idx in range(18):
                ws_price.column_dimensions[get_column_letter(idx + 1)].width = 16

            # Sheet 3: 每日行程
            if not df_itinerary.empty:
                df_itinerary.to_excel(writer, index=False, sheet_name='每日行程')
            else:
                pd.DataFrame(columns=['产品编号', '天数', '标题', '行程内容']
                ).to_excel(writer, index=False, sheet_name='每日行程')
            ws_itin = writer.sheets['每日行程']
            ws_itin.column_dimensions['A'].width = 16
            ws_itin.column_dimensions['B'].width = 8
            ws_itin.column_dimensions['C'].width = 30
            ws_itin.column_dimensions['D'].width = 60

        output.seek(0)
        filename = f'tour_products_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        flash(f'导出失败：{str(e)}', 'danger')
        return redirect(url_for('tour_products.product_list'))


@tour_products_bp.route('/import/excel', methods=['POST'])
@login_required
@staff_only
def import_excel():
    try:
        if 'file' not in request.files:
            flash('请选择要上传的文件', 'danger')
            return redirect(url_for('tour_products.product_list'))

        file = request.files['file']
        if file.filename == '':
            flash('未选择文件', 'danger')
            return redirect(url_for('tour_products.product_list'))

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('只支持 Excel 文件（.xlsx, .xls）', 'danger')
            return redirect(url_for('tour_products.product_list'))

        # 读取所有 Sheet
        xls = pd.ExcelFile(file, engine='openpyxl')
        sheet_names = xls.sheet_names

        # Sheet 1: 产品数据（第一个 Sheet）
        df = pd.read_excel(xls, sheet_name=0)

        imported_count = 0
        updated_count = 0
        price_imported = 0
        price_updated = 0
        itinerary_imported = 0
        itinerary_updated = 0
        errors = []

        # 用于记录产品编号到产品对象的映射（供 Sheet 2/3 使用）
        product_code_map = {}

        for index, row in df.iterrows():
            try:
                supplier = None
                if pd.notna(row.get('供应商')):
                    supplier = CustomerCompany.query.filter(
                        CustomerCompany.company_name == str(row['供应商']).strip(),
                        CustomerCompany.is_supplier == True
                    ).first()

                product = None
                if pd.notna(row.get('ID')):
                    product = Product.query.get(int(row['ID']))
                elif pd.notna(row.get('产品编号')) and str(row['产品编号']).strip():
                    product = Product.query.filter_by(product_code=str(row['产品编号']).strip()).first()

                # 处理 city_id（从 city_name 获取，如果城市不存在则自动创建）
                city_name = str(row['城市']).strip() if pd.notna(row.get('城市')) else None
                city_id = None
                if city_name:
                    from App_new.business.tour.models.Packagemodels import ProductCity
                    city = ProductCity.query.filter_by(city_name=city_name).first()
                    if not city:
                        # 自动创建城市记录
                        country_name = str(row['国家']).strip() if pd.notna(row.get('国家')) else '未知'
                        city = ProductCity(
                            city_name=city_name,
                            display_name=city_name,
                            country_name=country_name
                        )
                        db.session.add(city)
                        db.session.flush()
                    city_id = city.id

                # 处理 product_code，如果为空则自动生成（使用统一编号规则）
                from App_new.shared.models.business_types import generate_product_code as gen_code
                product_code_raw = str(row['产品编号']).strip() if pd.notna(row.get('产品编号')) else ''
                product_code = product_code_raw if product_code_raw else gen_code('tour')

                product_data = {
                    'supplier_id': supplier.id if supplier else None,
                    'product_code': product_code,
                    'product_name': str(row['产品名称']).strip(),
                    'product_type': str(row['产品类型']).strip() if pd.notna(row.get('产品类型')) else None,
                    'city_id': city_id,
                    'city_name': city_name,
                    'departure_city': str(row['出发城市']).strip() if pd.notna(row.get('出发城市')) else None,
                    'destination_city': str(row['目的地城市']).strip() if pd.notna(row.get('目的地城市')) else None,
                    'duration_days': int(row['行程天数']) if pd.notna(row.get('行程天数')) else None,
                    'min_pax': int(row['最少人数']) if pd.notna(row.get('最少人数')) else 1,
                    'max_pax': int(row['最多人数']) if pd.notna(row.get('最多人数')) else None,
                    'base_price': float(row['成人价格']) if pd.notna(row.get('成人价格')) else None,
                    'child_price': float(row['儿童价格']) if pd.notna(row.get('儿童价格')) else None,
                    'infant_price': float(row['婴儿价格']) if pd.notna(row.get('婴儿价格')) else None,
                    'single_room_supplement': float(row['单房差']) if pd.notna(row.get('单房差')) else None,
                    'currency': str(row['货币']).strip() if pd.notna(row.get('货币')) else 'SGD',
                    'product_description': str(row['产品描述']).strip() if pd.notna(row.get('产品描述')) else None,
                    'included_services': str(row['包含服务']).strip() if pd.notna(row.get('包含服务')) else None,
                    'excluded_services': str(row['不包含服务']).strip() if pd.notna(row.get('不包含服务')) else None,
                    'important_notes': str(row['重要提示']).strip() if pd.notna(row.get('重要提示')) else None,
                    'suitable_season': str(row['适合季节']).strip() if pd.notna(row.get('适合季节')) else None,
                    'difficulty_level': str(row['难度等级']).strip() if pd.notna(row.get('难度等级')) else None,
                    'tags': str(row['标签']).strip() if pd.notna(row.get('标签')) else None,
                    'product_status': str(row['产品状态']).strip() if pd.notna(row.get('产品状态')) else 'draft',
                }

                if pd.notna(row.get('有效期从')):
                    try:
                        product_data['valid_from'] = pd.to_datetime(row['有效期从']).date()
                    except:
                        pass

                if pd.notna(row.get('有效期至')):
                    try:
                        product_data['valid_until'] = pd.to_datetime(row['有效期至']).date()
                    except:
                        pass

                if product:
                    for key, value in product_data.items():
                        setattr(product, key, value)
                    product.updated_at = datetime.utcnow()
                    updated_count += 1
                else:
                    product = Product(**product_data)
                    product.created_by = f'EXCEL_IMPORT_{current_user.username}'
                    db.session.add(product)
                    imported_count += 1

                # 记录产品编号映射
                if product_code:
                    product_code_map[product_code] = product

            except Exception as e:
                errors.append(f'产品第 {index + 2} 行导入失败: {str(e)}')
                continue

        # 先 flush 确保新产品有 ID
        db.session.flush()

        # Sheet 2: 价格方案（可选）
        if '价格方案' in sheet_names:
            df_price = pd.read_excel(xls, sheet_name='价格方案')
            for index, row in df_price.iterrows():
                try:
                    code = str(row['产品编号']).strip() if pd.notna(row.get('产品编号')) else ''
                    if not code:
                        errors.append(f'价格方案第 {index + 2} 行：产品编号为空，跳过')
                        continue

                    # 先从本次导入映射中查找，再从数据库查找
                    product = product_code_map.get(code)
                    if not product:
                        product = Product.query.filter_by(product_code=code).first()
                    if not product:
                        errors.append(f'价格方案第 {index + 2} 行：产品编号 {code} 未找到，跳过')
                        continue

                    variant_name = str(row['方案名称']).strip() if pd.notna(row.get('方案名称')) else ''
                    if not variant_name:
                        errors.append(f'价格方案第 {index + 2} 行：方案名称为空，跳过')
                        continue

                    # 匹配规则：同产品 + 同方案名称 → 更新
                    variant = ProductPriceVariant.query.filter_by(
                        product_id=product.id, variant_name=variant_name
                    ).first()

                    variant_data = {
                        'product_id': product.id,
                        'variant_name': variant_name,
                        'min_pax': int(row['最少人数']) if pd.notna(row.get('最少人数')) else None,
                        'max_pax': int(row['最多人数']) if pd.notna(row.get('最多人数')) else None,
                        'single_price': float(row['单人房价格']) if pd.notna(row.get('单人房价格')) else None,
                        'twin_price': float(row['双人房价格']) if pd.notna(row.get('双人房价格')) else None,
                        'third_pax_price': float(row['第三人价格']) if pd.notna(row.get('第三人价格')) else None,
                        'child_no_bed_price': float(row['儿童不占床价格']) if pd.notna(row.get('儿童不占床价格')) else None,
                        'cost_single_price': float(row['成本_单人房']) if pd.notna(row.get('成本_单人房')) else None,
                        'cost_twin_price': float(row['成本_双人房']) if pd.notna(row.get('成本_双人房')) else None,
                        'cost_third_pax_price': float(row['成本_第三人']) if pd.notna(row.get('成本_第三人')) else None,
                        'cost_child_no_bed_price': float(row['成本_儿童不占床']) if pd.notna(row.get('成本_儿童不占床')) else None,
                        'currency': str(row['货币']).strip() if pd.notna(row.get('货币')) else 'SGD',
                        'profit_per_person': float(row['每人利润']) if pd.notna(row.get('每人利润')) else None,
                        'is_primary': str(row.get('是否主要', '')).strip() == '是',
                        'is_active': str(row.get('是否启用', '是')).strip() != '否',
                    }

                    # 处理日期
                    if pd.notna(row.get('开始日期')):
                        try:
                            variant_data['start_date'] = pd.to_datetime(row['开始日期']).date()
                        except:
                            pass
                    if pd.notna(row.get('结束日期')):
                        try:
                            variant_data['end_date'] = pd.to_datetime(row['结束日期']).date()
                        except:
                            pass

                    if variant:
                        for key, value in variant_data.items():
                            setattr(variant, key, value)
                        variant.updated_at = datetime.utcnow()
                        price_updated += 1
                    else:
                        variant = ProductPriceVariant(**variant_data)
                        db.session.add(variant)
                        price_imported += 1

                except Exception as e:
                    errors.append(f'价格方案第 {index + 2} 行导入失败: {str(e)}')
                    continue

        # Sheet 3: 每日行程（可选）
        if '每日行程' in sheet_names:
            df_itin = pd.read_excel(xls, sheet_name='每日行程')
            for index, row in df_itin.iterrows():
                try:
                    code = str(row['产品编号']).strip() if pd.notna(row.get('产品编号')) else ''
                    if not code:
                        errors.append(f'每日行程第 {index + 2} 行：产品编号为空，跳过')
                        continue

                    product = product_code_map.get(code)
                    if not product:
                        product = Product.query.filter_by(product_code=code).first()
                    if not product:
                        errors.append(f'每日行程第 {index + 2} 行：产品编号 {code} 未找到，跳过')
                        continue

                    day_number = int(row['天数']) if pd.notna(row.get('天数')) else None
                    if not day_number:
                        errors.append(f'每日行程第 {index + 2} 行：天数为空，跳过')
                        continue

                    # 匹配规则：同产品 + 同天数 → 更新
                    itinerary = ProductItinerary.query.filter_by(
                        product_id=product.id, day_number=day_number
                    ).first()

                    itin_data = {
                        'product_id': product.id,
                        'day_number': day_number,
                        'day_title': str(row['标题']).strip() if pd.notna(row.get('标题')) else None,
                        'content': str(row['行程内容']).strip() if pd.notna(row.get('行程内容')) else None,
                    }

                    if itinerary:
                        for key, value in itin_data.items():
                            setattr(itinerary, key, value)
                        itinerary.updated_at = datetime.utcnow()
                        itinerary_updated += 1
                    else:
                        itinerary = ProductItinerary(**itin_data)
                        db.session.add(itinerary)
                        itinerary_imported += 1

                except Exception as e:
                    errors.append(f'每日行程第 {index + 2} 行导入失败: {str(e)}')
                    continue

        total_changes = imported_count + updated_count + price_imported + price_updated + itinerary_imported + itinerary_updated
        if total_changes > 0:
            db.session.commit()
            msg_parts = []
            if imported_count or updated_count:
                msg_parts.append(f'产品：新增 {imported_count}，更新 {updated_count}')
            if price_imported or price_updated:
                msg_parts.append(f'价格方案：新增 {price_imported}，更新 {price_updated}')
            if itinerary_imported or itinerary_updated:
                msg_parts.append(f'每日行程：新增 {itinerary_imported}，更新 {itinerary_updated}')
            flash(f'导入成功！{"；".join(msg_parts)}', 'success')
        else:
            flash('没有导入任何数据', 'warning')

        if errors:
            for error in errors[:5]:
                flash(error, 'danger')
            if len(errors) > 5:
                flash(f'还有 {len(errors) - 5} 个错误未显示', 'warning')

    except Exception as e:
        db.session.rollback()
        flash(f'导入失败：{str(e)}', 'danger')

    return redirect(url_for('tour_products.product_list'))


@tour_products_bp.route('/download/template')
@login_required
@staff_only
def download_template():
    """下载 Excel 导入模板"""
    try:
        template_data = {
            'ID': ['', '留空则新建，填写则更新'],
            '供应商': ['供应商名称', '必须是系统中已存在的供应商'],
            '产品编号': ['SG-CITY-001', '可选，唯一标识'],
            '产品名称': ['新加坡3天2晚自由行', '必填'],
            '产品类型': ['自由行', '可选：跟团游/自由行/定制游/当地游'],
            '国家': ['新加坡', '必填'],
            '城市': ['新加坡', '可选'],
            '出发城市': ['北京', '可选'],
            '目的地城市': ['新加坡', '可选'],
            '行程天数': [3, '必填，数字'],
            '最少人数': [2, '可选，默认1'],
            '最多人数': [40, '可选'],
            '成人价格': [1200, '可选，数字'],
            '儿童价格': [800, '可选，数字'],
            '婴儿价格': [0, '可选，数字'],
            '单房差': [200, '可选，数字'],
            '货币': ['SGD', '可选，默认SGD'],
            '产品描述': ['体验新加坡的现代与传统', '可选'],
            '产品亮点': ['鱼尾狮公园\n滨海湾花园', '可选，多行用换行分隔'],
            '包含服务': ['往返机票\n酒店住宿', '可选'],
            '不包含服务': ['签证费用\n个人消费', '可选'],
            '重要提示': ['请确保护照有效期6个月以上', '可选'],
            '适合季节': ['全年', '可选'],
            '难度等级': ['简单', '可选：简单/中等/困难'],
            '标签': ['蜜月,豪华,亲子', '可选，逗号分隔'],
            '产品状态': ['active', '可选：active/draft/inactive，默认draft'],
            '有效期从': ['2025-01-01', '可选，格式：YYYY-MM-DD'],
            '有效期至': ['2025-12-31', '可选，格式：YYYY-MM-DD'],
            '创建时间': ['', '系统自动生成'],
            '更新时间': ['', '系统自动生成'],
        }

        df = pd.DataFrame(template_data)

        # 价格方案模板数据
        price_template = {
            '产品编号': ['SG-CITY-001', '填写产品编号，必须匹配产品数据Sheet'],
            '方案名称': ['标准价格', '必填，同产品+同名称则更新'],
            '开始日期': ['2026-01-01', '可选，格式：YYYY-MM-DD'],
            '结束日期': ['2026-12-31', '可选，格式：YYYY-MM-DD'],
            '最少人数': [2, '可选，数字'],
            '最多人数': [10, '可选，数字'],
            '单人房价格': [1500, '可选，数字'],
            '双人房价格': [1200, '可选，数字'],
            '第三人价格': [1000, '可选，数字'],
            '儿童不占床价格': [800, '可选，数字'],
            '成本_单人房': [1200, '可选，成本价'],
            '成本_双人房': [900, '可选，成本价'],
            '成本_第三人': [750, '可选，成本价'],
            '成本_儿童不占床': [600, '可选，成本价'],
            '货币': ['SGD', '可选，默认SGD'],
            '每人利润': [100, '可选，数字'],
            '是否主要': ['是', '是/否，默认否'],
            '是否启用': ['是', '是/否，默认是'],
        }
        df_price = pd.DataFrame(price_template)

        # 每日行程模板数据
        itinerary_template = {
            '产品编号': ['SG-CITY-001', '填写产品编号，必须匹配产品数据Sheet'],
            '天数': [1, '必填，第几天（数字）'],
            '标题': ['抵达新加坡', '可选，当日标题'],
            '行程内容': ['抵达樟宜机场，接机前往酒店办理入住', '可选，行程详细内容'],
        }
        df_itinerary = pd.DataFrame(itinerary_template)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: 产品数据模板
            df.to_excel(writer, index=False, sheet_name='导入模板')
            worksheet = writer.sheets['导入模板']
            for idx, col in enumerate(df.columns):
                col_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[col_letter].width = 20

            # Sheet 2: 价格方案模板
            df_price.to_excel(writer, index=False, sheet_name='价格方案')
            ws_price = writer.sheets['价格方案']
            for idx in range(len(price_template)):
                ws_price.column_dimensions[get_column_letter(idx + 1)].width = 18

            # Sheet 3: 每日行程模板
            df_itinerary.to_excel(writer, index=False, sheet_name='每日行程')
            ws_itin = writer.sheets['每日行程']
            ws_itin.column_dimensions['A'].width = 16
            ws_itin.column_dimensions['B'].width = 8
            ws_itin.column_dimensions['C'].width = 30
            ws_itin.column_dimensions['D'].width = 60

        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='tour_products_template.xlsx'
        )

    except Exception as e:
        flash(f'下载模板失败：{str(e)}', 'danger')
        return redirect(url_for('tour_products.product_list'))
