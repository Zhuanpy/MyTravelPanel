import os
from datetime import datetime
from flask import Blueprint, render_template, url_for, flash, redirect, request, jsonify, current_app, Response, send_file
from flask_login import login_required, current_user
from App_new.utils.decorators import staff_only
from sqlalchemy import or_, and_
from sqlalchemy.exc import SQLAlchemyError
from App_new.exts import db, csrf
from ..models.PackageBudget import BudgetHeader, BudgetItem
from ..models.TourProject import TourProject
from App_new.shared.models.business_types import BusinessType
import urllib.parse


# 创建蓝图
package_budget = Blueprint('package_budget', __name__, url_prefix='/package_budget')


@package_budget.route('/')
@package_budget.route('/list')
@login_required
@staff_only
def list_budgets():
    """预算单列表页面"""
    try:
        # 获取查询参数
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        search = request.args.get('search', '').strip()
        status_filter = request.args.get('status', '')
        is_template_filter = request.args.get('is_template', '')
        
        # 构建查询
        query = BudgetHeader.query
        
        # 搜索过滤
        if search:
            query = query.filter(
                or_(
                    BudgetHeader.package_name.ilike(f'%{search}%'),
                    BudgetHeader.remarks.ilike(f'%{search}%'),
                    BudgetHeader.created_by.ilike(f'%{search}%')
                )
            )
        
        # 状态过滤
        if status_filter:
            query = query.filter(BudgetHeader.status == status_filter)
        
        # 类型过滤
        if is_template_filter != '':
            is_template_bool = is_template_filter == '1'
            query = query.filter(BudgetHeader.is_template == is_template_bool)
        
        # 排序
        query = query.order_by(BudgetHeader.created_at.desc())
        
        # 分页
        pagination = query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        budgets = pagination.items
        
        return render_template('business/tour/package/TourBudget/list.html', 
                             budgets=budgets, 
                             pagination=pagination)
    
    except Exception as e:
        current_app.logger.error(f"Error in list_budgets: {e}")
        flash('获取预算单列表时发生错误', 'error')
        return render_template('business/tour/package/TourBudget/list.html', budgets=[], pagination=None)


@package_budget.route('/create', methods=['GET', 'POST'])
@login_required
@staff_only
@csrf.exempt
def create():
    """创建新预算单"""
    if request.method == 'POST':
        try:
            # 获取表单数据
            package_name = request.form.get('package_name', '').strip()
            adult_count = request.form.get('adult_count', type=int)
            child_count = request.form.get('child_count', type=int) or 0
            currency = request.form.get('currency', 'SGD')
            status = request.form.get('status', 'draft')
            is_template = request.form.get('is_template') == '1'
            remarks = request.form.get('remarks', '').strip()
            project_id = request.form.get('project_id', type=int)
            
            # 验证必填字段
            if not package_name:
                flash('套餐名称不能为空', 'error')
                return render_template('business/tour/package/TourBudget/create.html', form=request.form)
            
            if not adult_count or adult_count < 1:
                flash('成人数量必须大于0', 'error')
                return render_template('business/tour/package/TourBudget/create.html', form=request.form)
            
            # 获取当前用户名字
            if current_user.is_authenticated and current_user.profile:
                created_by_name = f"{current_user.profile.first_name or ''} {current_user.profile.last_name or ''}".strip()
                if not created_by_name:
                    created_by_name = current_user.email
            elif current_user.is_authenticated:
                created_by_name = current_user.email
            else:
                created_by_name = 'admin'
            
            # 创建预算单
            budget = BudgetHeader(
                package_name=package_name,
                adult_count=adult_count,
                child_count=child_count,
                currency=currency,
                status=status,
                is_template=is_template,
                remarks=remarks,
                created_by=created_by_name,
                created_at=datetime.utcnow(),
                project_id=project_id
            )
            
            db.session.add(budget)

            # 双向同步：新建预算若关联项目，人数回写到该项目团队
            if project_id:
                from App_new.business.tour.models.TourProject import TourGroup
                TourGroup.query.filter_by(project_id=project_id).update(
                    {'adult_count': adult_count, 'child_count': child_count,
                     'pax': adult_count + child_count},
                    synchronize_session=False)

            db.session.commit()

            flash('预算单创建成功！', 'success')
            return redirect(url_for('package_budget.detail', budget_id=budget.id))
            
        except SQLAlchemyError as e:
            db.session.rollback()
            current_app.logger.error(f"Database error in create budget: {e}")
            flash('创建预算单时发生数据库错误', 'error')
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error in create budget: {e}")
            flash('创建预算单时发生错误', 'error')
    
    # GET: 支持通过查询参数预填（来自旅游项目编辑页）
    if request.method == 'GET':
        project_id = request.args.get('project_id', type=int)
        project = None
        
        # 如果有project_id，获取项目信息用于预填
        if project_id:
            project = TourProject.query.get(project_id)
        
        # 尝试从项目关联的团队获取人数信息
        adult_count = ''
        child_count = '0'
        if project and hasattr(project, 'groups') and project.groups:
            first_group = project.groups[0] if project.groups else None
            if first_group:
                adult_count = first_group.pax or ''
        
        prefill = {
            'package_name': request.args.get('package_name', project.project_name if project else ''),
            'adult_count': request.args.get('adult_count', adult_count),
            'child_count': request.args.get('child_count', child_count),
            'currency': request.args.get('currency', project.currency if project else ''),
            'status': request.args.get('status', ''),
            'is_template': request.args.get('is_template', ''),
            'remarks': request.args.get('remarks', ''),
            'project_id': project_id,
        }
        return render_template('business/tour/package/TourBudget/create.html', form=prefill, project=project)

    return render_template('business/tour/package/TourBudget/create.html')


@package_budget.route('/quick_create/<int:project_id>', methods=['GET', 'POST'])
@login_required
@staff_only
@csrf.exempt
def quick_create_for_project(project_id):
    """为旅游项目一键快速创建预算单（无需手填表单）

    - 幂等：项目已有预算 → 直接进最新那个，不重复创建
    - 否则按项目信息自动建单（套餐名=项目名、人数取首个团、币种取项目），
      建好直接跳到预算详情页让用户加明细
    """
    project = TourProject.query.get_or_404(project_id)

    existing = BudgetHeader.query.filter_by(project_id=project_id)\
        .order_by(BudgetHeader.created_at.desc()).first()
    if existing:
        return redirect(url_for('package_budget.detail', budget_id=existing.id))

    # 人数：优先团队的大人/小孩，回退总人数 pax，再兜底 1
    adult_count, child_count = 1, 0
    first_group = project.groups.first() if hasattr(project, 'groups') else None
    if first_group:
        adult_count = first_group.adult_count or first_group.pax or 1
        child_count = first_group.child_count or 0

    if current_user.is_authenticated and current_user.profile:
        created_by_name = f"{current_user.profile.first_name or ''} {current_user.profile.last_name or ''}".strip() \
            or current_user.email
    elif current_user.is_authenticated:
        created_by_name = current_user.email
    else:
        created_by_name = 'admin'

    try:
        budget = BudgetHeader(
            package_name=project.project_name or f"项目{project_id}预算",
            adult_count=adult_count,
            child_count=child_count,
            currency=project.currency or 'SGD',
            status='active',
            is_template=False,
            created_by=created_by_name,
            created_at=datetime.utcnow(),
            project_id=project_id,
        )
        db.session.add(budget)
        db.session.commit()
        flash('已为该项目快速创建预算单，请继续添加预算明细', 'success')
        return redirect(url_for('package_budget.detail', budget_id=budget.id))
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"quick_create_for_project failed: {e}")
        flash('快速创建预算单失败', 'error')
        return redirect(url_for('package_budget.budgets_by_project', project_id=project_id))


@package_budget.route('/<int:budget_id>')
@login_required
@staff_only
def detail(budget_id):
    """预算单详情页面"""
    try:
        budget = BudgetHeader.query.get_or_404(budget_id)
        # 获取业务类型作为“类别”选项
        business_types = BusinessType.query.filter_by(is_active=True).order_by(BusinessType.sort_order.asc(), BusinessType.id.asc()).all()
        
        # 计算分类统计
        category_totals = {}
        adult_total = 0
        child_total = 0
        
        for item in budget.items:
            # 分类统计
            category = item.category or '未分类'
            if category not in category_totals:
                category_totals[category] = 0
            category_totals[category] += item.subtotal or 0
            
            # 计算成人费用
            if item.count_adult_apply:
                adult_count = item.adult_count_override or budget.adult_count
                adult_unit_price = item.adult_unit_price or 0
                adult_total += adult_unit_price * adult_count
            
            # 计算儿童费用
            if item.count_child_apply:
                child_count = item.child_count_override or budget.child_count
                child_unit_price = item.child_unit_price or 0
                child_total += child_unit_price * child_count
        
        return render_template('business/tour/package/TourBudget/detail.html',
                             budget=budget,
                             category_totals=category_totals,
                             adult_total=adult_total,
                             child_total=child_total,
                             business_types=business_types)
    
    except Exception as e:
        current_app.logger.error(f"Error in budget detail: {e}")
        flash('获取预算单详情时发生错误', 'error')
        return redirect(url_for('package_budget.list_budgets'))


@package_budget.route('/<int:budget_id>/edit', methods=['GET', 'POST'])
@login_required
@staff_only
@csrf.exempt
def edit(budget_id):
    """编辑预算单"""
    try:
        budget = BudgetHeader.query.get_or_404(budget_id)
        
        # 获取所有项目用于下拉选择
        projects = TourProject.query.order_by(TourProject.created_at.desc()).all()
        
        if request.method == 'POST':
            # 获取表单数据
            package_name = request.form.get('package_name', '').strip()
            adult_count = request.form.get('adult_count', type=int)
            child_count = request.form.get('child_count', type=int) or 0
            currency = request.form.get('currency', 'SGD')
            status = request.form.get('status', 'draft')
            is_template = request.form.get('is_template') == '1'
            remarks = request.form.get('remarks', '').strip()
            project_id = request.form.get('project_id', type=int) or None
            
            # 验证必填字段
            if not package_name:
                flash('套餐名称不能为空', 'error')
                return render_template('business/tour/package/TourBudget/edit.html', budget=budget, projects=projects)
            
            if not adult_count or adult_count < 1:
                flash('成人数量必须大于0', 'error')
                return render_template('business/tour/package/TourBudget/edit.html', budget=budget, projects=projects)
            
            # 更新预算单
            budget.package_name = package_name
            budget.adult_count = adult_count
            budget.child_count = child_count
            budget.currency = currency
            budget.status = status
            budget.is_template = is_template
            budget.remarks = remarks
            budget.project_id = project_id
            budget.updated_at = datetime.utcnow()

            # 双向同步：预算人数回写到关联项目的团队（pax 自动汇总）
            if project_id:
                from App_new.business.tour.models.TourProject import TourGroup
                TourGroup.query.filter_by(project_id=project_id).update(
                    {'adult_count': adult_count, 'child_count': child_count,
                     'pax': adult_count + child_count},
                    synchronize_session=False)

            db.session.commit()

            flash('预算单更新成功！', 'success')
            return redirect(url_for('package_budget.detail', budget_id=budget.id))
        
        return render_template('business/tour/package/TourBudget/edit.html', budget=budget, projects=projects)
    
    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"Database error in edit budget: {e}")
        flash('更新预算单时发生数据库错误', 'error')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in edit budget: {e}")
        flash('更新预算单时发生错误', 'error')
    
    return redirect(url_for('package_budget.detail', budget_id=budget_id))


@package_budget.route('/<int:budget_id>/delete', methods=['POST'])
@login_required
@staff_only
@csrf.exempt
def delete(budget_id):
    """删除预算单"""
    try:
        budget = BudgetHeader.query.get_or_404(budget_id)
        
        # 删除关联的项目
        for item in budget.items:
            db.session.delete(item)
        
        # 删除预算单
        db.session.delete(budget)
        db.session.commit()
        
        flash('预算单删除成功！', 'success')
        return jsonify({'success': True})
    
    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"Database error in delete budget: {e}")
        return jsonify({'success': False, 'error': '数据库错误'}), 500
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in delete budget: {e}")
        return jsonify({'success': False, 'error': '删除失败'}), 500


@package_budget.route('/<int:budget_id>/add_item', methods=['POST'])
@login_required
@staff_only
@csrf.exempt
def add_item(budget_id):
    """添加预算项目"""
    header = BudgetHeader.query.get_or_404(budget_id)
    
    try:
        # 获取表单数据
        category = request.form.get('category', '').strip()
        item_name = request.form.get('item_name', '').strip()
        item_details = request.form.get('item_details', '').strip()
        pricing_method = request.form.get('pricing_method', 'person_based')
        
        # 验证必填字段
        if not category:
            flash('类别不能为空', 'error')
            return redirect(url_for('package_budget.detail', budget_id=budget_id))
        
        if not item_name:
            flash('项目名称不能为空', 'error')
            return redirect(url_for('package_budget.detail', budget_id=budget_id))
        
        # 创建新项目
        item = BudgetItem(
            header_id=budget_id,
            category=category,
            item_name=item_name,
            item_details=item_details,
            pricing_method=pricing_method
        )
        
        # 根据计价方式处理价格数据
        if pricing_method == 'item_based':
            # 模板1：物品计价方式
            item_unit_price = request.form.get('item_unit_price')
            item_quantity = request.form.get('item_quantity', 1)
            
            if item_unit_price and item_quantity:
                item.item_unit_price = float(item_unit_price)
                item.item_quantity = int(item_quantity)
            else:
                flash('物品计价方式需要填写物品单价和件数', 'error')
                return redirect(url_for('package_budget.detail', budget_id=budget_id))
                
        else:
            # 模板2：人均计价方式
            adult_price = request.form.get('adult_price')
            child_price = request.form.get('child_price')
            
            if adult_price:
                item.adult_price = float(adult_price)
            if child_price:
                item.child_price = float(child_price)
                
            if not adult_price and not child_price:
                flash('人均计价方式需要至少填写成人单价或儿童单价', 'error')
                return redirect(url_for('package_budget.detail', budget_id=budget_id))
        
        # 处理人数设置
        item.count_adult_apply = 'count_adult_apply' in request.form
        item.count_child_apply = 'count_child_apply' in request.form
        item.is_optional = 'is_optional' in request.form
        
        # 处理可选的人数覆盖
        adult_count_override = request.form.get('adult_count_override')
        child_count_override = request.form.get('child_count_override')
        
        if adult_count_override:
            item.adult_count_override = int(adult_count_override)
        if child_count_override:
            item.child_count_override = int(child_count_override)
        
        # 处理备注
        remarks = request.form.get('remarks', '').strip()
        if remarks:
            item.remarks = remarks
        
        # 设置排序
        max_order = db.session.query(db.func.max(BudgetItem.sort_order)).filter_by(header_id=budget_id).scalar() or 0
        item.sort_order = max_order + 1
        
        db.session.add(item)
        db.session.commit()
        
        flash('项目添加成功！', 'success')
        
    except ValueError as e:
        flash(f'数据格式错误: {str(e)}', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'添加失败: {str(e)}', 'error')
    
    return redirect(url_for('package_budget.detail', budget_id=budget_id))


@package_budget.route('/<int:budget_id>/item/<int:item_id>/edit', methods=['GET', 'POST'])
@login_required
@staff_only
@csrf.exempt
def edit_item(budget_id, item_id):
    """编辑预算项目"""
    header = BudgetHeader.query.get_or_404(budget_id)
    item = BudgetItem.query.get_or_404(item_id)
    
    if request.method == 'POST':
        try:
            # 获取表单数据
            category = request.form.get('category', '').strip()
            item_name = request.form.get('item_name', '').strip()
            item_details = request.form.get('item_details', '').strip()
            pricing_method = request.form.get('pricing_method', 'person_based')
            
            # 验证必填字段
            if not category:
                flash('类别不能为空', 'error')
                return redirect(url_for('package_budget.edit_item', budget_id=budget_id, item_id=item_id))
            
            if not item_name:
                flash('项目名称不能为空', 'error')
                return redirect(url_for('package_budget.edit_item', budget_id=budget_id, item_id=item_id))
            
            # 更新基本信息
            item.category = category
            item.item_name = item_name
            item.item_details = item_details
            item.pricing_method = pricing_method
            
            # 根据计价方式处理价格数据
            if pricing_method == 'item_based':
                # 模板1：物品计价方式
                item_unit_price = request.form.get('item_unit_price')
                item_quantity = request.form.get('item_quantity', 1)
                
                if item_unit_price and item_quantity:
                    item.item_unit_price = float(item_unit_price)
                    item.item_quantity = int(item_quantity)
                    # 清空人均价格字段
                    item.adult_price = None
                    item.child_price = None
                else:
                    flash('物品计价方式需要填写物品单价和件数', 'error')
                    return redirect(url_for('package_budget.edit_item', budget_id=budget_id, item_id=item_id))
                    
            else:
                # 模板2：人均计价方式
                adult_price = request.form.get('adult_price')
                child_price = request.form.get('child_price')
                
                if adult_price:
                    item.adult_price = float(adult_price)
                else:
                    item.adult_price = None
                    
                if child_price:
                    item.child_price = float(child_price)
                else:
                    item.child_price = None
                    
                if not adult_price and not child_price:
                    flash('人均计价方式需要至少填写成人单价或儿童单价', 'error')
                    return redirect(url_for('package_budget.edit_item', budget_id=budget_id, item_id=item_id))
                
                # 清空物品价格字段
                item.item_unit_price = None
                item.item_quantity = 1
            
            # 处理人数设置
            item.count_adult_apply = 'count_adult_apply' in request.form
            item.count_child_apply = 'count_child_apply' in request.form
            item.is_optional = 'is_optional' in request.form
            
            # 处理可选的人数覆盖
            adult_count_override = request.form.get('adult_count_override')
            child_count_override = request.form.get('child_count_override')
            
            if adult_count_override:
                item.adult_count_override = int(adult_count_override)
            else:
                item.adult_count_override = None
                
            if child_count_override:
                item.child_count_override = int(child_count_override)
            else:
                item.child_count_override = None
            
            # 处理备注
            remarks = request.form.get('remarks', '').strip()
            item.remarks = remarks if remarks else None
            
            db.session.commit()
            return redirect(url_for('package_budget.detail', budget_id=budget_id))
            
        except ValueError as e:
            flash(f'数据格式错误: {str(e)}', 'error')
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败: {str(e)}', 'error')
    
    return render_template('business/tour/package/TourBudget/edit_item.html', budget=header, item=item)


@package_budget.route('/<int:budget_id>/item/<int:item_id>/delete', methods=['POST'])
@login_required
@staff_only
@csrf.exempt
def delete_item(budget_id, item_id):
    """删除预算项目"""
    try:
        budget = BudgetHeader.query.get_or_404(budget_id)
        item = BudgetItem.query.get_or_404(item_id)
        
        if item.header_id != budget_id:
            return jsonify({'success': False, 'error': '项目不属于此预算单'}), 400
        
        # 检查CSRF token（如果使用JSON请求）
        if request.is_json:
            # 从JSON数据中获取CSRF token
            data = request.get_json()
            csrf_token = data.get('csrf_token')
            if not csrf_token:
                return jsonify({'success': False, 'error': 'CSRF token缺失'}), 400
        else:
            # 从表单数据中获取CSRF token
            csrf_token = request.form.get('csrf_token')
            if not csrf_token:
                return jsonify({'success': False, 'error': 'CSRF token缺失'}), 400
        
        db.session.delete(item)
        db.session.commit()
        
        flash('项目删除成功！', 'success')
        return jsonify({'success': True})
    
    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"Database error in delete item: {e}")
        return jsonify({'success': False, 'error': '数据库错误'}), 500
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in delete item: {e}")
        return jsonify({'success': False, 'error': '删除失败'}), 500


@package_budget.route('/<int:budget_id>/item/<int:item_id>/copy', methods=['POST'])
@login_required
@staff_only
@csrf.exempt
def copy_item(budget_id, item_id):
    """复制预算明细项目"""
    try:
        budget = BudgetHeader.query.get_or_404(budget_id)
        item = BudgetItem.query.get_or_404(item_id)

        if item.header_id != budget_id:
            return jsonify({'success': False, 'error': '项目不属于此预算单'}), 400

        # 获取最大排序号
        max_order = db.session.query(db.func.max(BudgetItem.sort_order)).filter_by(header_id=budget_id).scalar() or 0

        # 复制所有字段
        new_item = BudgetItem(
            header_id=budget_id,
            category=item.category,
            item_name=item.item_name + '（副本）',
            item_details=item.item_details,
            pricing_method=item.pricing_method,
            item_unit_price=item.item_unit_price,
            item_quantity=item.item_quantity,
            adult_price=item.adult_price,
            child_price=item.child_price,
            count_adult_apply=item.count_adult_apply,
            count_child_apply=item.count_child_apply,
            adult_count_override=item.adult_count_override,
            child_count_override=item.child_count_override,
            total_override=item.total_override,
            sort_order=max_order + 1,
            tax_rate=item.tax_rate,
            tax_amount=item.tax_amount,
            is_optional=item.is_optional,
            remarks=item.remarks,
        )

        db.session.add(new_item)
        db.session.commit()

        flash('项目复制成功！', 'success')
        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in copy item: {e}")
        return jsonify({'success': False, 'error': '复制失败'}), 500


@package_budget.route('/<int:budget_id>/duplicate', methods=['POST'])
@login_required
@staff_only
@csrf.exempt
def duplicate(budget_id):
    """复制预算单"""
    try:
        original_budget = BudgetHeader.query.get_or_404(budget_id)
        
        # 创建新的预算单
        # 获取当前用户名字
        if current_user.is_authenticated and current_user.profile:
            created_by_name = f"{current_user.profile.first_name or ''} {current_user.profile.last_name or ''}".strip()
            if not created_by_name:
                created_by_name = current_user.email
        elif current_user.is_authenticated:
            created_by_name = current_user.email
        else:
            created_by_name = 'admin'
        
        new_budget = BudgetHeader(
            package_name=f"{original_budget.package_name} - 副本",
            adult_count=original_budget.adult_count,
            child_count=original_budget.child_count,
            currency=original_budget.currency,
            status='draft',
            is_template=False,
            remarks=original_budget.remarks,
            created_by=created_by_name,
            created_at=datetime.utcnow(),
            project_id=original_budget.project_id  # 继承项目关联
        )
        
        db.session.add(new_budget)
        db.session.flush()  # 获取新预算单的ID
        
        # 复制所有项目
        for original_item in original_budget.items:
            new_item = BudgetItem(
                header_id=new_budget.id,
                category=original_item.category,
                item_name=original_item.item_name,
                adult_price=original_item.adult_price,
                child_price=original_item.child_price,
                total_override=original_item.total_override,
                count_adult_apply=original_item.count_adult_apply,
                count_child_apply=original_item.count_child_apply,
                adult_count_override=original_item.adult_count_override,
                child_count_override=original_item.child_count_override,
                sort_order=original_item.sort_order,
                tax_rate=original_item.tax_rate,
                tax_amount=original_item.tax_amount,
                is_optional=original_item.is_optional,
                remarks=original_item.remarks
            )
            db.session.add(new_item)
        
        db.session.commit()
        
        flash('预算单复制成功！', 'success')
        return redirect(url_for('package_budget.detail', budget_id=new_budget.id))
    
    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"Database error in duplicate budget: {e}")
        flash('复制预算单时发生数据库错误', 'error')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in duplicate budget: {e}")
        flash('复制预算单时发生错误', 'error')
    
    return redirect(url_for('package_budget.detail', budget_id=budget_id))


@package_budget.route('/<int:budget_id>/export', methods=['GET'])
@login_required
@staff_only
def export_budget(budget_id):
    """导出预算单为JSON格式"""
    try:
        budget = BudgetHeader.query.get_or_404(budget_id)
        
        # 构建导出数据
        export_data = {
            'budget': {
                'id': budget.id,
                'package_name': budget.package_name,
                'adult_count': budget.adult_count,
                'child_count': budget.child_count,
                'currency': budget.currency,
                'status': budget.status,
                'is_template': budget.is_template,
                'remarks': budget.remarks,
                'created_at': budget.created_at.isoformat() if budget.created_at else None,
                'updated_at': budget.updated_at.isoformat() if budget.updated_at else None,
                'created_by': budget.created_by,
                'total_price': float(budget.total_price)
            },
            'items': []
        }
        
        for item in budget.items:
            export_data['items'].append({
                'id': item.id,
                'category': item.category,
                'item_name': item.item_name,
                'adult_price': float(item.adult_price) if item.adult_price else None,
                'child_price': float(item.child_price) if item.child_price else None,
                'total_override': float(item.total_override) if item.total_override else None,
                'count_adult_apply': item.count_adult_apply,
                'count_child_apply': item.count_child_apply,
                'adult_count_override': item.adult_count_override,
                'child_count_override': item.child_count_override,
                'sort_order': item.sort_order,
                'tax_rate': float(item.tax_rate) if item.tax_rate else None,
                'tax_amount': float(item.tax_amount) if item.tax_amount else None,
                'is_optional': item.is_optional,
                'remarks': item.remarks,
                'subtotal': float(item.subtotal)
            })
        
        return jsonify(export_data)
    
    except Exception as e:
        current_app.logger.error(f"Error in export budget: {e}")
        return jsonify({'error': '导出失败'}), 500


@package_budget.route('/import', methods=['GET', 'POST'])
@login_required
@staff_only
@csrf.exempt
def import_budget():
    """导入预算单"""
    if request.method == 'POST':
        try:
            # 这里可以实现从JSON文件导入预算单的功能
            # 暂时返回提示信息
            flash('导入功能正在开发中', 'info')
            return redirect(url_for('package_budget.list_budgets'))
        
        except Exception as e:
            current_app.logger.error(f"Error in import budget: {e}")
            flash('导入预算单时发生错误', 'error')
    
    return render_template('business/tour/package/TourBudget/import.html')


@package_budget.route('/templates')
@login_required
@staff_only
def list_templates():
    """模板列表页面"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        search = request.args.get('search', '').strip()
        
        # 查询模板
        query = BudgetHeader.query.filter(BudgetHeader.is_template == True)
        
        # 搜索过滤
        if search:
            query = query.filter(
                or_(
                    BudgetHeader.package_name.ilike(f'%{search}%'),
                    BudgetHeader.remarks.ilike(f'%{search}%')
                )
            )
        
        # 排序
        query = query.order_by(BudgetHeader.created_at.desc())
        
        # 分页
        pagination = query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        templates = pagination.items
        
        return render_template('business/tour/package/TourBudget/templates.html', 
                             templates=templates, 
                             pagination=pagination)
    
    except Exception as e:
        current_app.logger.error(f"Error in list templates: {e}")
        flash('获取模板列表时发生错误', 'error')
        return render_template('business/tour/package/TourBudget/templates.html', templates=[], pagination=None)


@package_budget.route('/<int:budget_id>/download_txt', methods=['GET'])
@login_required
@staff_only
def download_budget_txt(budget_id):
    """下载预算项目为txt文件（顾客版本）"""
    try:
        budget = BudgetHeader.query.get_or_404(budget_id)

        # 下载货币：优先用 URL 参数选择的货币（人民币/新币），否则用预算单默认货币
        # 注意：仅切换货币标签，不做汇率换算
        currency = (request.args.get('currency') or budget.currency or 'SGD').strip()

        # 检查预算单是否存在项目
        if not budget.items:
            flash('预算单中没有项目，无法生成下载文件', 'warning')
            return redirect(url_for('package_budget.detail', budget_id=budget_id))
        
        # 生成txt内容
        content = []
        
        # 标题部分
        content.append("=" * 60)
        content.append(f"旅游配套详情")
        content.append("=" * 60)
        content.append("")
        
        # 基本信息
        content.append("【配套信息】")
        content.append(f"配套名称：{budget.package_name}")
        content.append(f"成人人数：{budget.adult_count}人")
        content.append(f"儿童人数：{budget.child_count}人")
        content.append("")
        
        # 项目明细
        content.append("【包含项目】")
        content.append("-" * 40)
        content.append("")
        
        total_price = 0
        price_lines = []  # 每项详细价格，单独成段放到文件后面

        for i, item in enumerate(budget.items, 1):
            # 项目标题
            content.append(f"{i:2d}. {item.item_name}")

            # 项目详情
            if item.item_details:
                details_lines = item.item_details.split('\n')
                for line in details_lines:
                    if line.strip():
                        content.append(f"    {line.strip()}")

            # 可选项目标记
            if item.is_optional:
                content.append("    [可选项目]")

            # 备注（如果有重要信息）
            if item.remarks:
                content.append(f"    备注：{item.remarks}")

            content.append("")

            # 计算项目总价
            item_total = item.subtotal or 0
            total_price += item_total

            # 组装该项价格明细（单价构成 + 小计），放到后面的【费用明细】段
            adult_cnt = item.adult_count_override if item.adult_count_override is not None else budget.adult_count
            child_cnt = item.child_count_override if item.child_count_override is not None else budget.child_count
            if item.pricing_method == 'item_based':
                breakdown = f"单价 {float(item.item_unit_price or 0):.2f} × {item.item_quantity or 1}件"
            else:
                parts = []
                if item.count_adult_apply and (item.adult_price or 0):
                    parts.append(f"成人 {float(item.adult_price or 0):.2f} × {adult_cnt}人")
                if item.count_child_apply and (item.child_price or 0):
                    parts.append(f"儿童 {float(item.child_price or 0):.2f} × {child_cnt}人")
                breakdown = " + ".join(parts) if parts else "-"
            optional_tag = "（可选）" if item.is_optional else ""
            price_lines.append(f"{i:2d}. {item.item_name}{optional_tag}")
            price_lines.append(f"    {breakdown}")
            price_lines.append(f"    小计：{item_total:.2f} {currency}")
            price_lines.append("")

        # 费用明细（每项详细价格，放在后面）
        content.append("=" * 60)
        content.append("【费用明细】")
        content.append("-" * 40)
        content.append("")
        content.extend(price_lines)

        # 总价
        content.append("=" * 60)
        content.append("【总价】")
        content.append("-" * 40)
        content.append(f"总价：{total_price:.2f} {currency}")
        content.append("=" * 60)
        content.append("")
        
        # 备注信息
        if budget.remarks:
            content.append("【备注】")
            content.append("-" * 40)
            content.append(budget.remarks)
            content.append("")
        
        # 生成安全的文件名，避免编码问题
        safe_package_name = "".join(c for c in budget.package_name if c.isalnum() or c in (' ', '-', '_')).strip()
        if not safe_package_name:
            safe_package_name = "package"
        filename = f"{safe_package_name}_{currency}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        # 使用RFC 5987标准编码文件名，支持中文
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        
        # 使用UTF-8编码生成响应
        content_str = '\n'.join(content)
        response = Response(content_str.encode('utf-8'), mimetype='text/plain; charset=utf-8')
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        return response
    except Exception as e:
        current_app.logger.error(f"Error in download_budget_txt for budget {budget_id}: {e}")
        flash(f'下载失败: {str(e)}', 'error')
        return redirect(url_for('package_budget.detail', budget_id=budget_id))


@package_budget.route('/<int:budget_id>/update_currency', methods=['POST'])
@login_required
@staff_only
@csrf.exempt
def update_currency(budget_id):
    """更新预算单货币（仅切换货币标签，不做汇率换算），用于报价页自动保存货币选择"""
    try:
        budget = BudgetHeader.query.get_or_404(budget_id)
        currency = (request.form.get('currency') or '').strip().upper()
        if currency not in ('SGD', 'CNY'):
            return jsonify({'success': False, 'error': '不支持的货币'}), 400
        budget.currency = currency
        db.session.commit()
        return jsonify({'success': True, 'currency': currency})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"update_currency failed for budget {budget_id}: {e}")
        return jsonify({'success': False, 'error': '保存失败'}), 500


# 明细项目 Excel 导入/导出的列定义：(模型字段, 表头, 列宽)
EXCEL_ITEM_COLUMNS = [
    ('id', 'ID', 8),
    ('category', '类别', 14),
    ('item_name', '项目名称', 22),
    ('item_details', '详情', 30),
    ('pricing_method', '计价方式(人均/物品)', 18),
    ('adult_price', '成人单价', 10),
    ('child_price', '儿童单价', 10),
    ('item_unit_price', '物品单价', 10),
    ('item_quantity', '物品件数', 10),
    ('count_adult_apply', '计成人(是/否)', 12),
    ('count_child_apply', '计儿童(是/否)', 12),
    ('adult_count_override', '成人人数(可空)', 14),
    ('child_count_override', '儿童人数(可空)', 14),
    ('is_optional', '可选(是/否)', 12),
    ('remarks', '备注', 20),
]


@package_budget.route('/<int:budget_id>/items/export', methods=['GET'])
@login_required
@staff_only
def export_items(budget_id):
    """导出明细项目为Excel（含ID，可改后重新上传更新；小计为只读列）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    budget = BudgetHeader.query.get_or_404(budget_id)

    wb = Workbook()
    ws = wb.active
    ws.title = '明细项目'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 表头（末尾追加只读“小计”列，导入时忽略）
    headers = [h for (_, h, _) in EXCEL_ITEM_COLUMNS] + [f'小计({budget.currency})']
    widths = [w for (_, _, w) in EXCEL_ITEM_COLUMNS] + [12]
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    def b2s(v):
        return '是' if v else '否'

    for row_idx, item in enumerate(sorted(budget.items, key=lambda x: x.sort_order or 0), 2):
        pricing = '物品' if item.pricing_method == 'item_based' else '人均'
        row = [
            item.id,
            item.category or '',
            item.item_name or '',
            item.item_details or '',
            pricing,
            float(item.adult_price) if item.adult_price is not None else '',
            float(item.child_price) if item.child_price is not None else '',
            float(item.item_unit_price) if item.item_unit_price is not None else '',
            item.item_quantity if item.item_quantity is not None else '',
            b2s(item.count_adult_apply),
            b2s(item.count_child_apply),
            item.adult_count_override if item.adult_count_override is not None else '',
            item.child_count_override if item.child_count_override is not None else '',
            b2s(item.is_optional),
            item.remarks or '',
            round(item.subtotal or 0, 2),
        ]
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c for c in (budget.package_name or '') if c.isalnum() or c in (' ', '-', '_')).strip() or 'budget'
    download_name = f"{safe_name}_明细_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=download_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@package_budget.route('/<int:budget_id>/items/import', methods=['POST'])
@login_required
@staff_only
@csrf.exempt
def import_items(budget_id):
    """从Excel导入明细项目（有ID则更新，无ID则新增）"""
    from openpyxl import load_workbook

    budget = BudgetHeader.query.get_or_404(budget_id)

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'}), 400
    file = request.files['file']
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '请上传 .xlsx 格式的Excel文件'}), 400

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'success': False, 'message': f'无法读取Excel文件: {e}'}), 400

    # 表头 -> 列索引（按表头前缀匹配，忽略括号说明）
    headers = [cell.value for cell in ws[1]]
    col_map = {}
    for field, header, _ in EXCEL_ITEM_COLUMNS:
        prefix = header.split('(')[0]
        for h_idx, h_val in enumerate(headers):
            if h_val and str(h_val).strip().startswith(prefix):
                col_map[field] = h_idx
                break

    if 'item_name' not in col_map:
        return jsonify({'success': False, 'message': '未找到"项目名称"列，请使用导出的Excel作为模板'}), 400

    def cell_val(row, field):
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def to_bool(v, default):
        if v is None or v == '':
            return default
        return str(v).strip().lower() in ('是', 'true', '1', 'yes', 'y', 't')

    def to_float(v):
        if v is None or v == '':
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    def to_int(v):
        if v is None or v == '':
            return None
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return None

    created, updated, skipped = 0, 0, 0
    try:
        max_order = db.session.query(db.func.max(BudgetItem.sort_order)).filter_by(header_id=budget_id).scalar() or 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            name = cell_val(row, 'item_name')
            if name is None or str(name).strip() == '':
                skipped += 1
                continue

            row_id = to_int(cell_val(row, 'id'))
            pricing_raw = str(cell_val(row, 'pricing_method') or '').strip()
            pricing = 'item_based' if pricing_raw in ('物品', 'item_based', 'item') else 'person_based'

            item = None
            if row_id:
                item = BudgetItem.query.filter_by(id=row_id, header_id=budget_id).first()
            if item is None:
                max_order += 1
                item = BudgetItem(header_id=budget_id, sort_order=max_order)
                db.session.add(item)
                created += 1
            else:
                updated += 1

            item.category = (str(cell_val(row, 'category') or '').strip()) or None
            item.item_name = str(name).strip()
            item.item_details = (str(cell_val(row, 'item_details') or '').strip()) or None
            item.pricing_method = pricing
            item.adult_price = to_float(cell_val(row, 'adult_price'))
            item.child_price = to_float(cell_val(row, 'child_price'))
            item.item_unit_price = to_float(cell_val(row, 'item_unit_price'))
            qty = to_int(cell_val(row, 'item_quantity'))
            item.item_quantity = qty if qty is not None else 1
            item.count_adult_apply = to_bool(cell_val(row, 'count_adult_apply'), True)
            item.count_child_apply = to_bool(cell_val(row, 'count_child_apply'), True)
            item.adult_count_override = to_int(cell_val(row, 'adult_count_override'))
            item.child_count_override = to_int(cell_val(row, 'child_count_override'))
            item.is_optional = to_bool(cell_val(row, 'is_optional'), False)
            item.remarks = (str(cell_val(row, 'remarks') or '').strip()) or None

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"import_items failed for budget {budget_id}: {e}")
        return jsonify({'success': False, 'message': f'导入失败: {e}'}), 500

    return jsonify({
        'success': True,
        'message': f'导入完成：新增 {created} 项，更新 {updated} 项，跳过 {skipped} 项',
        'created': created, 'updated': updated, 'skipped': skipped,
    })


@package_budget.route('/project/<int:project_id>')
@login_required
@staff_only
def budgets_by_project(project_id):
    """根据项目ID获取关联的预算单列表"""
    try:
        project = TourProject.query.get_or_404(project_id)
        budgets = BudgetHeader.query.filter_by(project_id=project_id).order_by(BudgetHeader.created_at.desc()).all()
        
        return render_template('business/tour/package/TourBudget/project_budgets.html',
                             project=project,
                             budgets=budgets)
    
    except Exception as e:
        current_app.logger.error(f"Error in budgets_by_project: {e}")
        flash('获取项目预算时发生错误', 'error')
        return redirect(url_for('tour_projects.list_tour_projects'))


@package_budget.route('/project/<int:project_id>/json')
@login_required
@staff_only
def budgets_by_project_json(project_id):
    """根据项目ID获取关联的预算单列表（JSON格式）"""
    try:
        budgets = BudgetHeader.query.filter_by(project_id=project_id).order_by(BudgetHeader.created_at.desc()).all()
        
        result = []
        for budget in budgets:
            result.append({
                'id': budget.id,
                'package_name': budget.package_name,
                'adult_count': budget.adult_count,
                'child_count': budget.child_count,
                'currency': budget.currency,
                'status': budget.status,
                'total_price': float(budget.total_price),
                'created_at': budget.created_at.strftime('%Y-%m-%d %H:%M') if budget.created_at else None,
                'created_by': budget.created_by
            })
        
        return jsonify({'success': True, 'budgets': result, 'count': len(result)})
    
    except Exception as e:
        current_app.logger.error(f"Error in budgets_by_project_json: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500