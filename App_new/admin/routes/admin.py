"""
管理员功能路由
管理员仪表板、用户管理、权限管理、系统配置等功能
"""

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app
from flask_login import login_required, current_user
from App_new.utils.decorators import admin_only, permission_required
from App_new.auth.models.auth import AuthUser, Role, UserProfile, InvitationCode
from App_new.utils.permissions import get_all_permissions, get_all_roles, ROLE_PERMISSIONS
from App_new.exts import db
from datetime import datetime, timedelta
from sqlalchemy import func
import json

# 创建管理员蓝图
admin = Blueprint('admin', __name__, url_prefix='/admin')

@admin.route('/dashboard')
@login_required
@admin_only
def dashboard():
    """管理员仪表板"""
    try:
        # 获取系统统计信息
        stats = {
            'total_users': AuthUser.query.count(),
            'total_roles': Role.query.count(),
            'active_users': AuthUser.query.filter_by(is_active=True).count(),
            'new_users_this_month': get_new_users_count(),
        }
        
        # 获取业务数据统计（使用优化后的方法）
        try:
            from App_new.business.projects.services.project_stats import ProjectStatsService
            project_stats_service = ProjectStatsService()

            # 使用优化后的统计方法
            business_stats = project_stats_service.get_optimized_total_stats()

            # 添加业务统计数据
            stats.update({
                'total_projects': business_stats.get('total_projects', 0),
                'active_projects': business_stats.get('active_projects', 0),
                'completed_projects': business_stats.get('completed_projects', 0),
                'total_revenue': round(business_stats.get('total_revenue', 0), 2),
                'total_cost': round(business_stats.get('total_cost', 0), 2),
                'total_profit': round(business_stats.get('total_profit', 0), 2),
                'total_received': round(business_stats.get('total_received', 0), 2),
                'profit_margin': round(business_stats.get('profit_margin', 0), 1),
                'payment_ratio': round(business_stats.get('payment_ratio', 0), 1),
            })

            # 获取预警统计
            warning_stats = project_stats_service.get_warning_stats()
            stats['warnings'] = warning_stats

            # 获取旅游团统计
            tour_stats = project_stats_service.get_tour_stats()
            stats['tour'] = tour_stats

            # 获取TOP客户
            top_customers = project_stats_service.get_top_customers(5)
            stats['top_customers'] = top_customers

        except Exception as e:
            current_app.logger.warning(f'获取业务统计数据失败: {str(e)}')
            # 如果业务统计失败，设置默认值
            stats.update({
                'total_projects': 0,
                'active_projects': 0,
                'completed_projects': 0,
                'total_revenue': 0,
                'total_cost': 0,
                'total_profit': 0,
                'total_received': 0,
                'warnings': {},
                'tour': {},
                'top_customers': [],
            })
        
        # 获取系统健康度
        try:
            stats['system_health'] = get_system_health()
        except Exception as e:
            current_app.logger.warning(f'获取系统健康度失败: {str(e)}')
            stats['system_health'] = {'status': 'unknown', 'percentage': 0}
        
        # 获取最近注册的用户
        recent_users = AuthUser.query.order_by(AuthUser.created_at.desc()).limit(5).all()
        
        # 获取系统活动
        recent_activities = get_recent_activities()
        
        # 获取用户增长数据
        user_growth_data = get_user_growth_data()
        
        # 获取角色分布数据
        role_distribution = get_role_distribution()
        
        return render_template('admin/dashboard.html', 
                             stats=stats,
                             recent_users=recent_users,
                             recent_activities=recent_activities,
                             user_growth_data=user_growth_data,
                             role_distribution=role_distribution,
                             now=datetime.now())
    except Exception as e:
        # 记录详细错误日志
        current_app.logger.error(f'加载管理员仪表板失败: {str(e)}', exc_info=True)
        flash(f'加载仪表板失败：{str(e)}', 'error')
        return render_template('admin/dashboard.html',
                             stats={}, recent_users=[], recent_activities=[],
                             user_growth_data={'labels': [], 'data': []},
                             role_distribution={},
                             now=datetime.now())

@admin.route('/users')
@login_required
@admin_only
def users():
    """用户管理"""
    try:
        # 获取所有用户
        all_users = AuthUser.query.order_by(AuthUser.created_at.desc()).all()
        
        # 统计各角色用户数量
        admin_count = AuthUser.query.join(Role).filter(Role.name == 'admin').count()
        staff_count = AuthUser.query.join(Role).filter(Role.name == 'staff').count()
        member_count = AuthUser.query.join(Role).filter(Role.name == 'member').count()
        total_count = AuthUser.query.count()
        
        return render_template('admin/users.html',
                             users=all_users,
                             admin_count=admin_count,
                             staff_count=staff_count,
                             member_count=member_count,
                             total_count=total_count)
    except Exception as e:
        flash(f'加载用户列表失败：{str(e)}', 'error')
        return render_template('admin/users.html',
                             users=[],
                             admin_count=0,
                             staff_count=0,
                             member_count=0,
                             total_count=0)

@admin.route('/user/<int:user_id>')
@login_required
@admin_only
def user_detail(user_id):
    """用户详情"""
    try:
        user = AuthUser.query.get_or_404(user_id)
        
        # 获取用户统计信息（模拟数据）
        user_stats = {
            'login_count': 0,
            'last_login': None,
            'projects_count': 0,
            'orders_count': 0
        }
        
        return render_template('admin/user_detail.html',
                             user=user,
                             user_stats=user_stats)
    except Exception as e:
        flash(f'加载用户详情失败：{str(e)}', 'error')
        return redirect(url_for('admin.users'))

@admin.route('/user/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_only
def edit_user(user_id):
    """编辑用户"""
    try:
        user = AuthUser.query.get_or_404(user_id)
        
        if request.method == 'POST':
            # 获取表单数据
            role_id = request.form.get('role_id', type=int)
            is_active = request.form.get('is_active') == 'on'
            
            # 更新用户信息
            if role_id:
                role = Role.query.get(role_id)
                if role:
                    user.role_id = role_id
                    user.role = role
            
            # 如果用户模型有is_active字段，更新它
            if hasattr(user, 'is_active'):
                user.is_active = is_active
            
            # 更新用户资料
            if user.profile:
                user.profile.first_name = request.form.get('first_name', '').strip()
                user.profile.last_name = request.form.get('last_name', '').strip()
                user.profile.phone = request.form.get('phone', '').strip()
                user.profile.address = request.form.get('address', '').strip()
                
                # 更新员工等级（仅对员工角色有效）
                if user.role and user.role.name == 'staff':
                    staff_level = request.form.get('staff_level', type=int)
                    if staff_level in [1, 2]:
                        user.profile.staff_level = staff_level
            
            db.session.commit()
            flash('用户信息更新成功', 'success')
            return redirect(url_for('admin.user_detail', user_id=user_id))
        
        # 获取所有角色
        roles = Role.query.all()
        
        return render_template('admin/edit_user.html',
                             user=user,
                             roles=roles)
    except Exception as e:
        db.session.rollback()
        flash(f'编辑用户失败：{str(e)}', 'error')
        return redirect(url_for('admin.users'))

@admin.route('/staff-levels')
@login_required
@admin_only
def staff_levels():
    """员工等级管理页面"""
    try:
        # 获取所有员工用户
        staff_users = AuthUser.query.join(Role).filter(Role.name == 'staff').all()
        
        # 统计各级别员工数量
        total_count = len(staff_users)
        level_1_count = sum(1 for user in staff_users if user.profile and (user.profile.staff_level or 1) == 1)
        level_2_count = sum(1 for user in staff_users if user.profile and (user.profile.staff_level or 1) == 2)
        
        return render_template('admin/staff_levels.html',
                             staff_users=staff_users,
                             total_count=total_count,
                             level_1_count=level_1_count,
                             level_2_count=level_2_count)
    except Exception as e:
        flash(f'加载员工等级管理页面失败：{str(e)}', 'error')
        return render_template('admin/staff_levels.html',
                             staff_users=[], total_count=0, level_1_count=0, level_2_count=0)

@admin.route('/update-staff-level', methods=['POST'])
@login_required
@admin_only
def update_staff_level():
    """更新员工等级"""
    try:
        # 获取JSON数据
        data = request.get_json()
        print(f"🔍 接收到的数据: {data}")
        
        if not data:
            return jsonify({'success': False, 'message': '无效的请求数据'}), 400
        
        user_id = data.get('user_id')
        new_level = data.get('level')
        print(f"🔍 解析的参数: user_id={user_id}, new_level={new_level}")
        
        # 转换为整数
        try:
            if user_id:
                user_id = int(user_id)
            if new_level:
                new_level = int(new_level)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '参数类型错误'}), 400
        
        if not user_id or new_level not in [1, 2]:
            return jsonify({'success': False, 'message': '无效的参数'}), 400
        
        user = AuthUser.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': '用户不存在'}), 404
        
        if not user.role or user.role.name != 'staff':
            return jsonify({'success': False, 'message': '只能修改员工用户的等级'}), 400
        
        # 确保用户有profile
        if not user.profile:
            user.profile = UserProfile(user_id=user.id)
            db.session.add(user.profile)
        
        user.profile.staff_level = new_level
        db.session.commit()
        
        print(f"✅ 员工等级更新成功: 用户{user_id} -> {new_level}级")
        return jsonify({'success': True, 'message': f'员工等级已更新为{new_level}级'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'更新失败：{str(e)}'}), 500

@admin.route('/roles')
@login_required
@admin_only
def roles():
    """角色管理"""
    try:
        roles = Role.query.all()
        
        # 为每个角色获取用户数量
        role_stats = {}
        for role in roles:
            role_stats[role.id] = AuthUser.query.filter_by(role_id=role.id).count()
        
        return render_template('admin/roles.html',
                             roles=roles,
                             role_stats=role_stats,
                             role_permissions=ROLE_PERMISSIONS)
    except Exception as e:
        flash(f'加载角色列表失败：{str(e)}', 'error')
        return render_template('admin/roles.html',
                             roles=[], role_stats={}, role_permissions={})

@admin.route('/role/<int:role_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_only
def edit_role(role_id):
    """编辑角色"""
    try:
        role = Role.query.get_or_404(role_id)
        
        if request.method == 'POST':
            # 获取表单数据
            role.description = request.form.get('description', '').strip()
            
            # 获取权限列表
            permissions = request.form.getlist('permissions')
            role.permissions = permissions
            
            db.session.commit()
            flash('角色权限更新成功', 'success')
            return redirect(url_for('admin.roles'))
        
        # 获取所有可用权限
        all_permissions = get_all_permissions()
        current_permissions = role.permissions or []
        
        return render_template('admin/edit_role.html',
                             role=role,
                             all_permissions=all_permissions,
                             current_permissions=current_permissions)
    except Exception as e:
        db.session.rollback()
        flash(f'编辑角色失败：{str(e)}', 'error')
        return redirect(url_for('admin.roles'))

@admin.route('/system')
@login_required
@admin_only
def system():
    """系统配置"""
    try:
        # 获取系统信息
        system_info = {
            'app_name': 'MyTravelPanel',
            'version': '1.0.0',
            'database_users': AuthUser.query.count(),
            'database_roles': Role.query.count(),
            'python_version': get_python_version(),
            'flask_version': get_flask_version(),
        }
        
        return render_template('admin/system.html',
                             system_info=system_info)
    except Exception as e:
        flash(f'加载系统配置失败：{str(e)}', 'error')
        return render_template('admin/system.html',
                             system_info={})

@admin.route('/analytics')
@login_required
@admin_only
def analytics():
    """数据分析"""
    try:
        # 获取分析数据
        analytics_data = {
            'user_growth': get_user_growth_data(),
            'role_distribution': get_role_distribution(),
            'activity_stats': get_activity_stats(),
        }
        
        return render_template('admin/analytics.html',
                             analytics_data=analytics_data)
    except Exception as e:
        flash(f'加载数据分析失败：{str(e)}', 'error')
        return render_template('admin/analytics.html',
                             analytics_data={})

@admin.route('/create-user', methods=['GET', 'POST'])
@login_required
@admin_only
def create_user():
    """管理员创建用户"""
    if request.method == 'POST':
        try:
            # 获取表单数据
            email = request.form.get('email', '').strip()
            password = request.form.get('password', '')
            role_name = request.form.get('role', '')
            first_name = request.form.get('first_name', '').strip()
            last_name = request.form.get('last_name', '').strip()
            phone = request.form.get('phone', '').strip()
            
            # 基础验证
            if not all([email, password, role_name, first_name]):
                flash('请填写所有必填字段', 'error')
                return render_template('admin/create_user.html', roles=get_all_roles())
            
            # 只允许创建员工和管理员（会员通过公开注册）
            if role_name not in ['staff', 'admin']:
                flash('只能创建员工或管理员账户，会员请通过公开注册', 'error')
                return render_template('admin/create_user.html', roles=get_all_roles())
            
            # 邮箱格式验证
            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, email):
                flash('请输入有效的邮箱地址', 'error')
                return render_template('admin/create_user.html', roles=get_all_roles())
            
            # 密码长度验证
            if len(password) < 6:
                flash('密码长度至少6位', 'error')
                return render_template('admin/create_user.html', roles=get_all_roles())
            
            # 检查邮箱是否已存在
            existing_user = AuthUser.query.filter_by(email=email).first()
            if existing_user:
                flash('该邮箱已被注册', 'error')
                return render_template('admin/create_user.html', roles=get_all_roles())
            
            # 获取角色
            role = Role.query.filter_by(name=role_name).first()
            if not role:
                flash(f'角色 {role_name} 不存在', 'error')
                return render_template('admin/create_user.html', roles=get_all_roles())
            
            # 创建新用户
            username = email.split('@')[0]
            base_username = username
            counter = 1
            while AuthUser.query.filter_by(username=username).first():
                username = f"{base_username}{counter}"
                counter += 1
            
            new_user = AuthUser(
                username=username,
                email=email,
                role_id=role.id
            )
            new_user.set_password(password)
            
            # 保存用户到数据库
            db.session.add(new_user)
            db.session.commit()
            
            # 创建用户资料
            user_profile = UserProfile(
                user_id=new_user.id,
                first_name=first_name,
                last_name=last_name,
                phone=phone
            )
            db.session.add(user_profile)
            db.session.commit()
            
            flash(f'成功创建{role.description}账户：{email}', 'success')
            return redirect(url_for('admin.users'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'创建用户失败：{str(e)}', 'error')
            return render_template('admin/create_user.html', roles=get_all_roles())
    
    # GET请求，显示创建用户表单
    return render_template('admin/create_user.html', roles=get_all_roles())

@admin.route('/invitation-codes')
@login_required
@admin_only
def invitation_codes():
    """邀请码管理页面"""
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    codes = InvitationCode.query.order_by(InvitationCode.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False)
    
    return render_template('admin/invitation_codes.html', codes=codes, now=datetime.utcnow())

@admin.route('/invitation-codes/create', methods=['GET', 'POST'])
@login_required
@admin_only
def create_invitation_code():
    """创建邀请码"""
    if request.method == 'POST':
        try:
            role_name = request.form.get('role_name', '').strip()
            expires_days = request.form.get('expires_days', type=int)
            count = request.form.get('count', 1, type=int)
            
            if role_name not in ['staff', 'admin']:
                flash('角色类型无效', 'error')
                return render_template('admin/create_invitation_code.html')
            
            if count < 1 or count > 10:
                flash('一次最多生成10个邀请码', 'error')
                return render_template('admin/create_invitation_code.html')
            
            # 计算过期时间
            expires_at = None
            if expires_days and expires_days > 0:
                expires_at = datetime.utcnow() + timedelta(days=expires_days)
            
            # 批量创建邀请码
            created_codes = []
            for _ in range(count):
                code = InvitationCode(
                    code=InvitationCode.generate_code(),
                    role_name=role_name,
                    created_by=current_user.id,
                    expires_at=expires_at
                )
                db.session.add(code)
                created_codes.append(code.code)
            
            db.session.commit()
            
            flash(f'成功创建 {count} 个{_get_role_display_name(role_name)}邀请码', 'success')
            
            # 如果只创建了一个，显示邀请码
            if count == 1:
                flash(f'邀请码：{created_codes[0]}', 'info')
            
            return redirect(url_for('admin.invitation_codes'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'创建邀请码失败：{str(e)}', 'error')
    
    return render_template('admin/create_invitation_code.html')

@admin.route('/invitation-codes/<int:code_id>/revoke', methods=['POST'])
@login_required
@admin_only
def revoke_invitation_code(code_id):
    """撤销邀请码"""
    try:
        code = InvitationCode.query.get_or_404(code_id)
        
        if code.is_used:
            flash('邀请码已被使用，无法撤销', 'error')
        else:
            code.is_used = True
            code.used_at = datetime.utcnow()
            db.session.commit()
            flash('邀请码已撤销', 'success')
            
    except Exception as e:
        db.session.rollback()
        flash(f'撤销邀请码失败：{str(e)}', 'error')
    
    return redirect(url_for('admin.invitation_codes'))

# 工具函数
def get_new_users_count():
    """获取本月新用户数量"""
    try:
        first_day = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return AuthUser.query.filter(AuthUser.created_at >= first_day).count()
    except (AttributeError, TypeError) as e:
        # 如果 created_at 字段不存在或查询失败，记录错误并返回0
        current_app.logger.warning(f'获取本月新用户数量失败: {str(e)}')
        return 0
    except Exception as e:
        # 其他数据库相关错误
        current_app.logger.error(f'获取本月新用户数量时发生错误: {str(e)}', exc_info=True)
        return 0

def get_recent_activities():
    """获取最近系统活动（从现有数据中提取）"""
    activities = []
    try:
        # 获取最近注册的用户（作为用户注册活动）
        recent_registered_users = AuthUser.query.order_by(
            AuthUser.created_at.desc()
        ).limit(5).all()
        
        for user in recent_registered_users:
            if user.created_at:
                role_name = user.role.name if user.role else '未知'
                activities.append({
                    'type': 'user_register',
                    'title': f'新用户注册',
                    'description': f'{user.email} 注册为{role_name}',
                    'time': user.created_at.strftime('%Y-%m-%d %H:%M'),
                    'icon': 'fa-user-plus',
                    'color': 'success'
                })
        
        # 获取最近登录的用户（如果有last_login字段）
        if hasattr(AuthUser, 'last_login'):
            recent_logins = AuthUser.query.filter(
                AuthUser.last_login.isnot(None)
            ).order_by(
                AuthUser.last_login.desc()
            ).limit(3).all()
            
            for user in recent_logins:
                if user.last_login:
                    activities.append({
                        'type': 'user_login',
                        'title': f'用户登录',
                        'description': f'{user.email} 登录系统',
                        'time': user.last_login.strftime('%Y-%m-%d %H:%M'),
                        'icon': 'fa-sign-in-alt',
                        'color': 'info'
                    })
        
        # 按时间排序，最新的在前
        activities.sort(key=lambda x: x['time'], reverse=True)
        
        # 只返回最近10条活动
        return activities[:10]
        
    except Exception as e:
        current_app.logger.error(f'获取系统活动失败: {str(e)}', exc_info=True)
        # 如果获取失败，返回一条默认活动
        return [{
            'type': 'system',
            'title': '系统运行中',
            'description': '系统正常运行',
            'time': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'icon': 'fa-check-circle',
            'color': 'success'
        }]

def get_python_version():
    """获取Python版本"""
    import sys
    return f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"

def get_flask_version():
    """获取Flask版本"""
    try:
        import flask
        return flask.__version__
    except ImportError:
        # Flask 未安装
        return "未知"
    except AttributeError:
        # Flask 版本信息不可用
        return "未知"
    except Exception as e:
        current_app.logger.warning(f'获取Flask版本失败: {str(e)}')
        return "未知"

def get_user_growth_data():
    """获取用户增长数据（真实数据）"""
    try:
        from sqlalchemy import func, extract
        
        # 获取最近6个月的数据
        six_months_ago = datetime.now() - timedelta(days=180)
        
        # 按月统计用户注册数
        monthly_stats = db.session.query(
            extract('year', AuthUser.created_at).label('year'),
            extract('month', AuthUser.created_at).label('month'),
            func.count(AuthUser.id).label('count')
        ).filter(
            AuthUser.created_at >= six_months_ago
        ).group_by(
            extract('year', AuthUser.created_at),
            extract('month', AuthUser.created_at)
        ).order_by(
            extract('year', AuthUser.created_at),
            extract('month', AuthUser.created_at)
        ).all()
        
        labels = []
        data = []
        
        for stat in monthly_stats:
            labels.append(f"{int(stat.year)}年{int(stat.month)}月")
            data.append(stat.count)
        
        # 如果数据不足6个月，补充前面的月份
        if len(labels) < 6:
            current_date = datetime.now()
            for i in range(6 - len(labels)):
                month_date = current_date - timedelta(days=30 * (6 - len(labels) - i))
                labels.insert(0, f"{month_date.year}年{month_date.month}月")
                data.insert(0, 0)
        
        return {
            'labels': labels[-6:],  # 只返回最近6个月
            'data': data[-6:]
        }
    except Exception as e:
        current_app.logger.error(f"获取用户增长数据失败: {str(e)}", exc_info=True)
        # 返回空数据而不是模拟数据
        return {'labels': [], 'data': []}

def get_role_distribution():
    """获取角色分布数据"""
    try:
        roles = Role.query.all()
        distribution = {}
        for role in roles:
            count = AuthUser.query.filter_by(role_id=role.id).count()
            distribution[role.name] = count
        return distribution
    except Exception as e:
        current_app.logger.error(f'获取角色分布数据失败: {str(e)}', exc_info=True)
        return {}

def get_activity_stats():
    """获取活动统计"""
    try:
        # 获取最近24小时内的登录次数（如果有last_login字段）
        from datetime import timedelta
        yesterday = datetime.now() - timedelta(days=1)
        recent_logins = AuthUser.query.filter(
            AuthUser.last_login >= yesterday
        ).count() if hasattr(AuthUser, 'last_login') else 0
        
        return {
            'total_logins': recent_logins,
            'active_sessions': 1,  # 当前管理员
            'system_uptime': '运行中'
        }
    except Exception as e:
        current_app.logger.warning(f'获取活动统计失败: {str(e)}')
        return {
            'total_logins': 0,
            'active_sessions': 1,
            'system_uptime': '运行中'
        }

def get_system_health():
    """获取系统健康度"""
    try:
        health_score = 100
        issues = []
        
        # 检查数据库连接
        try:
            db.session.execute(db.text('SELECT 1'))
            db.session.commit()
        except Exception as e:
            health_score -= 30
            issues.append('数据库连接异常')
            current_app.logger.error(f'数据库连接检查失败: {str(e)}')
        
        # 检查系统资源（如果psutil可用）
        try:
            import psutil
            cpu_percent = psutil.cpu_percent(interval=0.1)
            memory = psutil.virtual_memory()
            disk = psutil.disk_usage('/' if hasattr(psutil, 'disk_usage') else 'C:')
            
            # CPU使用率检查
            if cpu_percent > 90:
                health_score -= 10
                issues.append('CPU使用率过高')
            elif cpu_percent > 70:
                health_score -= 5
                issues.append('CPU使用率较高')
            
            # 内存使用率检查
            if memory.percent > 90:
                health_score -= 10
                issues.append('内存使用率过高')
            elif memory.percent > 70:
                health_score -= 5
                issues.append('内存使用率较高')
            
            # 磁盘使用率检查
            if disk.percent > 90:
                health_score -= 10
                issues.append('磁盘空间不足')
            elif disk.percent > 70:
                health_score -= 5
                issues.append('磁盘空间使用率较高')
        except ImportError:
            # psutil未安装，跳过资源检查
            pass
        except Exception as e:
            current_app.logger.warning(f'系统资源检查失败: {str(e)}')
        
        # 确定健康状态
        if health_score >= 90:
            status = 'excellent'
        elif health_score >= 70:
            status = 'good'
        elif health_score >= 50:
            status = 'warning'
        else:
            status = 'critical'
        
        return {
            'status': status,
            'percentage': max(0, min(100, health_score)),
            'issues': issues
        }
    except Exception as e:
        current_app.logger.error(f'获取系统健康度失败: {str(e)}', exc_info=True)
        return {
            'status': 'unknown',
            'percentage': 0,
            'issues': ['无法获取系统健康度']
        }

def get_all_roles():
    """获取所有角色（仅用于管理员创建用户）"""
    return Role.query.filter(Role.name.in_(['staff', 'admin'])).all()

def _get_role_display_name(role_name):
    """获取角色显示名称"""
    role_names = {
        'member': '会员',
        'staff': '员工',
        'admin': '管理员'
    }
    return role_names.get(role_name, role_name)

# API 路由
@admin.route('/api/stats')
@login_required
@admin_only
def api_stats():
    """获取管理员统计数据"""
    try:
        stats = {
            'total_users': AuthUser.query.count(),
            'total_roles': Role.query.count(),
            'active_users': AuthUser.query.count(),  # 简化版本
            'new_users_today': 0,
            'system_health': 'excellent'
        }
        
        return jsonify({
            'success': True,
            'data': stats
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@admin.route('/api/user/<int:user_id>/toggle-status', methods=['POST'])
@login_required
@admin_only
def api_toggle_user_status(user_id):
    """切换用户状态"""
    try:
        user = AuthUser.query.get_or_404(user_id)
        
        # 如果用户模型有is_active字段
        if hasattr(user, 'is_active'):
            user.is_active = not user.is_active
            db.session.commit()
            
            status = '激活' if user.is_active else '禁用'
            return jsonify({
                'success': True,
                'message': f'用户已{status}',
                'is_active': user.is_active
            })
        else:
            return jsonify({
                'success': False,
                'message': '用户状态切换功能未实现'
            }), 400
            
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@admin.route('/api/user/<int:user_id>/delete', methods=['DELETE'])
@login_required
@admin_only
def api_delete_user(user_id):
    """删除用户"""
    try:
        user = AuthUser.query.get_or_404(user_id)
        
        # 不能删除自己
        if user.id == current_user.id:
            return jsonify({
                'success': False,
                'message': '不能删除自己的账户'
            }), 400
        
        # 不能删除管理员（除非当前用户是超级管理员）
        if user.role and user.role.name == 'admin':
            return jsonify({
                'success': False,
                'message': '不能删除管理员账户'
            }), 400
        
        # 删除用户资料（如果存在）
        if user.profile:
            db.session.delete(user.profile)
        
        # 删除用户
        db.session.delete(user)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '用户已删除'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@admin.route('/api/system/info')
@login_required
@admin_only
def api_system_info():
    """获取系统信息"""
    try:
        import psutil
        import platform
        
        system_info = {
            'platform': platform.system(),
            'python_version': get_python_version(),
            'flask_version': get_flask_version(),
            'cpu_percent': psutil.cpu_percent(),
            'memory_percent': psutil.virtual_memory().percent,
            'disk_usage': psutil.disk_usage('/').percent if platform.system() != 'Windows' else psutil.disk_usage('C:').percent
        }
        
        return jsonify({
            'success': True,
            'data': system_info
        })
    except ImportError:
        # 如果psutil未安装，返回基础信息
        return jsonify({
            'success': True,
            'data': {
                'python_version': get_python_version(),
                'flask_version': get_flask_version(),
                'status': '运行中'
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500 

@admin.route('/users/locked')
@admin_only
def locked_users():
    """查看被锁定的用户列表"""
    try:
        from App_new.auth.models import AuthUser
        
        # 获取所有被锁定的用户
        locked_users = AuthUser.query.filter_by(is_locked=True).all()
        
        return render_template('admin/locked_users.html', users=locked_users)
    except Exception as e:
        flash(f'获取锁定用户列表失败：{str(e)}', 'error')
        return redirect(url_for('admin.dashboard'))

@admin.route('/users/<int:user_id>/unlock', methods=['POST'])
@admin_only
def unlock_user(user_id):
    """解锁被锁定的用户"""
    try:
        from App_new.auth.models import AuthUser
        
        user = AuthUser.query.get_or_404(user_id)
        
        if not user.is_locked:
            flash('该用户账户未被锁定', 'info')
        else:
            user.unlock_account()
            flash(f'用户 {user.username} 的账户已成功解锁', 'success')
        
        return redirect(url_for('admin.locked_users'))
    except Exception as e:
        flash(f'解锁用户失败：{str(e)}', 'error')
        return redirect(url_for('admin.locked_users'))

@admin.route('/users/<int:user_id>/reset-attempts', methods=['POST'])
@admin_only
def reset_user_attempts(user_id):
    """重置用户的登录失败次数"""
    try:
        from App_new.auth.models import AuthUser
        
        user = AuthUser.query.get_or_404(user_id)
        
        user.login_attempts = 0
        db.session.commit()
        
        flash(f'用户 {user.username} 的登录失败次数已重置', 'success')
        return redirect(url_for('admin.locked_users'))
    except Exception as e:
        flash(f'重置失败次数失败：{str(e)}', 'error')
        return redirect(url_for('admin.locked_users'))

@admin.route('/change-password', methods=['GET', 'POST'])
@login_required
@admin_only
def change_password():
    """管理员更改密码"""
    if request.method == 'POST':
        try:
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            # 验证表单数据
            if not all([current_password, new_password, confirm_password]):
                flash('请填写所有字段', 'error')
                return render_template('admin/change_password.html')
            
            if new_password != confirm_password:
                flash('新密码和确认密码不匹配', 'error')
                return render_template('admin/change_password.html')
            
            if len(new_password) < 6:
                flash('新密码长度至少为6位', 'error')
                return render_template('admin/change_password.html')
            
            # 验证当前密码
            # 检查数据库中存储的密码（支持明文密码）
            if current_password == current_user.password_hash:
                # 更新密码（临时使用明文存储，避免哈希问题）
                old_password = current_user.password_hash
                current_user.password_hash = new_password
                
                try:
                    db.session.commit()
                    
                    # 验证密码是否真的被更新了
                    db.session.refresh(current_user)
                    if current_user.password_hash == new_password:
                        current_app.logger.info(f"管理员密码修改成功: {current_user.email}")
                        flash('密码修改成功！', 'success')
                        return redirect(url_for('admin.dashboard'))
                    else:
                        current_app.logger.error(f"密码更新验证失败: {current_user.email}")
                        flash('密码修改失败：更新验证失败', 'error')
                        return render_template('admin/change_password.html')
                except Exception as db_error:
                    db.session.rollback()
                    current_app.logger.error(f"数据库提交失败: {db_error}")
                    flash('密码修改失败：数据库更新错误', 'error')
                    return render_template('admin/change_password.html')
            else:
                flash('当前密码错误', 'error')
                return render_template('admin/change_password.html')
                
        except Exception as e:
            flash(f'密码修改失败：{str(e)}', 'error')
            return render_template('admin/change_password.html')
    
    return render_template('admin/change_password.html')

@admin.route('/center')
@login_required
@admin_only
def admin_center():
    """管理员中心"""
    return render_template('admin/admin_center.html')


@admin.route('/export/stats')
@login_required
@admin_only
def export_stats():
    """导出统计报表"""
    try:
        import io
        from flask import send_file

        # 检查是否安装了 openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            flash('请先安装 openpyxl: pip install openpyxl', 'error')
            return redirect(url_for('admin.dashboard'))

        from App_new.business.projects.services.project_stats import ProjectStatsService
        project_stats_service = ProjectStatsService()

        # 获取统计数据
        business_stats = project_stats_service.get_optimized_total_stats()
        warning_stats = project_stats_service.get_warning_stats()
        tour_stats = project_stats_service.get_tour_stats()
        top_customers = project_stats_service.get_top_customers(10)

        # 创建工作簿
        wb = Workbook()

        # 样式定义
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: 业务统计概览
        ws1 = wb.active
        ws1.title = '业务统计概览'

        # 添加标题
        ws1['A1'] = '业务统计报表'
        ws1['A1'].font = Font(bold=True, size=16)
        ws1['A2'] = f'导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

        # 项目统计
        ws1['A4'] = '项目统计'
        ws1['A4'].font = Font(bold=True, size=12)
        data_rows = [
            ['指标', '数值'],
            ['总项目数', business_stats.get('total_projects', 0)],
            ['活跃项目', business_stats.get('active_projects', 0)],
            ['已完成项目', business_stats.get('completed_projects', 0)],
            ['草稿项目', business_stats.get('draft_projects', 0)],
        ]
        for i, row in enumerate(data_rows, start=5):
            for j, val in enumerate(row, start=1):
                cell = ws1.cell(row=i, column=j, value=val)
                cell.border = border
                if i == 5:
                    cell.font = header_font
                    cell.fill = header_fill

        # 财务统计
        ws1['A11'] = '财务统计'
        ws1['A11'].font = Font(bold=True, size=12)
        finance_rows = [
            ['指标', '金额 (S$)'],
            ['总收入', round(business_stats.get('total_revenue', 0), 2)],
            ['总成本', round(business_stats.get('total_cost', 0), 2)],
            ['总利润', round(business_stats.get('total_profit', 0), 2)],
            ['已收款', round(business_stats.get('total_received', 0), 2)],
            ['利润率', f"{round(business_stats.get('profit_margin', 0), 1)}%"],
            ['回款率', f"{round(business_stats.get('payment_ratio', 0), 1)}%"],
        ]
        for i, row in enumerate(finance_rows, start=12):
            for j, val in enumerate(row, start=1):
                cell = ws1.cell(row=i, column=j, value=val)
                cell.border = border
                if i == 12:
                    cell.font = header_font
                    cell.fill = header_fill

        # 调整列宽
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 20

        # Sheet 2: 预警统计
        ws2 = wb.create_sheet('预警统计')
        ws2['A1'] = '预警统计'
        ws2['A1'].font = Font(bold=True, size=14)

        warning_rows = [
            ['预警类型', '数量', '说明'],
            ['超期项目', warning_stats.get('overdue_projects', 0), '活跃超过30天未完成'],
            ['逾期应收款', warning_stats.get('overdue_receivables', 0), f"金额: S${warning_stats.get('overdue_amount', 0)}"],
            ['草稿项目', warning_stats.get('draft_projects', 0), '待跟进处理'],
            ['低利润项目', warning_stats.get('low_profit_projects', 0), '利润率低于10%'],
        ]
        for i, row in enumerate(warning_rows, start=3):
            for j, val in enumerate(row, start=1):
                cell = ws2.cell(row=i, column=j, value=val)
                cell.border = border
                if i == 3:
                    cell.font = header_font
                    cell.fill = header_fill

        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 10
        ws2.column_dimensions['C'].width = 25

        # Sheet 3: 旅游团统计
        ws3 = wb.create_sheet('旅游团统计')
        ws3['A1'] = '旅游团运营统计'
        ws3['A1'].font = Font(bold=True, size=14)

        tour_rows = [
            ['指标', '数值'],
            ['总团数', tour_stats.get('total_groups', 0)],
            ['本月出发', tour_stats.get('departing_this_month', 0)],
            ['近7天出发', tour_stats.get('upcoming_7days', 0)],
            ['进行中', tour_stats.get('ongoing_tours', 0)],
            ['待出发', tour_stats.get('pending_departure', 0)],
            ['已完成', tour_stats.get('completed_tours', 0)],
            ['总人数', tour_stats.get('total_pax', 0)],
            ['平均人数/团', tour_stats.get('avg_pax', 0)],
        ]
        for i, row in enumerate(tour_rows, start=3):
            for j, val in enumerate(row, start=1):
                cell = ws3.cell(row=i, column=j, value=val)
                cell.border = border
                if i == 3:
                    cell.font = header_font
                    cell.fill = header_fill

        ws3.column_dimensions['A'].width = 15
        ws3.column_dimensions['B'].width = 15

        # Sheet 4: TOP客户
        ws4 = wb.create_sheet('TOP客户')
        ws4['A1'] = 'TOP客户排名'
        ws4['A1'].font = Font(bold=True, size=14)

        customer_rows = [['排名', '客户名称', '项目数', '总收入 (S$)']]
        for i, customer in enumerate(top_customers, start=1):
            customer_rows.append([
                i,
                customer.get('name', ''),
                customer.get('project_count', 0),
                round(customer.get('total_revenue', 0), 2)
            ])

        for i, row in enumerate(customer_rows, start=3):
            for j, val in enumerate(row, start=1):
                cell = ws4.cell(row=i, column=j, value=val)
                cell.border = border
                if i == 3:
                    cell.font = header_font
                    cell.fill = header_fill

        ws4.column_dimensions['A'].width = 8
        ws4.column_dimensions['B'].width = 30
        ws4.column_dimensions['C'].width = 10
        ws4.column_dimensions['D'].width = 15

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 返回文件
        filename = f'统计报表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.error(f'导出报表失败: {str(e)}', exc_info=True)
        flash(f'导出报表失败：{str(e)}', 'error')
        return redirect(url_for('admin.dashboard'))

